from datetime import date

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q, Sum, F
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView, FormView

from apps.core.barcode import generate_qr_image
from apps.core.mixins import ManagerRequiredMixin
from .models import (
    AssetAudit, AssetAuditItem, AssetCategory, AssetTransfer,
    Certification, DepreciationRecord, FixedAsset, LeaseContract, Location,
)
from .forms import (
    AssetAuditForm, AssetAuditItemForm, AssetCategoryForm, AssetDisposalForm,
    AssetRegisterFilterForm, AssetTransferForm, CertificationForm,
    DepreciationRunForm, FixedAssetForm, LeaseContractForm, LocationForm,
)


# === 고정자산 ===

class AssetListView(ManagerRequiredMixin, ListView):
    model = FixedAsset
    template_name = 'asset/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('category', 'department', 'responsible_person')

        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(name__icontains=q)

        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(category_id=category)

        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['categories'] = AssetCategory.objects.filter(is_active=True)
        ctx['status_choices'] = FixedAsset.Status.choices
        return ctx


class AssetCreateView(ManagerRequiredMixin, CreateView):
    model = FixedAsset
    form_class = FixedAssetForm
    template_name = 'asset/asset_form.html'
    success_url = reverse_lazy('asset:asset_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '고정자산이 등록되었습니다.')
        return super().form_valid(form)


class AssetDetailView(ManagerRequiredMixin, DetailView):
    model = FixedAsset
    template_name = 'asset/asset_detail.html'
    context_object_name = 'asset'
    slug_field = 'asset_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'category', 'department', 'responsible_person',
            'source_receipt_item__goods_receipt',
            'disposal_approval',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['depreciation_records'] = self.object.depreciation_records.all()[:24]
        ctx['transfers'] = self.object.transfers.select_related(
            'from_department', 'to_department', 'from_person', 'to_person',
        ).filter(is_active=True)[:20]
        ctx['lease_contracts'] = self.object.lease_contracts.select_related(
            'lessor',
        ).filter(is_active=True)[:10]
        return ctx


class AssetUpdateView(ManagerRequiredMixin, UpdateView):
    model = FixedAsset
    form_class = FixedAssetForm
    template_name = 'asset/asset_form.html'
    slug_field = 'asset_number'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('asset:asset_list')

    def form_valid(self, form):
        messages.success(self.request, '고정자산이 수정되었습니다.')
        return super().form_valid(form)


class AssetDisposalView(ManagerRequiredMixin, UpdateView):
    model = FixedAsset
    form_class = AssetDisposalForm
    template_name = 'asset/asset_disposal_form.html'
    slug_field = 'asset_number'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('asset:asset_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = '자산 처분'
        return ctx

    def form_valid(self, form):
        asset = self.object
        disposal_date = form.cleaned_data.get('disposal_date')

        if disposal_date and asset.status == FixedAsset.Status.ACTIVE:
            with transaction.atomic():
                # Catch-up depreciation: 처분월까지 빠진 월에 대한 감가상각 실행
                last_record = asset.depreciation_records.order_by('-year', '-month').first()
                if last_record:
                    start_year, start_month = last_record.year, last_record.month
                    # 다음 월부터
                    if start_month == 12:
                        start_year += 1
                        start_month = 1
                    else:
                        start_month += 1
                else:
                    start_year = asset.acquisition_date.year
                    start_month = asset.acquisition_date.month

                end_year = disposal_date.year
                end_month = disposal_date.month

                y, m = start_year, start_month
                while (y, m) <= (end_year, end_month):
                    asset.refresh_from_db()
                    if asset.is_fully_depreciated:
                        break
                    if DepreciationRecord.objects.filter(asset=asset, year=y, month=m).exists():
                        if m == 12:
                            y, m = y + 1, 1
                        else:
                            m += 1
                        continue

                    dep_amount = asset.monthly_depreciation
                    if dep_amount <= 0:
                        break
                    max_depreciable = asset.book_value - asset.residual_value
                    if dep_amount > max_depreciable:
                        dep_amount = int(max_depreciable)
                    if dep_amount <= 0:
                        break

                    new_accumulated = asset.accumulated_depreciation + dep_amount
                    new_book_value = asset.acquisition_cost - new_accumulated

                    DepreciationRecord.objects.create(
                        asset=asset,
                        year=y,
                        month=m,
                        depreciation_amount=dep_amount,
                        accumulated_amount=new_accumulated,
                        book_value_after=new_book_value,
                        created_by=self.request.user,
                    )
                    if m == 12:
                        y, m = y + 1, 1
                    else:
                        m += 1

                asset.refresh_from_db()

        messages.success(self.request, '자산 처분이 처리되었습니다.')
        return super().form_valid(form)


class AssetDepreciationRunView(ManagerRequiredMixin, FormView):
    form_class = DepreciationRunForm
    template_name = 'asset/depreciation_run.html'
    success_url = reverse_lazy('asset:asset_list')

    def form_valid(self, form):
        year = form.cleaned_data['year']
        month = form.cleaned_data['month']

        assets = FixedAsset.objects.filter(
            is_active=True,
            status=FixedAsset.Status.ACTIVE,
        )

        created_count = 0
        skipped_count = 0

        with transaction.atomic():
            for asset in assets:
                # 이미 해당 기간 감가상각 내역이 있으면 건너뜀
                if DepreciationRecord.objects.filter(asset=asset, year=year, month=month).exists():
                    skipped_count += 1
                    continue

                # 완전 상각된 자산은 건너뜀
                if asset.is_fully_depreciated:
                    skipped_count += 1
                    continue

                dep_amount = asset.monthly_depreciation
                if dep_amount <= 0:
                    skipped_count += 1
                    continue

                # 잔존가치 이하로 내려가지 않도록 조정
                max_depreciable = asset.book_value - asset.residual_value
                if dep_amount > max_depreciable:
                    dep_amount = int(max_depreciable)

                if dep_amount <= 0:
                    skipped_count += 1
                    continue

                new_accumulated = asset.accumulated_depreciation + dep_amount
                new_book_value = asset.acquisition_cost - new_accumulated

                DepreciationRecord.objects.create(
                    asset=asset,
                    year=year,
                    month=month,
                    depreciation_amount=dep_amount,
                    accumulated_amount=new_accumulated,
                    book_value_after=new_book_value,
                    created_by=self.request.user,
                )

                # F() 표현식으로 원자적 업데이트
                FixedAsset.all_objects.filter(pk=asset.pk).update(
                    accumulated_depreciation=F('accumulated_depreciation') + dep_amount,
                    book_value=F('book_value') - dep_amount,
                )

                created_count += 1

        messages.success(
            self.request,
            f'{year}년 {month}월 감가상각 완료: {created_count}건 처리, {skipped_count}건 건너뜀',
        )
        return super().form_valid(form)


# === 자산 이관 ===

class AssetTransferListView(ManagerRequiredMixin, ListView):
    model = AssetTransfer
    template_name = 'asset/transfer_list.html'
    context_object_name = 'transfers'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related(
            'asset', 'from_department', 'to_department',
            'from_person', 'to_person',
        )

        asset = self.request.GET.get('asset')
        if asset:
            qs = qs.filter(
                Q(asset__asset_number__icontains=asset) | Q(asset__name__icontains=asset)
            )

        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(transfer_date__gte=date_from)

        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(transfer_date__lte=date_to)

        department = self.request.GET.get('department')
        if department:
            qs = qs.filter(
                Q(from_department_id=department) | Q(to_department_id=department)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.hr.models import Department
        ctx['departments'] = Department.objects.filter(is_active=True)
        return ctx


class AssetTransferCreateView(ManagerRequiredMixin, CreateView):
    model = AssetTransfer
    form_class = AssetTransferForm
    template_name = 'asset/asset_transfer_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.asset = get_object_or_404(
            FixedAsset, asset_number=kwargs['slug'], status=FixedAsset.Status.ACTIVE,
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['asset'] = self.asset
        return ctx

    def get_initial(self):
        initial = super().get_initial()
        initial['transfer_date'] = date.today().isoformat()
        return initial

    def form_valid(self, form):
        form.instance.asset = self.asset
        form.instance.from_department = self.asset.department
        form.instance.from_person = self.asset.responsible_person
        form.instance.from_location = self.asset.location or ''
        form.instance.from_managed_location = self.asset.managed_location
        form.instance.created_by = self.request.user
        messages.success(self.request, '자산 이관이 처리되었습니다.')
        response = super().form_valid(form)
        return response

    def get_success_url(self):
        return reverse('asset:asset_detail', kwargs={'slug': self.asset.asset_number})


# === 자산대장 리포트 ===

class AssetRegisterReportView(ManagerRequiredMixin, TemplateView):
    template_name = 'asset/register_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        form = AssetRegisterFilterForm(self.request.GET or None)
        ctx['form'] = form

        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            category = form.cleaned_data.get('category')
            department = form.cleaned_data.get('department')

            base_qs = FixedAsset.objects.filter(is_active=True)
            if category:
                base_qs = base_qs.filter(category=category)
            if department:
                base_qs = base_qs.filter(department=department)

            # 기초잔액: 기간 시작일 전 취득, 처분일이 없거나 기간 내/후
            opening_qs = base_qs.filter(
                acquisition_date__lt=start_date,
            ).filter(
                Q(disposal_date__isnull=True) | Q(disposal_date__gte=start_date)
            )
            opening_balance = opening_qs.aggregate(s=Sum('book_value'))['s'] or 0

            # 당기취득
            acquired_qs = base_qs.filter(
                acquisition_date__gte=start_date, acquisition_date__lte=end_date,
            )
            acquisitions = acquired_qs.aggregate(s=Sum('acquisition_cost'))['s'] or 0

            # 당기처분
            disposed_qs = base_qs.filter(
                disposal_date__gte=start_date, disposal_date__lte=end_date,
            )
            disposals = disposed_qs.aggregate(s=Sum('book_value'))['s'] or 0

            # 당기감가상각
            dep_qs = DepreciationRecord.objects.filter(
                asset__in=base_qs, is_active=True,
            )
            # year/month를 날짜 범위로 필터
            dep_qs = dep_qs.filter(
                Q(year__gt=start_date.year) | Q(year=start_date.year, month__gte=start_date.month),
            ).filter(
                Q(year__lt=end_date.year) | Q(year=end_date.year, month__lte=end_date.month),
            )
            depreciation = dep_qs.aggregate(s=Sum('depreciation_amount'))['s'] or 0

            # 기말잔액
            closing_balance = opening_balance + acquisitions - disposals - depreciation

            ctx.update({
                'has_data': True,
                'start_date': start_date,
                'end_date': end_date,
                'opening_balance': opening_balance,
                'acquisitions': acquisitions,
                'disposals': disposals,
                'depreciation': depreciation,
                'closing_balance': closing_balance,
                'acquired_assets': acquired_qs.select_related('category')[:50],
                'disposed_assets': disposed_qs.select_related('category')[:50],
            })
        return ctx


class AssetRegisterExcelView(ManagerRequiredMixin, TemplateView):
    template_name = 'asset/register_report.html'

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment

        form = AssetRegisterFilterForm(request.GET or None)
        if not form.is_valid():
            messages.error(request, '날짜를 올바르게 입력해주세요.')
            return HttpResponse(status=400)

        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        category = form.cleaned_data.get('category')
        department = form.cleaned_data.get('department')

        base_qs = FixedAsset.objects.filter(is_active=True)
        if category:
            base_qs = base_qs.filter(category=category)
        if department:
            base_qs = base_qs.filter(department=department)

        assets = base_qs.filter(
            Q(acquisition_date__lte=end_date),
            Q(disposal_date__isnull=True) | Q(disposal_date__gte=start_date),
        ).select_related('category', 'department').order_by('asset_number')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '자산대장'

        headers = ['자산번호', '자산명', '분류', '사용부서', '취득일', '취득원가', '감가상각누계', '장부가액', '상태']
        bold = Font(bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row, asset in enumerate(assets, 2):
            ws.cell(row=row, column=1, value=asset.asset_number)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=str(asset.category) if asset.category else '')
            ws.cell(row=row, column=4, value=str(asset.department) if asset.department else '')
            ws.cell(row=row, column=5, value=asset.acquisition_date.isoformat() if asset.acquisition_date else '')
            ws.cell(row=row, column=6, value=int(asset.acquisition_cost))
            ws.cell(row=row, column=7, value=int(asset.accumulated_depreciation))
            ws.cell(row=row, column=8, value=int(asset.book_value))
            ws.cell(row=row, column=9, value=asset.get_status_display())

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=asset_register_{start_date}_{end_date}.xlsx'
        wb.save(response)
        return response


# === 위치 관리 ===

class LocationListView(ManagerRequiredMixin, ListView):
    model = Location
    template_name = 'asset/location_list.html'
    context_object_name = 'locations'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related('parent')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(building__icontains=q))
        return qs


class LocationCreateView(ManagerRequiredMixin, CreateView):
    model = Location
    form_class = LocationForm
    template_name = 'asset/location_form.html'
    success_url = reverse_lazy('asset:location_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '자산 위치가 등록되었습니다.')
        return super().form_valid(form)


class LocationUpdateView(ManagerRequiredMixin, UpdateView):
    model = Location
    form_class = LocationForm
    template_name = 'asset/location_form.html'
    success_url = reverse_lazy('asset:location_list')

    def form_valid(self, form):
        messages.success(self.request, '자산 위치가 수정되었습니다.')
        return super().form_valid(form)


# === 자산 분류 ===

class AssetCategoryListView(ManagerRequiredMixin, ListView):
    model = AssetCategory
    template_name = 'asset/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class AssetCategoryCreateView(ManagerRequiredMixin, CreateView):
    model = AssetCategory
    form_class = AssetCategoryForm
    template_name = 'asset/category_form.html'
    success_url = reverse_lazy('asset:category_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '자산 분류가 등록되었습니다.')
        return super().form_valid(form)


# === 자산 현황 요약 ===

class AssetSummaryView(ManagerRequiredMixin, TemplateView):
    template_name = 'asset/summary.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # 분류별 현황
        category_summary = AssetCategory.objects.filter(
            is_active=True,
        ).prefetch_related('assets').all()

        summary_data = []
        total_acquisition = 0
        total_accumulated = 0
        total_book_value = 0

        for cat in category_summary:
            active_assets = cat.assets.filter(is_active=True, status=FixedAsset.Status.ACTIVE)
            agg = active_assets.aggregate(
                acq=Sum('acquisition_cost'),
                acc=Sum('accumulated_depreciation'),
                bv=Sum('book_value'),
            )
            acq = agg['acq'] or 0
            acc = agg['acc'] or 0
            bv = agg['bv'] or 0
            count = active_assets.count()

            if count > 0:
                summary_data.append({
                    'category': cat,
                    'count': count,
                    'acquisition_cost': acq,
                    'accumulated_depreciation': acc,
                    'book_value': bv,
                })
                total_acquisition += acq
                total_accumulated += acc
                total_book_value += bv

        ctx['summary_data'] = summary_data
        ctx['total_acquisition'] = total_acquisition
        ctx['total_accumulated'] = total_accumulated
        ctx['total_book_value'] = total_book_value

        # 상태별 자산 수
        ctx['active_count'] = FixedAsset.objects.filter(is_active=True, status=FixedAsset.Status.ACTIVE).count()
        ctx['disposed_count'] = FixedAsset.objects.filter(is_active=True, status=FixedAsset.Status.DISPOSED).count()
        ctx['scrapped_count'] = FixedAsset.objects.filter(is_active=True, status=FixedAsset.Status.SCRAPPED).count()

        return ctx


class AssetDepartmentSummaryView(ManagerRequiredMixin, TemplateView):
    """부서별 자산 현황 + 차트 대시보드"""
    template_name = 'asset/department_summary.html'

    def get_context_data(self, **kwargs):
        import json
        from django.db.models import Count
        ctx = super().get_context_data(**kwargs)
        from apps.hr.models import Department

        active_assets = FixedAsset.objects.filter(is_active=True, status=FixedAsset.Status.ACTIVE)

        # 1) 부서별 집계
        dept_data = []
        departments = Department.objects.filter(is_active=True).order_by('name')
        for dept in departments:
            agg = active_assets.filter(department=dept).aggregate(
                count=Count('id'),
                acq=Sum('acquisition_cost'),
                bv=Sum('book_value'),
            )
            if agg['count'] > 0:
                dept_data.append({
                    'department': dept,
                    'count': agg['count'],
                    'acquisition_cost': agg['acq'] or 0,
                    'book_value': agg['bv'] or 0,
                })

        # 미배정 자산
        unassigned = active_assets.filter(department__isnull=True).aggregate(
            count=Count('id'),
            acq=Sum('acquisition_cost'),
            bv=Sum('book_value'),
        )
        if unassigned['count'] > 0:
            dept_data.append({
                'department': None,
                'count': unassigned['count'],
                'acquisition_cost': unassigned['acq'] or 0,
                'book_value': unassigned['bv'] or 0,
            })

        ctx['dept_data'] = dept_data

        # 2) 카테고리별 집계 (파이차트)
        cat_agg = active_assets.values('category__name').annotate(
            count=Count('id'),
            bv=Sum('book_value'),
        ).order_by('-bv')
        ctx['category_labels_json'] = json.dumps([c['category__name'] or '미분류' for c in cat_agg])
        ctx['category_values_json'] = json.dumps([int(c['bv'] or 0) for c in cat_agg])
        ctx['category_counts_json'] = json.dumps([c['count'] for c in cat_agg])

        # 3) 상태별 집계 (도넛차트)
        all_assets = FixedAsset.objects.filter(is_active=True)
        status_agg = all_assets.values('status').annotate(count=Count('id')).order_by('status')
        status_labels = []
        status_values = []
        status_map = dict(FixedAsset.Status.choices)
        for s in status_agg:
            status_labels.append(status_map.get(s['status'], s['status']))
            status_values.append(s['count'])
        ctx['status_labels_json'] = json.dumps(status_labels)
        ctx['status_values_json'] = json.dumps(status_values)

        # 4) 부서별 장부가 (바차트용)
        dept_labels = [d['department'].name if d['department'] else '미배정' for d in dept_data]
        dept_book_values = [int(d['book_value']) for d in dept_data]
        dept_counts = [d['count'] for d in dept_data]
        ctx['dept_labels_json'] = json.dumps(dept_labels)
        ctx['dept_bv_json'] = json.dumps(dept_book_values)
        ctx['dept_counts_json'] = json.dumps(dept_counts)

        # 5) 위치별 자산 수
        location_agg = active_assets.filter(
            managed_location__isnull=False,
        ).values('managed_location__name').annotate(
            count=Count('id'),
        ).order_by('-count')[:10]
        ctx['location_labels_json'] = json.dumps([loc['managed_location__name'] for loc in location_agg])
        ctx['location_counts_json'] = json.dumps([loc['count'] for loc in location_agg])

        # 총계
        totals = active_assets.aggregate(
            total_count=Count('id'),
            total_acq=Sum('acquisition_cost'),
            total_bv=Sum('book_value'),
            total_dep=Sum('accumulated_depreciation'),
        )
        ctx['total_count'] = totals['total_count'] or 0
        ctx['total_acq'] = totals['total_acq'] or 0
        ctx['total_bv'] = totals['total_bv'] or 0
        ctx['total_dep'] = totals['total_dep'] or 0

        return ctx


# === 인증 관리 (G1) ===

class CertificationListView(ManagerRequiredMixin, ListView):
    model = Certification
    template_name = 'asset/certification_list.html'
    context_object_name = 'certifications'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related('product', 'asset')
        cert_type = self.request.GET.get('cert_type')
        if cert_type:
            qs = qs.filter(cert_type=cert_type)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cert_types'] = Certification.CertType.choices
        return ctx


class CertificationCreateView(ManagerRequiredMixin, CreateView):
    model = Certification
    form_class = CertificationForm
    template_name = 'asset/certification_form.html'
    success_url = reverse_lazy('asset:certification_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '인증이 등록되었습니다.')
        return super().form_valid(form)


class CertificationDetailView(ManagerRequiredMixin, DetailView):
    model = Certification
    template_name = 'asset/certification_detail.html'
    context_object_name = 'cert'

    def get_queryset(self):
        return super().get_queryset().select_related('product', 'asset')


# === 리스 계약 (G2) ===

class LeaseContractListView(ManagerRequiredMixin, ListView):
    model = LeaseContract
    template_name = 'asset/lease_list.html'
    context_object_name = 'contracts'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related('asset', 'lessor')
        lease_type = self.request.GET.get('lease_type')
        if lease_type:
            qs = qs.filter(lease_type=lease_type)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['lease_types'] = LeaseContract.LeaseType.choices
        return ctx


class LeaseContractCreateView(ManagerRequiredMixin, CreateView):
    model = LeaseContract
    form_class = LeaseContractForm
    template_name = 'asset/lease_form.html'
    success_url = reverse_lazy('asset:lease_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '리스 계약이 등록되었습니다.')
        return super().form_valid(form)


class LeaseContractDetailView(ManagerRequiredMixin, DetailView):
    model = LeaseContract
    template_name = 'asset/lease_detail.html'
    context_object_name = 'contract'

    def get_queryset(self):
        return super().get_queryset().select_related('asset', 'lessor')


# === 자산 실사 (G3) ===

class AssetAuditListView(ManagerRequiredMixin, ListView):
    model = AssetAudit
    template_name = 'asset/audit_list.html'
    context_object_name = 'audits'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).select_related('department', 'auditor')


class AssetAuditCreateView(ManagerRequiredMixin, CreateView):
    model = AssetAudit
    form_class = AssetAuditForm
    template_name = 'asset/audit_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        # 대상부서의 ACTIVE 자산으로 실사항목 자동 생성
        audit = self.object
        asset_qs = FixedAsset.objects.filter(is_active=True, status=FixedAsset.Status.ACTIVE)
        if audit.department:
            asset_qs = asset_qs.filter(department=audit.department)

        items = [
            AssetAuditItem(
                audit=audit,
                asset=asset,
                created_by=self.request.user,
            )
            for asset in asset_qs
        ]
        AssetAuditItem.objects.bulk_create(items)

        messages.success(self.request, f'자산실사가 등록되었습니다. ({len(items)}건 항목 생성)')
        return response

    def get_success_url(self):
        return reverse('asset:audit_detail', kwargs={'pk': self.object.pk})


class AssetAuditDetailView(ManagerRequiredMixin, DetailView):
    model = AssetAudit
    template_name = 'asset/audit_detail.html'
    context_object_name = 'audit'

    def get_queryset(self):
        return super().get_queryset().select_related('department', 'auditor')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = self.object.items.select_related('asset__category', 'asset__department').all()
        ctx['status_choices'] = AssetAuditItem.AuditStatus.choices
        ctx['condition_choices'] = AssetAuditItem.Condition.choices
        total = self.object.items.count()
        confirmed = self.object.items.exclude(status=AssetAuditItem.AuditStatus.FOUND).count()
        ctx['progress'] = int(confirmed / total * 100) if total > 0 else 0
        ctx['total_count'] = total
        ctx['confirmed_count'] = confirmed
        return ctx


class AssetAuditExecuteView(ManagerRequiredMixin, DetailView):
    """실사 실행 — 항목별 결과 일괄 입력"""
    model = AssetAudit
    template_name = 'asset/audit_execute.html'
    context_object_name = 'audit'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = self.object.items.select_related('asset__category', 'asset__department').all()
        ctx['status_choices'] = AssetAuditItem.AuditStatus.choices
        ctx['condition_choices'] = AssetAuditItem.Condition.choices
        return ctx

    def post(self, request, *args, **kwargs):
        audit = self.get_object()
        updated_count = 0

        with transaction.atomic():
            for item in audit.items.all():
                status_val = request.POST.get(f'status_{item.pk}')
                location = request.POST.get(f'location_{item.pk}', '')
                condition = request.POST.get(f'condition_{item.pk}', '')
                remark = request.POST.get(f'remark_{item.pk}', '')

                if status_val and status_val != item.status:
                    item.status = status_val
                    item.actual_location = location
                    item.condition = condition or AssetAuditItem.Condition.GOOD
                    item.remark = remark
                    item.save(update_fields=[
                        'status', 'actual_location', 'condition',
                        'remark', 'updated_at',
                    ])
                    updated_count += 1

        messages.success(request, f'{updated_count}건 실사결과가 업데이트되었습니다.')
        return self.get(request, *args, **kwargs)


# === 자산 QR 코드 ===

class AssetQRView(LoginRequiredMixin, DetailView):
    """자산 QR 이미지 PNG 응답"""
    model = FixedAsset
    slug_field = 'asset_number'
    slug_url_kwarg = 'slug'

    def get(self, request, *args, **kwargs):
        import base64
        asset = self.get_object()
        qr_data = f'{asset.asset_number}|{asset.name}'
        qr_b64 = generate_qr_image(qr_data)
        image_bytes = base64.b64decode(qr_b64)
        return HttpResponse(image_bytes, content_type='image/png')


class AssetQRPrintView(LoginRequiredMixin, TemplateView):
    """QR 라벨 인쇄 페이지 (3열 격자)"""
    template_name = 'asset/qr_print.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ids_param = self.request.GET.get('ids', '')
        if ids_param:
            id_list = [i.strip() for i in ids_param.split(',') if i.strip().isdigit()]
            assets = FixedAsset.objects.filter(pk__in=id_list, is_active=True)
        else:
            assets = FixedAsset.objects.none()

        labels = []
        for asset in assets:
            qr_data = f'{asset.asset_number}|{asset.name}'
            labels.append({
                'asset': asset,
                'qr_image': generate_qr_image(qr_data),
            })
        ctx['labels'] = labels
        return ctx


class AssetQRScanView(LoginRequiredMixin, TemplateView):
    """QR 스캔 페이지"""
    template_name = 'asset/qr_scan.html'
