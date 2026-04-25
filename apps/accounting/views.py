import json
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.views import View

from apps.core.import_views import BaseImportView
from apps.core.mixins import AdminRequiredMixin, ManagerRequiredMixin
from django.db.models import Count, F, Q, Sum
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView


def safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

from apps.inventory.models import Product
from apps.sales.models import Order, OrderItem
from .models import (
    Currency, ExchangeRate,
    TaxRate, TaxInvoice, FixedCost, WithholdingTax, AccountCode, Voucher, VoucherLine,
    AccountReceivable, AccountPayable, Payment, BankAccount,
    AccountTransfer, PaymentDistribution,
    CostSettlement, CostSettlementItem,
    SalesSettlement, SalesSettlementOrder,
    Budget, ClosingPeriod,
    CreditCard, CardTransaction, CardBilling,
    CostCenter,
    CashReceipt, CashReceiptItem,
    PlatformFinancialConfig,
    AdvanceReceived, AdvancePaid,
)
from .models_baddebt import BadDebtAllowance, AgingBucket
from .forms import (
    CurrencyForm, ExchangeRateForm,
    TaxRateForm, TaxInvoiceForm, FixedCostForm, WithholdingTaxForm,
    AccountCodeForm, VoucherForm, VoucherLineFormSet,
    AccountReceivableForm, AccountPayableForm, PaymentForm, BankAccountForm,
    AccountTransferForm, PaymentDistributionFormSet,
    CreditCardForm, CardTransactionForm,
    CostCenterForm,
    CashReceiptForm, CashReceiptCancelForm, CashReceiptItemFormSet,
    PlatformFinancialConfigForm,
)


class AccountingDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        year = today.year

        # 월별 매출 (올해)
        monthly_revenue = []
        monthly_costs = []
        monthly_fixed = []
        monthly_variable = []
        monthly_purchases = []
        months = []
        for m in range(1, 13):
            months.append(f'{m}월')
            revenue = Order.objects.filter(
                is_active=True,
                order_date__year=year, order_date__month=m,
                status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED', 'CLOSED'],
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            monthly_revenue.append(int(revenue))

            # 고정비: FixedCost
            fixed = FixedCost.objects.filter(
                is_active=True,
                month__year=year, month__month=m,
            ).aggregate(total=Sum('amount'))['total'] or 0

            # 변동비: 매출원가 (COGS — 판매 주문의 원가 × 수량)
            variable = OrderItem.objects.filter(
                is_active=True,
                order__is_active=True,
                order__order_date__year=year, order__order_date__month=m,
                order__status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED', 'CLOSED'],
            ).aggregate(
                total=Sum(F('cost_price') * F('quantity'))
            )['total'] or 0

            # 당기매입액: 해당 월 발주(입고완료 기준) 금액
            from apps.purchase.models import PurchaseOrder
            purchases = PurchaseOrder.objects.filter(
                is_active=True,
                order_date__year=year, order_date__month=m,
                status__in=['RECEIVED', 'PARTIAL_RECEIVED', 'CONFIRMED'],
            ).aggregate(total=Sum('grand_total'))['total'] or 0

            monthly_fixed.append(int(fixed))
            monthly_variable.append(int(variable))
            monthly_costs.append(int(fixed) + int(variable))
            monthly_purchases.append(int(purchases))

        ctx['months_json'] = months
        ctx['revenue_json'] = monthly_revenue
        ctx['costs_json'] = monthly_costs
        ctx['fixed_costs_json'] = monthly_fixed
        ctx['variable_costs_json'] = monthly_variable
        ctx['purchases_json'] = monthly_purchases

        # 올해 총 매출/비용/이익
        ctx['year_revenue'] = sum(monthly_revenue)
        ctx['year_costs'] = sum(monthly_costs)
        ctx['year_profit'] = ctx['year_revenue'] - ctx['year_costs']
        ctx['year_fixed'] = sum(monthly_fixed)
        ctx['year_variable'] = sum(monthly_variable)
        ctx['year_purchases'] = sum(monthly_purchases)

        # 제품별 이익률
        products = Product.objects.filter(product_type='FINISHED', unit_price__gt=0)
        product_margins = []
        for p in products[:10]:
            product_margins.append({
                'name': p.name,
                'margin': float(p.profit_margin),
                'unit_price': int(p.unit_price),
                'cost_price': int(p.cost_price),
            })
        ctx['product_margins_json'] = product_margins

        # 이번 분기 부가세
        quarter = (today.month - 1) // 3 + 1
        q_start_month = (quarter - 1) * 3 + 1
        sales_tax = TaxInvoice.objects.filter(
            is_active=True,
            invoice_type='SALES', issue_date__year=year,
            issue_date__month__gte=q_start_month,
            issue_date__month__lte=q_start_month + 2,
        ).aggregate(total=Sum('tax_amount'))['total'] or 0
        purchase_tax = TaxInvoice.objects.filter(
            is_active=True,
            invoice_type='PURCHASE', issue_date__year=year,
            issue_date__month__gte=q_start_month,
            issue_date__month__lte=q_start_month + 2,
        ).aggregate(total=Sum('tax_amount'))['total'] or 0
        ctx['quarter'] = quarter
        ctx['sales_tax'] = sales_tax
        ctx['purchase_tax'] = purchase_tax
        ctx['net_vat'] = sales_tax - purchase_tax

        return ctx


# === 세율 ===
class TaxRateListView(ManagerRequiredMixin, ListView):
    model = TaxRate
    template_name = 'accounting/taxrate_list.html'
    context_object_name = 'tax_rates'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class TaxRateCreateView(ManagerRequiredMixin, CreateView):
    model = TaxRate
    form_class = TaxRateForm
    template_name = 'accounting/taxrate_form.html'
    success_url = reverse_lazy('accounting:taxrate_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class TaxRateUpdateView(ManagerRequiredMixin, UpdateView):
    model = TaxRate
    form_class = TaxRateForm
    template_name = 'accounting/taxrate_form.html'
    success_url = reverse_lazy('accounting:taxrate_list')


# === 세금계산서 ===
class TaxInvoiceListView(ManagerRequiredMixin, ListView):
    model = TaxInvoice
    template_name = 'accounting/taxinvoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related(
            'partner', 'order',
        )
        inv_type = self.request.GET.get('type')
        if inv_type:
            qs = qs.filter(invoice_type=inv_type)
        return qs


class TaxInvoiceCreateView(ManagerRequiredMixin, CreateView):
    model = TaxInvoice
    form_class = TaxInvoiceForm
    template_name = 'accounting/taxinvoice_form.html'
    success_url = reverse_lazy('accounting:taxinvoice_list')

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['invoice_number'] = generate_document_number(TaxInvoice, 'invoice_number', 'TI')
        # Order 연동 시 sales_channel → PlatformFinancialConfig 조회해 issuer/tax 기본값 추론
        order_id = self.request.GET.get('order_id') or self.request.GET.get('order')
        if order_id:
            try:
                order = Order.objects.filter(pk=int(order_id), is_active=True).first()
            except (TypeError, ValueError):
                order = None
            if order:
                initial['order'] = order.pk
                initial['partner'] = order.partner_id
                initial['tax_type'] = getattr(order, 'tax_type', TaxInvoice.TaxType.TAXABLE)
                config = PlatformFinancialConfig.objects.filter(
                    code=getattr(order, 'sales_channel', ''), is_enabled=True, is_active=True,
                ).first()
                if config and config.tax_invoice_issuer == PlatformFinancialConfig.IssuerType.PLATFORM:
                    initial['issuer_type'] = TaxInvoice.IssuerType.PLATFORM
                    initial['platform_name'] = config.name
        return initial

    def form_valid(self, form):
        order = form.cleaned_data.get('order')
        if order is not None:
            channel = getattr(order, 'sales_channel', '')
            if channel:
                config = PlatformFinancialConfig.objects.filter(
                    code=channel, is_enabled=True, is_active=True,
                ).first()
                if (config
                        and config.tax_invoice_issuer == PlatformFinancialConfig.IssuerType.PLATFORM
                        and form.cleaned_data.get('issuer_type') != TaxInvoice.IssuerType.PLATFORM):
                    form.add_error(
                        None,
                        f'{config.name}이(가) 세금계산서를 이미 대행 발행하는 채널입니다. '
                        f'중복발행을 차단합니다. 발행주체를 "플랫폼대행"으로 변경하거나 '
                        f'해당 주문의 판매채널/플랫폼 설정을 확인하세요.',
                    )
                    return self.form_invalid(form)
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class TaxInvoiceDetailView(ManagerRequiredMixin, DetailView):
    model = TaxInvoice
    template_name = 'accounting/taxinvoice_detail.html'
    slug_field = 'invoice_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'partner', 'order',
        )


class TaxInvoiceUpdateView(ManagerRequiredMixin, UpdateView):
    model = TaxInvoice
    form_class = TaxInvoiceForm
    template_name = 'accounting/taxinvoice_form.html'
    slug_field = 'invoice_number'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('accounting:taxinvoice_list')


# === 부가세 집계 ===
class VATSummaryView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/vat_summary.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        ctx['year'] = year
        ctx['years'] = list(range(date.today().year, date.today().year - 5, -1))

        quarters = []
        for q in range(1, 5):
            m_start = (q - 1) * 3 + 1
            m_end = m_start + 2
            sales = TaxInvoice.objects.filter(
                is_active=True,
                invoice_type='SALES', issue_date__year=year,
                issue_date__month__gte=m_start, issue_date__month__lte=m_end,
            ).aggregate(
                supply=Sum('supply_amount'), tax=Sum('tax_amount'),
            )
            purchase = TaxInvoice.objects.filter(
                is_active=True,
                invoice_type='PURCHASE', issue_date__year=year,
                issue_date__month__gte=m_start, issue_date__month__lte=m_end,
            ).aggregate(
                supply=Sum('supply_amount'), tax=Sum('tax_amount'),
            )
            s_tax = sales['tax'] or 0
            p_tax = purchase['tax'] or 0
            quarters.append({
                'quarter': q,
                'sales_supply': sales['supply'] or 0,
                'sales_tax': s_tax,
                'purchase_supply': purchase['supply'] or 0,
                'purchase_tax': p_tax,
                'net_vat': s_tax - p_tax,
            })
        ctx['quarters'] = quarters
        ctx['annual_total'] = {
            'sales_tax': sum(q['sales_tax'] for q in quarters),
            'purchase_tax': sum(q['purchase_tax'] for q in quarters),
            'net_vat': sum(q['net_vat'] for q in quarters),
        }
        return ctx


# === 고정비 ===
class FixedCostListView(ManagerRequiredMixin, ListView):
    model = FixedCost
    template_name = 'accounting/fixedcost_list.html'
    context_object_name = 'costs'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True)
        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(category=category)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        ctx['monthly_total'] = FixedCost.objects.filter(
            is_active=True,
            month__year=today.year, month__month=today.month,
        ).aggregate(total=Sum('amount'))['total'] or 0
        ctx['categories'] = FixedCost.CostCategory.choices
        return ctx


class FixedCostCreateView(ManagerRequiredMixin, CreateView):
    model = FixedCost
    form_class = FixedCostForm
    template_name = 'accounting/fixedcost_form.html'
    success_url = reverse_lazy('accounting:fixedcost_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class FixedCostUpdateView(ManagerRequiredMixin, UpdateView):
    model = FixedCost
    form_class = FixedCostForm
    template_name = 'accounting/fixedcost_form.html'
    success_url = reverse_lazy('accounting:fixedcost_list')


# === 결제계좌 ===

class BankAccountListView(ManagerRequiredMixin, ListView):
    model = BankAccount
    template_name = 'accounting/bankaccount_list.html'
    context_object_name = 'accounts'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).select_related(
            'account_code', 'employee', 'partner',
        )


class BankAccountCreateView(ManagerRequiredMixin, CreateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = 'accounting/bankaccount_form.html'
    success_url = reverse_lazy('accounting:bankaccount_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class BankAccountUpdateView(ManagerRequiredMixin, UpdateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = 'accounting/bankaccount_form.html'
    success_url = reverse_lazy('accounting:bankaccount_list')


# === 손익분기점 ===
class BreakEvenView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/breakeven.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()

        # 월간 고정비 합계
        monthly_fixed = FixedCost.objects.filter(
            is_active=True,
            month__year=today.year, month__month=today.month,
        ).aggregate(total=Sum('amount'))['total'] or 0
        ctx['monthly_fixed_cost'] = monthly_fixed

        # 완제품별 손익분기점
        from apps.production.models import BOM
        products = Product.objects.filter(product_type='FINISHED', unit_price__gt=0)
        # BOM 일괄 조회 (N+1 방지)
        default_boms = {
            b.product_id: b
            for b in BOM.objects.filter(
                product__in=products, is_default=True,
            ).select_related('product').prefetch_related('items__material')
        }
        breakeven_data = []
        for p in products:
            # 변동비 = BOM 원가 (자재 원가)
            bom = default_boms.get(p.pk)
            variable_cost = int(bom.total_material_cost) if bom else int(p.cost_price)
            contribution = int(p.unit_price) - variable_cost

            if contribution > 0:
                bep_qty = int(monthly_fixed / contribution) if monthly_fixed else 0
                bep_revenue = bep_qty * int(p.unit_price)
            else:
                bep_qty = 0
                bep_revenue = 0

            breakeven_data.append({
                'name': p.name,
                'unit_price': int(p.unit_price),
                'variable_cost': variable_cost,
                'contribution': contribution,
                'bep_quantity': bep_qty,
                'bep_revenue': bep_revenue,
            })

        ctx['breakeven_data'] = breakeven_data
        ctx['breakeven_json'] = breakeven_data
        ctx['fixed_cost_breakdown'] = FixedCost.objects.filter(
            is_active=True,
            month__year=today.year, month__month=today.month,
        ).values('category').annotate(total=Sum('amount')).order_by('-total')
        return ctx


# === 월별 손익 ===
class MonthlyPLView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/monthly_pl.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        ctx['year'] = year
        ctx['years'] = list(range(date.today().year, date.today().year - 5, -1))

        monthly_data = []
        for m in range(1, 13):
            # 매출
            revenue = Order.objects.filter(
                is_active=True,
                order_date__year=year, order_date__month=m,
                status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED', 'CLOSED'],
            ).aggregate(total=Sum('total_amount'))['total'] or 0

            # 매출원가 (판매된 제품의 원가) — DB 레벨 집계
            cogs = OrderItem.objects.filter(
                is_active=True,
                order__is_active=True,
                order__order_date__year=year, order__order_date__month=m,
                order__status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED', 'CLOSED'],
            ).aggregate(
                total=Sum(F('cost_price') * F('quantity'))
            )['total'] or 0
            cogs = int(cogs)

            gross_profit = int(revenue) - cogs

            # 고정비
            fixed = FixedCost.objects.filter(
                is_active=True,
                month__year=year, month__month=m,
            ).aggregate(total=Sum('amount'))['total'] or 0

            net_profit = gross_profit - int(fixed)

            monthly_data.append({
                'month': m,
                'revenue': int(revenue),
                'cogs': cogs,
                'gross_profit': gross_profit,
                'fixed_costs': int(fixed),
                'net_profit': net_profit,
            })

        ctx['monthly_data'] = monthly_data
        ctx['monthly_json'] = monthly_data

        # 연간 합계
        ctx['annual'] = {
            'revenue': sum(d['revenue'] for d in monthly_data),
            'cogs': sum(d['cogs'] for d in monthly_data),
            'gross_profit': sum(d['gross_profit'] for d in monthly_data),
            'fixed_costs': sum(d['fixed_costs'] for d in monthly_data),
            'net_profit': sum(d['net_profit'] for d in monthly_data),
        }
        return ctx


class IncomeStatementView(ManagerRequiredMixin, TemplateView):
    """정식 손익계산서 — VoucherLine 기반 K-GAAP 9단계.

    계정과목 prefix 규약 (한국 기업회계기준):
      4xx   매출 (REVENUE)
      501~509 매출원가 (EXPENSE)
      52x~58x 판매비와관리비 (EXPENSE)
      47x   영업외수익 (REVENUE, 이자수익/수입임대료 등)
      91x~92x 영업외비용 (EXPENSE, 이자비용/기부금 등)
      998   법인세비용 (EXPENSE)

    위 규약은 관례이며 실제 AccountCode는 사용자가 자유롭게 등록한다.
    prefix 매칭에 해당하지 않는 REVENUE/EXPENSE는 '기타'로 분류한다.
    """

    template_name = 'accounting/income_statement.html'

    REVENUE_PREFIX = ('4',)
    COGS_PREFIXES = ('501', '502', '503', '504', '505', '506', '507', '508', '509')
    NONOP_REVENUE_PREFIXES = ('47',)
    NONOP_EXPENSE_PREFIXES = ('91', '92')
    INCOME_TAX_PREFIXES = ('998', '999')

    def _bucket(self, code: str, account_type: str) -> str:
        if account_type == 'REVENUE':
            for p in self.NONOP_REVENUE_PREFIXES:
                if code.startswith(p):
                    return 'nonop_revenue'
            for p in self.REVENUE_PREFIX:
                if code.startswith(p):
                    return 'sales'
            return 'nonop_revenue'
        if account_type == 'EXPENSE':
            for p in self.INCOME_TAX_PREFIXES:
                if code.startswith(p):
                    return 'income_tax'
            for p in self.COGS_PREFIXES:
                if code.startswith(p):
                    return 'cogs'
            for p in self.NONOP_EXPENSE_PREFIXES:
                if code.startswith(p):
                    return 'nonop_expense'
            return 'sga'
        return 'skip'

    def _aggregate_period(self, period_start, period_end):
        qs = VoucherLine.objects.filter(
            is_active=True, voucher__is_active=True,
            voucher__approval_status='APPROVED',
            voucher__voucher_date__gte=period_start,
            voucher__voucher_date__lte=period_end,
            account__account_type__in=['REVENUE', 'EXPENSE'],
        ).values(
            'account__code', 'account__name', 'account__account_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('account__code')

        buckets = {
            'sales': [], 'cogs': [], 'sga': [],
            'nonop_revenue': [], 'nonop_expense': [], 'income_tax': [],
        }
        totals = {k: 0 for k in buckets}

        for row in qs:
            code = row['account__code']
            name = row['account__name']
            atype = row['account__account_type']
            debit = int(row['total_debit'] or 0)
            credit = int(row['total_credit'] or 0)
            if atype == 'REVENUE':
                amount = credit - debit
            else:
                amount = debit - credit
            bucket = self._bucket(code, atype)
            if bucket == 'skip':
                continue
            buckets[bucket].append({
                'code': code, 'name': name, 'amount': amount,
            })
            totals[bucket] += amount

        step1_revenue = totals['sales']
        step2_cogs = totals['cogs']
        step3_gross = step1_revenue - step2_cogs
        step4_sga = totals['sga']
        step5_operating = step3_gross - step4_sga
        step6_nonop_rev = totals['nonop_revenue']
        step6_nonop_exp = totals['nonop_expense']
        step7_pretax = step5_operating + step6_nonop_rev - step6_nonop_exp
        step8_tax = totals['income_tax']
        step9_net = step7_pretax - step8_tax

        return {
            'buckets': buckets,
            'step1_revenue': step1_revenue,
            'step2_cogs': step2_cogs,
            'step3_gross': step3_gross,
            'step4_sga': step4_sga,
            'step5_operating': step5_operating,
            'step6_nonop_rev': step6_nonop_rev,
            'step6_nonop_exp': step6_nonop_exp,
            'step7_pretax': step7_pretax,
            'step8_tax': step8_tax,
            'step9_net': step9_net,
        }

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        year = safe_int(self.request.GET.get('year'), today.year)
        month = self.request.GET.get('month')
        month_int = safe_int(month, 0) if month else 0
        compare = self.request.GET.get('compare', '1') == '1'

        ctx['year'] = year
        ctx['month'] = month_int
        ctx['compare'] = compare
        ctx['years'] = list(range(today.year, today.year - 5, -1))
        ctx['months'] = list(range(1, 13))

        if month_int:
            period_start = date(year, month_int, 1)
            if month_int == 12:
                period_end = date(year, 12, 31)
            else:
                period_end = date(year, month_int + 1, 1) - timedelta(days=1)
            ctx['period_label'] = f'{year}년 {month_int}월'
            prior_start = date(year - 1, month_int, 1)
            if month_int == 12:
                prior_end = date(year - 1, 12, 31)
            else:
                prior_end = date(year - 1, month_int + 1, 1) - timedelta(days=1)
            prior_label = f'{year - 1}년 {month_int}월'
        else:
            period_start = date(year, 1, 1)
            period_end = date(year, 12, 31)
            ctx['period_label'] = f'{year}년 연간'
            prior_start = date(year - 1, 1, 1)
            prior_end = date(year - 1, 12, 31)
            prior_label = f'{year - 1}년 연간'

        ctx['period_start'] = period_start
        ctx['period_end'] = period_end
        ctx['prior_period_label'] = prior_label

        current = self._aggregate_period(period_start, period_end)
        ctx['buckets'] = current['buckets']
        for key in ('step1_revenue', 'step2_cogs', 'step3_gross', 'step4_sga',
                    'step5_operating', 'step6_nonop_rev', 'step6_nonop_exp',
                    'step7_pretax', 'step8_tax', 'step9_net'):
            ctx[key] = current[key]

        if compare:
            prior = self._aggregate_period(prior_start, prior_end)
            ctx['prior'] = prior
            # YoY 증감률 (전년 대비)
            def _pct(cur, prev):
                if not prev:
                    return None
                return round((cur - prev) / abs(prev) * 100, 1)
            ctx['yoy'] = {
                'revenue': _pct(current['step1_revenue'], prior['step1_revenue']),
                'gross': _pct(current['step3_gross'], prior['step3_gross']),
                'operating': _pct(current['step5_operating'], prior['step5_operating']),
                'net': _pct(current['step9_net'], prior['step9_net']),
            }

        step1_revenue = current['step1_revenue']
        if step1_revenue:
            ctx['gross_margin_pct'] = round(current['step3_gross'] / step1_revenue * 100, 2)
            ctx['operating_margin_pct'] = round(current['step5_operating'] / step1_revenue * 100, 2)
            ctx['net_margin_pct'] = round(current['step9_net'] / step1_revenue * 100, 2)
        else:
            ctx['gross_margin_pct'] = 0
            ctx['operating_margin_pct'] = 0
            ctx['net_margin_pct'] = 0
        return ctx


# === 원천징수 ===
class WithholdingTaxListView(ManagerRequiredMixin, ListView):
    model = WithholdingTax
    template_name = 'accounting/withholding_list.html'
    context_object_name = 'withholdings'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class WithholdingTaxCreateView(ManagerRequiredMixin, CreateView):
    model = WithholdingTax
    form_class = WithholdingTaxForm
    template_name = 'accounting/withholding_form.html'
    success_url = reverse_lazy('accounting:withholding_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class WithholdingTaxUpdateView(ManagerRequiredMixin, UpdateView):
    model = WithholdingTax
    form_class = WithholdingTaxForm
    template_name = 'accounting/withholding_form.html'
    success_url = reverse_lazy('accounting:withholding_list')


class WithholdingTaxReportView(ManagerRequiredMixin, TemplateView):
    """원천세 월별 납부/신고서 — 세목별 집계 + 익월 10일 납부 기한 표시.

    `?export=excel|pdf` 로 다운로드 지원.
    """
    template_name = 'accounting/withholding_report.html'

    def get(self, request, *args, **kwargs):
        export = request.GET.get('export', '').lower()
        if export == 'excel':
            return self._render_excel(request)
        if export == 'pdf':
            return self._render_pdf(request)
        return super().get(request, *args, **kwargs)

    def _render_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font
        ctx = self.get_context_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '원천세신고서'
        bold = Font(bold=True)

        ws.cell(row=1, column=1, value=f'{ctx["year"]}년 {ctx["month"]:02d}월 원천세 신고서').font = bold
        ws.cell(row=2, column=1, value=f'납부기한: {ctx["due_date"]}')

        ws.cell(row=4, column=1, value='【세목별 집계】').font = bold
        for col, h in enumerate(['세목', '건수', '지급액', '원천징수액', '실지급액'], 1):
            ws.cell(row=5, column=col, value=h).font = bold
        for i, row in enumerate(ctx['by_type'], 6):
            ws.cell(row=i, column=1, value=row['label'])
            ws.cell(row=i, column=2, value=row['count'])
            ws.cell(row=i, column=3, value=int(row['gross']))
            ws.cell(row=i, column=4, value=int(row['tax']))
            ws.cell(row=i, column=5, value=int(row['net']))
        total_row = 6 + len(ctx['by_type'])
        ws.cell(row=total_row, column=1, value='합계').font = bold
        ws.cell(row=total_row, column=2, value=ctx['total_count']).font = bold
        ws.cell(row=total_row, column=3, value=int(ctx['total_gross'])).font = bold
        ws.cell(row=total_row, column=4, value=int(ctx['total_tax'])).font = bold
        ws.cell(row=total_row, column=5, value=int(ctx['total_net'])).font = bold

        for col in [1, 2, 3, 4, 5]:
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename=withholding_{ctx["year"]}{ctx["month"]:02d}.xlsx'
        )
        wb.save(response)
        return response

    def _render_pdf(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from apps.core.pdf import _get_font

        ctx = self.get_context_data()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename=withholding_{ctx["year"]}{ctx["month"]:02d}.pdf'
        )

        doc = SimpleDocTemplate(response, pagesize=A4)
        font = _get_font()
        styles = getSampleStyleSheet()
        styles['Title'].fontName = font
        styles['BodyText'].fontName = font

        flow = [
            Paragraph(
                f'{ctx["year"]}년 {ctx["month"]:02d}월 원천세 신고서',
                styles['Title'],
            ),
            Paragraph(f'납부기한: {ctx["due_date"]}', styles['BodyText']),
            Spacer(1, 12),
        ]
        data = [['세목', '건수', '지급액', '원천징수액', '실지급액']]
        for row in ctx['by_type']:
            data.append([
                row['label'], row['count'],
                f'{int(row["gross"]):,}', f'{int(row["tax"]):,}', f'{int(row["net"]):,}',
            ])
        data.append([
            '합계', ctx['total_count'],
            f'{int(ctx["total_gross"]):,}',
            f'{int(ctx["total_tax"]):,}',
            f'{int(ctx["total_net"]):,}',
        ])
        t = Table(data, colWidths=[120, 60, 90, 90, 90])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
        ]))
        flow.append(t)
        doc.build(flow)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        month = safe_int(self.request.GET.get('month'), date.today().month)
        month = max(1, min(12, month))

        import calendar as _cal
        period_start = date(year, month, 1)
        period_end = date(year, month, _cal.monthrange(year, month)[1])
        # 납부기한 — 지급월 익월 10일
        if month == 12:
            due_date = date(year + 1, 1, 10)
        else:
            due_date = date(year, month + 1, 10)

        base_qs = WithholdingTax.objects.filter(
            is_active=True,
            payment_date__gte=period_start,
            payment_date__lte=period_end,
        )

        by_type = []
        for key, label in WithholdingTax.TaxType.choices:
            agg = base_qs.filter(tax_type=key).aggregate(
                gross=Sum('gross_amount'),
                tax=Sum('tax_amount'),
                net=Sum('net_amount'),
                cnt=Count('pk'),
            )
            by_type.append({
                'code': key, 'label': label,
                'gross': agg['gross'] or 0,
                'tax': agg['tax'] or 0,
                'net': agg['net'] or 0,
                'count': agg['cnt'] or 0,
            })

        totals = base_qs.aggregate(
            gross=Sum('gross_amount'),
            tax=Sum('tax_amount'),
            net=Sum('net_amount'),
            cnt=Count('pk'),
        )

        ctx.update({
            'year': year,
            'month': month,
            'period_start': period_start,
            'period_end': period_end,
            'due_date': due_date,
            'by_type': by_type,
            'total_gross': totals['gross'] or 0,
            'total_tax': totals['tax'] or 0,
            'total_net': totals['net'] or 0,
            'total_count': totals['cnt'] or 0,
            'years': list(range(date.today().year, date.today().year - 5, -1)),
            'months': list(range(1, 13)),
            'items': base_qs.order_by('payment_date', 'pk'),
        })
        return ctx


# === 계정과목 ===
class AccountCodeListView(ManagerRequiredMixin, ListView):
    model = AccountCode
    template_name = 'accounting/accountcode_list.html'
    context_object_name = 'account_codes'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).select_related('parent')


class AccountCodeCreateView(ManagerRequiredMixin, CreateView):
    model = AccountCode
    form_class = AccountCodeForm
    template_name = 'accounting/accountcode_form.html'
    success_url = reverse_lazy('accounting:accountcode_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class AccountCodeUpdateView(ManagerRequiredMixin, UpdateView):
    model = AccountCode
    form_class = AccountCodeForm
    template_name = 'accounting/accountcode_form.html'
    success_url = reverse_lazy('accounting:accountcode_list')


# === 전표 ===
class VoucherListView(ManagerRequiredMixin, ListView):
    model = Voucher
    template_name = 'accounting/voucher_list.html'
    context_object_name = 'vouchers'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related(
            'approved_by',
        )
        voucher_type = self.request.GET.get('type')
        status = self.request.GET.get('status')
        if voucher_type:
            qs = qs.filter(voucher_type=voucher_type)
        if status:
            qs = qs.filter(approval_status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['type_choices'] = Voucher.VoucherType.choices
        ctx['status_choices'] = Voucher.ApprovalStatus.choices
        return ctx


class VoucherCreateView(ManagerRequiredMixin, CreateView):
    model = Voucher
    form_class = VoucherForm
    template_name = 'accounting/voucher_form.html'
    success_url = reverse_lazy('accounting:voucher_list')

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['voucher_number'] = generate_document_number(Voucher, 'voucher_number', 'VC')
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = VoucherLineFormSet(self.request.POST)
        else:
            ctx['formset'] = VoucherLineFormSet()
        return ctx

    def form_valid(self, form):
        voucher_date = form.cleaned_data.get('voucher_date')
        if voucher_date:
            closed = ClosingPeriod.objects.filter(
                year=voucher_date.year,
                month=voucher_date.month,
                is_closed=True,
                is_active=True,
            ).exists()
            if closed:
                messages.error(
                    self.request,
                    f'{voucher_date.year}년 {voucher_date.month}월은 결산 마감되어 전표를 생성할 수 없습니다.',
                )
                return self.form_invalid(form)

        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            try:
                with transaction.atomic():
                    self.object = form.save(commit=False)
                    self.object.created_by = self.request.user
                    self.object.save()
                    formset.instance = self.object
                    formset.save()
                    if not self.object.is_balanced:
                        raise ValueError('unbalanced')
            except ValueError:
                messages.error(
                    self.request,
                    '차변 합계와 대변 합계가 일치하지 않습니다.',
                )
                return self.form_invalid(form)

            # 예산 초과 경고 (차단하지 않음)
            voucher_date = self.object.voucher_date
            if voucher_date:
                for line in self.object.lines.select_related('account').all():
                    if line.account.account_type == 'EXPENSE' and line.debit > 0:
                        budget = Budget.objects.filter(
                            account=line.account,
                            year=voucher_date.year,
                            month=voucher_date.month,
                            is_active=True,
                        ).first()
                        if budget and budget.actual_amount > int(budget.budget_amount):
                            messages.warning(
                                self.request,
                                f'{line.account.name} 예산 초과: '
                                f'예산 {int(budget.budget_amount):,}원 / '
                                f'실적 {budget.actual_amount:,}원 '
                                f'(집행률 {budget.execution_rate}%)',
                            )

            return redirect(self.get_success_url())
        return self.form_invalid(form)


class VoucherDetailView(ManagerRequiredMixin, DetailView):
    model = Voucher
    template_name = 'accounting/voucher_detail.html'
    slug_field = 'voucher_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['lines'] = self.object.lines.select_related(
            'account',
        ).all()
        return ctx


class VoucherUpdateView(ManagerRequiredMixin, UpdateView):
    model = Voucher
    form_class = VoucherForm
    template_name = 'accounting/voucher_form.html'
    slug_field = 'voucher_number'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('accounting:voucher_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = VoucherLineFormSet(self.request.POST, instance=self.object)
        else:
            ctx['formset'] = VoucherLineFormSet(instance=self.object)
        return ctx

    def form_valid(self, form):
        voucher_date = form.cleaned_data.get('voucher_date')
        if voucher_date:
            closed = ClosingPeriod.objects.filter(
                year=voucher_date.year,
                month=voucher_date.month,
                is_closed=True,
                is_active=True,
            ).exists()
            if closed:
                messages.error(
                    self.request,
                    f'{voucher_date.year}년 {voucher_date.month}월은 결산 마감되어 전표를 수정할 수 없습니다.',
                )
                return self.form_invalid(form)

        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            try:
                with transaction.atomic():
                    self.object = form.save()
                    formset.instance = self.object
                    formset.save()
                    if not self.object.is_balanced:
                        raise ValueError('unbalanced')
            except ValueError:
                messages.error(
                    self.request,
                    '차변 합계와 대변 합계가 일치하지 않습니다.',
                )
                return self.form_invalid(form)
            return redirect(self.get_success_url())
        return self.form_invalid(form)


# === 세금계산서 PDF ===
class TaxInvoicePDFView(ManagerRequiredMixin, DetailView):
    model = TaxInvoice
    slug_field = 'invoice_number'
    slug_url_kwarg = 'slug'

    def get(self, request, *args, **kwargs):
        invoice = self.get_object()
        from apps.core.pdf import generate_tax_invoice_pdf
        return generate_tax_invoice_pdf(invoice)


# === 매출정산 PDF ===
class SalesSettlementPDFView(ManagerRequiredMixin, DetailView):
    model = SalesSettlement
    slug_field = 'settlement_number'
    slug_url_kwarg = 'slug'

    def get(self, request, *args, **kwargs):
        settlement = self.get_object()
        from apps.core.pdf import generate_settlement_pdf
        return generate_settlement_pdf(settlement)


# === 미수금 (Accounts Receivable) ===
class ARListView(ManagerRequiredMixin, ListView):
    model = AccountReceivable
    template_name = 'accounting/ar_list.html'
    context_object_name = 'receivables'
    paginate_by = 20

    def get_queryset(self):
        # 자동 연체 전환: due_date가 지난 미완납 AR을 OVERDUE로 갱신
        today = date.today()
        AccountReceivable.objects.filter(
            is_active=True,
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL'],
        ).update(status='OVERDUE')

        qs = super().get_queryset().filter(is_active=True).select_related('partner', 'order')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        partner_id = self.request.GET.get('partner')
        if partner_id:
            qs = qs.filter(partner_id=partner_id)
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(partner__name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = AccountReceivable.Status.choices
        from apps.sales.models import Partner
        ctx['partners'] = Partner.objects.filter(is_active=True)
        total = AccountReceivable.objects.filter(
            is_active=True,
            status__in=['PENDING', 'PARTIAL', 'OVERDUE']
        ).aggregate(total=Sum('amount'), paid=Sum('paid_amount'))
        ctx['total_remaining'] = (total['total'] or 0) - (total['paid'] or 0)
        return ctx


class ARDetailView(ManagerRequiredMixin, DetailView):
    model = AccountReceivable
    template_name = 'accounting/ar_detail.html'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'partner', 'order',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['payments'] = (
            self.object.payments
            .select_related('partner')
            .order_by('-payment_date')
        )
        return ctx


class ARCreateView(ManagerRequiredMixin, CreateView):
    model = AccountReceivable
    form_class = AccountReceivableForm
    template_name = 'accounting/ar_form.html'
    success_url = reverse_lazy('accounting:ar_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class PaymentCreateView(ManagerRequiredMixin, CreateView):
    """입금 등록 - AR에 대한 결제"""
    model = Payment
    form_class = PaymentForm
    template_name = 'accounting/payment_form.html'

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['payment_number'] = generate_document_number(Payment, 'payment_number', 'PM')
        ar = AccountReceivable.objects.filter(pk=self.kwargs['pk']).first()
        if ar:
            initial['partner'] = ar.partner_id
            initial['payment_type'] = 'RECEIPT'
            initial['amount'] = ar.remaining_amount
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['receivable'] = get_object_or_404(AccountReceivable, pk=self.kwargs['pk'])
        ctx['is_receipt'] = True
        return ctx

    def form_valid(self, form):
        with transaction.atomic():
            ar = AccountReceivable.objects.select_for_update().get(pk=self.kwargs['pk'])
            amount = form.cleaned_data['amount']
            if amount > ar.remaining_amount:
                form.add_error('amount', f'잔액({ar.remaining_amount:,}원)을 초과할 수 없습니다.')
                return self.form_invalid(form)

            payment = form.save(commit=False)
            payment.receivable = ar
            payment.partner = ar.partner
            payment.payment_type = 'RECEIPT'
            payment.created_by = self.request.user
            payment.save()

            ar.paid_amount = F('paid_amount') + amount
            ar.save(update_fields=['paid_amount', 'updated_at'])
            ar.refresh_from_db()

            if ar.paid_amount >= ar.amount:
                ar.status = 'PAID'
            elif ar.paid_amount > 0:
                ar.status = 'PARTIAL'
            ar.save(update_fields=['status', 'updated_at'])

        messages.success(self.request, f'{amount:,}원 입금 처리 완료')
        return redirect('accounting:ar_detail', pk=ar.pk)

    def get_success_url(self):
        return reverse_lazy('accounting:ar_detail', args=[self.kwargs['pk']])


# === 미지급금 (Accounts Payable) ===
class APListView(ManagerRequiredMixin, ListView):
    model = AccountPayable
    template_name = 'accounting/ap_list.html'
    context_object_name = 'payables'
    paginate_by = 20

    def get_queryset(self):
        # 자동 연체 전환: due_date가 지난 미완납 AP를 OVERDUE로 갱신
        today = date.today()
        AccountPayable.objects.filter(
            is_active=True,
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL'],
        ).update(status='OVERDUE')

        qs = super().get_queryset().filter(is_active=True).select_related(
            'partner', 'purchase_order',
        )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        partner_id = self.request.GET.get('partner')
        if partner_id:
            qs = qs.filter(partner_id=partner_id)
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(partner__name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = AccountPayable.Status.choices
        from apps.sales.models import Partner
        ctx['partners'] = Partner.objects.filter(is_active=True)
        total = AccountPayable.objects.filter(
            is_active=True,
            status__in=['PENDING', 'PARTIAL', 'OVERDUE']
        ).aggregate(total=Sum('amount'), paid=Sum('paid_amount'))
        ctx['total_remaining'] = (total['total'] or 0) - (total['paid'] or 0)
        return ctx


class APDetailView(ManagerRequiredMixin, DetailView):
    model = AccountPayable
    template_name = 'accounting/ap_detail.html'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'partner', 'purchase_order',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['payments'] = (
            self.object.payments
            .select_related('partner')
            .order_by('-payment_date')
        )
        return ctx


class APCreateView(ManagerRequiredMixin, CreateView):
    model = AccountPayable
    form_class = AccountPayableForm
    template_name = 'accounting/ap_form.html'
    success_url = reverse_lazy('accounting:ap_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class DisbursementCreateView(ManagerRequiredMixin, CreateView):
    """출금 등록 - AP에 대한 결제"""
    model = Payment
    form_class = PaymentForm
    template_name = 'accounting/payment_form.html'

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['payment_number'] = generate_document_number(Payment, 'payment_number', 'PM')
        ap = AccountPayable.objects.filter(pk=self.kwargs['pk']).first()
        if ap:
            initial['partner'] = ap.partner_id
            initial['payment_type'] = 'DISBURSEMENT'
            initial['amount'] = ap.remaining_amount
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['payable'] = get_object_or_404(AccountPayable, pk=self.kwargs['pk'])
        ctx['is_receipt'] = False
        return ctx

    def form_valid(self, form):
        with transaction.atomic():
            ap = AccountPayable.objects.select_for_update().get(pk=self.kwargs['pk'])
            amount = form.cleaned_data['amount']
            if amount > ap.remaining_amount:
                form.add_error('amount', f'잔액({ap.remaining_amount:,}원)을 초과할 수 없습니다.')
                return self.form_invalid(form)

            payment = form.save(commit=False)
            payment.payable = ap
            payment.partner = ap.partner
            payment.payment_type = 'DISBURSEMENT'
            payment.created_by = self.request.user
            payment.save()

            ap.paid_amount = F('paid_amount') + amount
            ap.save(update_fields=['paid_amount', 'updated_at'])
            ap.refresh_from_db()

            if ap.paid_amount >= ap.amount:
                ap.status = 'PAID'
            elif ap.paid_amount > 0:
                ap.status = 'PARTIAL'
            ap.save(update_fields=['status', 'updated_at'])

        messages.success(self.request, f'{amount:,}원 출금 처리 완료')
        return redirect('accounting:ap_detail', pk=ap.pk)

    def get_success_url(self):
        return reverse_lazy('accounting:ap_detail', args=[self.kwargs['pk']])


# === 계좌 상세 ===
class BankAccountDetailView(ManagerRequiredMixin, DetailView):
    model = BankAccount
    template_name = 'accounting/bankaccount_detail.html'

    def get_queryset(self):
        return super().get_queryset().select_related('account_code')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        account = self.object
        ctx['payments'] = Payment.objects.filter(
            bank_account=account, is_active=True,
        ).select_related('partner').order_by('-payment_date')[:20]
        ctx['transfers'] = AccountTransfer.objects.filter(
            Q(from_account=account) | Q(to_account=account),
            is_active=True,
        ).select_related('from_account', 'to_account').order_by('-transfer_date')[:20]
        ctx['distributions'] = PaymentDistribution.objects.filter(
            bank_account=account, is_active=True,
        ).select_related('payment').order_by('-pk')[:20]
        return ctx


# === 계좌 대시보드 ===
class BankAccountDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/bankaccount_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        accounts = BankAccount.objects.filter(is_active=True)
        ctx['accounts'] = accounts
        ctx['total_balance'] = accounts.aggregate(
            total=Sum('balance'),
        )['total'] or 0
        return ctx


# === 계좌이체 ===
class AccountTransferListView(ManagerRequiredMixin, ListView):
    model = AccountTransfer
    template_name = 'accounting/transfer_list.html'
    context_object_name = 'transfers'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(
            is_active=True,
        ).select_related('from_account', 'to_account')


class AccountTransferCreateView(ManagerRequiredMixin, CreateView):
    model = AccountTransfer
    form_class = AccountTransferForm
    template_name = 'accounting/transfer_form.html'
    success_url = reverse_lazy('accounting:transfer_list')

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['transfer_number'] = generate_document_number(
            AccountTransfer, 'transfer_number', 'BT'
        )
        initial['transfer_date'] = date.today()
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


# === 결제 분배 ===
class PaymentDistributeView(ManagerRequiredMixin, DetailView):
    model = Payment
    template_name = 'accounting/payment_distribute.html'
    slug_field = 'payment_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related('partner', 'bank_account')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = PaymentDistributionFormSet(
                self.request.POST, instance=self.object,
            )
        else:
            ctx['formset'] = PaymentDistributionFormSet(instance=self.object)
        return ctx

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            with transaction.atomic():
                distributions = formset.save(commit=False)
                for dist in distributions:
                    dist.created_by = request.user
                    dist.save()
                for obj in formset.deleted_objects:
                    obj.delete()
            messages.success(request, '결제 분배가 저장되었습니다.')
            return redirect('accounting:payment_distribute', slug=self.object.payment_number)
        return self.render_to_response(ctx)


# ── 일괄 가져오기 ──

class FixedCostImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import FixedCostResource
            return export_resource_data(FixedCostResource(), '고정비_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '고정비 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('accounting:fixedcost_list')
        ctx['sample_url'] = reverse_lazy('accounting:fixedcost_import_sample')
        ctx['field_hints'] = [
            'category: RENT(임대료), LABOR(인건비), EQUIPMENT(장비), '
            'INSURANCE(보험), TELECOM(통신비), SUBSCRIPTION(구독), OTHER(기타)',
            'month: YYYY-MM-DD 형식 (해당월 1일)',
            '동일 (name, month) 조합이 있으면 금액이 수정됩니다.',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from apps.core.import_views import parse_import_file, build_preview, collect_errors
        from .resources import FixedCostResource

        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')

        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = FixedCostResource()
        try:
            data = parse_import_file(import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다.')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = build_preview(result, data)
            ctx['errors'] = collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = collect_errors(result)
                return self.render_to_response(ctx)
            total = result.totals.get('new', 0) + result.totals.get('update', 0)
            messages.success(request, f'고정비 {total}건이 성공적으로 가져오기 되었습니다.')
            return HttpResponseRedirect(reverse_lazy('accounting:fixedcost_list'))


class FixedCostImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('name', 25), ('category', 15), ('amount', 15),
            ('month', 12), ('is_recurring', 10),
        ]
        rows = [
            ['사무실 임대료', 'RENT', 1500000, '2026-03-01', True],
            ['서버 호스팅', 'SUBSCRIPTION', 100000, '2026-03-01', True],
        ]
        return export_to_excel(
            '고정비_가져오기_양식', headers, rows,
            filename='고정비_가져오기_양식.xlsx',
            required_columns=[0, 2, 3],  # name, amount, month
        )


class AccountCodeImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import AccountCodeResource
            return export_resource_data(AccountCodeResource(), '계정과목_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '계정과목 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('accounting:accountcode_list')
        ctx['sample_url'] = reverse_lazy(
            'accounting:accountcode_import_sample',
        )
        ctx['field_hints'] = [
            '계정코드(code)가 동일하면 기존 계정이 수정됩니다.',
            'account_type: ASSET(자산), LIABILITY(부채), '
            'EQUITY(자본), REVENUE(수익), EXPENSE(비용)',
            'parent_code: 상위 계정코드 (없으면 비워두기)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from apps.core.import_views import (
            parse_import_file, build_preview, collect_errors,
        )
        from .resources import AccountCodeResource

        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = AccountCodeResource()
        try:
            data = parse_import_file(import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다.')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = build_preview(result, data)
            ctx['errors'] = collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(request, f'{total}건 가져오기 완료.')
            return HttpResponseRedirect(
                str(reverse_lazy('accounting:accountcode_list')),
            )


class AccountCodeImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('code', 12), ('name', 25),
            ('account_type', 12), ('parent_code', 12),
        ]
        rows = [
            ['1000', '자산', 'ASSET', ''],
            ['1100', '유동자산', 'ASSET', '1000'],
            ['1110', '현금및현금성자산', 'ASSET', '1100'],
            ['4000', '매출', 'REVENUE', ''],
            ['5000', '매출원가', 'EXPENSE', ''],
        ]
        return export_to_excel(
            '계정과목_가져오기_양식', headers, rows,
            filename='계정과목_가져오기_양식.xlsx',
            required_columns=[0, 1, 2],  # code, name, account_type
        )


class TaxInvoiceImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import TaxInvoiceResource
            return export_resource_data(TaxInvoiceResource(), '세금계산서_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '세금계산서 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('accounting:taxinvoice_list')
        ctx['sample_url'] = reverse_lazy(
            'accounting:taxinvoice_import_sample',
        )
        ctx['field_hints'] = [
            '세금계산서번호(invoice_number)가 동일하면 수정됩니다.',
            'invoice_type: SALES(매출), PURCHASE(매입)',
            'partner_code: 거래처코드',
            '금액: 숫자만 입력 (원 단위)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from apps.core.import_views import (
            parse_import_file, build_preview, collect_errors,
        )
        from .resources import TaxInvoiceResource

        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = TaxInvoiceResource()
        try:
            data = parse_import_file(import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다.')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = build_preview(result, data)
            ctx['errors'] = collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(request, f'{total}건 가져오기 완료.')
            return HttpResponseRedirect(
                str(reverse_lazy('accounting:taxinvoice_list')),
            )


class TaxInvoiceImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('invoice_number', 18), ('invoice_type', 12),
            ('partner_code', 15), ('issue_date', 12),
            ('supply_amount', 15), ('tax_amount', 12),
            ('total_amount', 15), ('description', 25),
        ]
        rows = [
            ['INV-2026-001', 'SALES', 'PTN-001', '2026-03-01',
             1000000, 100000, 1100000, '3월 납품건'],
            ['INV-2026-002', 'PURCHASE', 'PTN-002', '2026-03-05',
             500000, 50000, 550000, '자재 구매'],
        ]
        return export_to_excel(
            '세금계산서_가져오기_양식', headers, rows,
            filename='세금계산서_가져오기_양식.xlsx',
            money_columns=[4, 5, 6],
            required_columns=[0, 1, 3, 4],  # invoice_number, invoice_type, issue_date, supply_amount
        )


# === 전표 일괄 가져오기 ===

class VoucherImportView(BaseImportView):
    resource_class = None
    page_title = '전표 일괄 가져오기'
    cancel_url = reverse_lazy('accounting:voucher_list')
    sample_url = reverse_lazy('accounting:voucher_import_sample')
    export_filename = '전표_데이터'
    field_hints = [
        '전표번호(voucher_number)가 동일하면 기존 전표가 수정됩니다.',
        'voucher_type: RECEIPT(입금), PAYMENT(출금), TRANSFER(대체)',
    ]

    def get_resource(self):
        from .resources import VoucherResource
        return VoucherResource()


class VoucherImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('voucher_number', 18), ('voucher_type', 12),
            ('voucher_date', 12), ('description', 30),
        ]
        rows = [
            ['VCH-2026-001', 'RECEIPT', '2026-03-01', '매출대금 입금'],
            ['VCH-2026-002', 'PAYMENT', '2026-03-05', '원자재 구매'],
        ]
        return export_to_excel(
            '전표_가져오기_양식', headers, rows,
            filename='전표_가져오기_양식.xlsx',
            required_columns=[0, 1, 2, 3],
        )


# === 원천징수 일괄 가져오기 ===

class WithholdingImportView(BaseImportView):
    resource_class = None
    page_title = '원천징수 일괄 가져오기'
    cancel_url = reverse_lazy('accounting:withholding_list')
    sample_url = reverse_lazy('accounting:withholding_import_sample')
    export_filename = '원천징수_데이터'
    field_hints = [
        '소득자명(payee_name) + 지급일(payment_date) 조합이 동일하면 수정됩니다.',
        'tax_type: INCOME(소득세), CORPORATE(법인세), RESIDENT(주민세)',
    ]

    def get_resource(self):
        from .resources import WithholdingTaxResource
        return WithholdingTaxResource()


class WithholdingImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('tax_type', 12), ('payee_name', 15),
            ('payment_date', 12), ('gross_amount', 15),
            ('tax_rate', 10), ('tax_amount', 15), ('net_amount', 15),
        ]
        rows = [
            ['INCOME', '홍길동', '2026-03-10', 3000000,
             3.3, 99000, 2901000],
        ]
        return export_to_excel(
            '원천징수_가져오기_양식', headers, rows,
            filename='원천징수_가져오기_양식.xlsx',
            money_columns=[3, 5, 6],
            required_columns=[0, 1, 2, 3, 4, 5, 6],
        )


# === 원가 정산 ===

class CostSettlementListView(ManagerRequiredMixin, ListView):
    model = CostSettlement
    template_name = 'accounting/settlement_list.html'
    context_object_name = 'settlements'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class CostSettlementDetailView(ManagerRequiredMixin, DetailView):
    model = CostSettlement
    template_name = 'accounting/settlement_detail.html'
    context_object_name = 'settlement'
    slug_field = 'settlement_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = (
            self.object.items
            .select_related('product')
            .order_by('product__code')
        )
        return ctx


class CostSettlementCreateView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/settlement_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        # 기본값: 전월
        if today.month == 1:
            default_year, default_month = today.year - 1, 12
        else:
            default_year, default_month = today.year, today.month - 1
        ctx['default_year'] = default_year
        ctx['default_month'] = default_month
        ctx['period_choices'] = CostSettlement.Period.choices
        return ctx

    def post(self, request, *args, **kwargs):
        from calendar import monthrange

        period_type = request.POST.get('period_type', 'MONTHLY')
        year = safe_int(request.POST.get('year'), date.today().year)
        month = safe_int(request.POST.get('month'), date.today().month)

        # 정산 기간 계산
        if period_type == 'MONTHLY':
            period_start = date(year, month, 1)
            _, last_day = monthrange(year, month)
            period_end = date(year, month, last_day)
        elif period_type == 'QUARTERLY':
            q_start_month = ((month - 1) // 3) * 3 + 1
            period_start = date(year, q_start_month, 1)
            q_end_month = q_start_month + 2
            _, last_day = monthrange(year, q_end_month)
            period_end = date(year, q_end_month, last_day)
        else:  # YEARLY
            period_start = date(year, 1, 1)
            period_end = date(year, 12, 31)

        # 중복 검증
        exists = CostSettlement.objects.filter(
            period_type=period_type,
            period_start=period_start,
            period_end=period_end,
            is_active=True,
        ).exists()
        if exists:
            messages.error(
                request,
                f'해당 기간({period_start}~{period_end}) 정산이 '
                f'이미 존재합니다.',
            )
            return self.get(request, *args, **kwargs)

        with transaction.atomic():
            settlement = CostSettlement.objects.create(
                period_type=period_type,
                period_start=period_start,
                period_end=period_end,
                created_by=request.user,
            )

            # 표준원가 자동 재계산: BOM 자재원가 → 노무비/간접비 → Product.cost_price 갱신
            from apps.production.models import StandardCost
            updated_count = 0
            for sc in StandardCost.objects.filter(is_current=True, is_active=True).select_related('product'):
                sc.calculate_material_cost()
                sc.save()  # labor_cost, overhead_cost, total 자동 재계산
                if sc.product.cost_price != sc.total_standard_cost:
                    sc.product.cost_price = sc.total_standard_cost
                    sc.product.save(update_fields=['cost_price', 'updated_at'])
                    updated_count += 1

            products = Product.objects.filter(is_active=True)
            total_value = 0
            items = []
            for p in products:
                value = p.current_stock * (p.cost_price or 0)
                items.append(CostSettlementItem(
                    settlement=settlement,
                    product=p,
                    stock_quantity=p.current_stock,
                    cost_price=p.cost_price or 0,
                    inventory_value=value,
                ))
                total_value += value

            CostSettlementItem.objects.bulk_create(items)
            settlement.total_inventory_value = total_value
            settlement.save(
                update_fields=['total_inventory_value', 'updated_at'],
            )

        messages.success(
            request,
            f'{period_start}~{period_end} 원가정산 완료 '
            f'(총 재고자산: {total_value:,}원, '
            f'표준원가 갱신: {updated_count}건)',
        )
        return redirect(
            'accounting:settlement_detail', slug=settlement.settlement_number,
        )


class CostSettlementRecalcView(ManagerRequiredMixin, View):
    """기존 정산 건의 표준원가 재계산 + 스냅샷 갱신"""

    def post(self, request, slug):
        settlement = get_object_or_404(
            CostSettlement, settlement_number=slug, is_active=True,
        )
        from apps.production.models import StandardCost

        with transaction.atomic():
            # 표준원가 재계산
            updated_count = 0
            for sc in StandardCost.objects.filter(
                is_current=True, is_active=True,
            ).select_related('product'):
                sc.calculate_material_cost()
                sc.save()
                if sc.product.cost_price != sc.total_standard_cost:
                    sc.product.cost_price = sc.total_standard_cost
                    sc.product.save(update_fields=['cost_price', 'updated_at'])
                    updated_count += 1

            # 정산 항목 스냅샷 갱신
            total_value = 0
            for item in settlement.items.select_related('product').all():
                item.cost_price = item.product.cost_price or 0
                item.stock_quantity = item.product.current_stock
                item.inventory_value = item.stock_quantity * item.cost_price
                item.save(update_fields=[
                    'cost_price', 'stock_quantity',
                    'inventory_value', 'updated_at',
                ])
                total_value += item.inventory_value

            settlement.total_inventory_value = total_value
            settlement.save(
                update_fields=['total_inventory_value', 'updated_at'],
            )

        messages.success(
            request,
            f'{settlement.settlement_number} 원가 재계산 완료 '
            f'(총 재고자산: {total_value:,}원, '
            f'표준원가 갱신: {updated_count}건)',
        )
        return redirect(
            'accounting:settlement_detail', slug=settlement.settlement_number,
        )


# ===== 매출 정산 =====

class SalesSettlementListView(ManagerRequiredMixin, ListView):
    model = SalesSettlement
    template_name = 'accounting/sales_settlement_list.html'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class SalesSettlementDetailView(ManagerRequiredMixin, DetailView):
    model = SalesSettlement
    template_name = 'accounting/sales_settlement_detail.html'
    slug_field = 'settlement_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'commission_bank_account', 'commission_deposit_account',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = (
            self.object.settlement_orders
            .select_related('order__partner', 'order__customer')
            .all()
        )
        # 주문별 입금 상태
        from apps.sales.models import Order
        order_ids = [i.order_id for i in ctx['items']]
        orders = Order.objects.filter(pk__in=order_ids)
        ctx['paid_count'] = orders.filter(is_paid=True).count()
        ctx['unpaid_count'] = orders.filter(is_paid=False).count()
        # 계좌 목록 (출금/입금 공용)
        ctx['bank_accounts'] = BankAccount.objects.filter(
            is_active=True,
        )
        # 최근 전표 (수동 지급 시 연결용)
        ctx['recent_vouchers'] = Voucher.objects.filter(
            is_active=True,
        ).order_by('-voucher_date', '-pk')[:20]
        return ctx


class SalesSettlementPaymentView(ManagerRequiredMixin, View):
    """정산 내 미입금 주문 일괄 입금 처리"""

    def post(self, request, slug):
        settlement = get_object_or_404(
            SalesSettlement, settlement_number=slug, is_active=True,
        )
        from apps.sales.signals import _auto_create_payment

        items = settlement.settlement_orders.select_related(
            'order',
        ).all()
        count = 0
        for item in items:
            order = item.order
            if not order.is_paid:
                try:
                    _auto_create_payment(order)
                    count += 1
                except Exception:
                    pass

        if count:
            messages.success(
                request,
                f'{count}건 입금 처리 완료',
            )
        else:
            messages.info(request, '처리할 미입금 주문이 없습니다.')
        return redirect(
            'accounting:sales_settlement_detail', slug=settlement.settlement_number,
        )


class SalesSettlementCommissionPayView(ManagerRequiredMixin, View):
    """정산 수수료 지급 처리"""

    def post(self, request, slug):
        settlement = get_object_or_404(
            SalesSettlement, settlement_number=slug, is_active=True,
        )
        if settlement.commission_paid:
            messages.warning(request, '이미 수수료 지급 완료된 정산입니다.')
            return redirect(
                'accounting:sales_settlement_detail', slug=settlement.settlement_number,
            )

        bank_id = request.POST.get('commission_bank_account')
        bank = None
        if bank_id:
            bank = BankAccount.objects.filter(
                pk=bank_id, is_active=True,
            ).first()

        deposit_bank_id = request.POST.get('commission_deposit_account')
        deposit_bank = None
        if deposit_bank_id:
            deposit_bank = BankAccount.objects.filter(
                pk=deposit_bank_id, is_active=True,
            ).first()

        commission_total = int(settlement.total_commission)
        if commission_total <= 0:
            messages.info(request, '지급할 수수료가 없습니다.')
            return redirect(
                'accounting:sales_settlement_detail', slug=settlement.settlement_number,
            )

        with transaction.atomic():
            # 정산 내 첫 번째 거래처 기준
            first_item = settlement.settlement_orders.select_related(
                'order__partner',
            ).first()
            partner = first_item.order.partner if first_item and first_item.order.partner else None

            # 수수료 지급 전표 생성
            voucher = Voucher.objects.create(
                voucher_type='PAYMENT',
                voucher_date=date.today(),
                description=(
                    f'수수료 지급 - {settlement.settlement_number}'
                ),
                approval_status='APPROVED',
                created_by=request.user,
            )

            acct_payable = AccountCode.objects.filter(
                code='205',
            ).first()  # 미지급금
            acct_deposit = AccountCode.objects.filter(
                code='103',
            ).first()  # 보통예금

            if not acct_payable or not acct_deposit:
                messages.warning(request, '수수료 전표 생성에 필요한 계정과목이 미등록되어 전표가 생성되지 않았습니다.')
            else:
                # 차변: 미지급금 (부채 감소)
                deposit_desc = f' → {deposit_bank.name}' if deposit_bank else ''
                VoucherLine.objects.create(
                    voucher=voucher,
                    account=acct_payable,
                    debit=commission_total,
                    credit=0,
                    description=(
                        f'{settlement.settlement_number} 수수료{deposit_desc}'
                    ),
                )
                # 대변: 보통예금 (자산 감소)
                VoucherLine.objects.create(
                    voucher=voucher,
                    account=acct_deposit,
                    debit=0,
                    credit=commission_total,
                    description=(
                        f'{bank.name} 출금' if bank
                        else '수수료 지급'
                    ),
                )

            # 출금 기록 (시그널에서 계좌 잔액 자동 차감)
            Payment.objects.create(
                payment_type='DISBURSEMENT',
                partner=partner,
                bank_account=bank,
                voucher=voucher,
                amount=commission_total,
                payment_date=date.today(),
                payment_method='BANK_TRANSFER',
                reference=(
                    f'수수료 {settlement.settlement_number}'
                ),
                created_by=request.user,
            )

            # 미지급금 생성/업데이트
            if first_item and first_item.order.partner:
                ap, created = AccountPayable.objects.get_or_create(
                    partner=first_item.order.partner,
                    notes=f'수수료 {settlement.settlement_number}',
                    defaults={
                        'amount': commission_total,
                        'paid_amount': commission_total,
                        'due_date': date.today(),
                        'status': AccountPayable.Status.PAID,
                        'created_by': request.user,
                    },
                )
                if not created:
                    ap.paid_amount = commission_total
                    ap.status = AccountPayable.Status.PAID
                    ap.save()

            # 입금계좌 잔액 증가 (선택된 경우)
            if deposit_bank:
                BankAccount.objects.filter(pk=deposit_bank.pk).update(
                    balance=F('balance') + commission_total,
                )

            # 정산 수수료 지급 상태 업데이트
            settlement.commission_bank_account = bank
            settlement.commission_deposit_account = deposit_bank
            settlement.commission_paid = True
            settlement.commission_paid_date = date.today()
            settlement.commission_paid_amount = commission_total
            settlement.commission_voucher = voucher
            settlement.save()

        messages.success(
            request,
            f'수수료 {commission_total:,}원 지급 완료',
        )
        return redirect(
            'accounting:sales_settlement_detail', slug=settlement.settlement_number,
        )


class SalesSettlementCommissionManualView(ManagerRequiredMixin, View):
    """수수료 수동 지급완료 처리 (전표/출금 없이 상태만 변경)"""

    def post(self, request, slug):
        settlement = get_object_or_404(
            SalesSettlement, settlement_number=slug, is_active=True,
        )
        if settlement.commission_paid:
            messages.warning(request, '이미 수수료 지급 완료된 정산입니다.')
            return redirect(
                'accounting:sales_settlement_detail', slug=settlement.settlement_number,
            )

        commission_total = int(settlement.total_commission)
        memo = request.POST.get('memo', '')
        voucher_id = request.POST.get('voucher')

        settlement.commission_paid = True
        settlement.commission_paid_date = date.today()
        settlement.commission_paid_amount = commission_total
        settlement.commission_memo = memo
        if voucher_id:
            v = Voucher.objects.filter(pk=voucher_id).first()
            if v:
                settlement.commission_voucher = v
        settlement.save()

        messages.success(
            request,
            f'수수료 {commission_total:,}원 수동 지급완료 처리'
            + (f' ({memo})' if memo else ''),
        )
        return redirect(
            'accounting:sales_settlement_detail', slug=settlement.settlement_number,
        )


class SalesSettlementCreateView(ManagerRequiredMixin, TemplateView):
    """주문 선택 → 매출 정산 생성"""
    template_name = 'accounting/sales_settlement_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # 미정산 + 종결 주문만 표시
        orders = (
            Order.objects.filter(
                is_active=True,
                is_settled=False,
                status='CLOSED',
            )
            .select_related('partner', 'customer')
            .prefetch_related('items__product')
            .order_by('-order_date')
        )
        # 주문별 원가 계산 + 객체에 동적 속성 추가
        order_costs = {}
        for order in orders:
            cost = sum(
                int(i.cost_price or 0) * int(i.quantity)
                for i in order.items.all()
            )
            order_costs[order.pk] = cost
            order.cost_total = cost
        # 수수료율 맵 (partner_id → total rate) — CommissionRate 항목 합산
        from apps.sales.models import Partner
        rates = {}
        for p in Partner.objects.filter(is_active=True):
            r = float(p.total_commission_rate)
            if r > 0:
                rates[p.pk] = r
        ctx['orders'] = orders
        ctx['order_costs'] = {str(k): v for k, v in order_costs.items()}
        ctx['commission_rates'] = {str(k): v for k, v in rates.items()}
        ctx['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        return ctx

    def post(self, request, *args, **kwargs):
        order_ids = request.POST.getlist('orders')
        if not order_ids:
            messages.error(request, '정산할 주문을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        description = request.POST.get('description', '')

        from apps.sales.models import Partner
        rates_map = {}
        for p in Partner.objects.filter(is_active=True):
            r = p.total_commission_rate
            if r > 0:
                rates_map[p.pk] = r

        with transaction.atomic():
            settlement = SalesSettlement.objects.create(
                settlement_date=date.today(),
                description=description,
                created_by=request.user,
            )

            total_revenue = 0
            total_cost = 0
            total_tax = 0
            total_shipping = 0
            total_platform_commission = 0
            total_commission = 0
            total_profit = 0
            total_cost_variance = 0

            orders = Order.objects.filter(
                pk__in=order_ids, is_settled=False,
            ).select_related('partner')

            for order in orders:
                revenue = int(order.total_amount)
                tax = int(order.tax_total)
                shipping = int(order.shipping_cost or 0)
                # 원가 계산: 주문시점 + 정산시점(현재)
                items = order.items.select_related('product').all()
                cost = sum(
                    int(i.cost_price or 0) * int(i.quantity)
                    for i in items
                )
                current_cost = sum(
                    int(i.product.cost_price or 0) * int(i.quantity)
                    for i in items
                )
                cost_variance = current_cost - cost
                cost_variance_rate = (
                    round(cost_variance / cost * 100, 2) if cost else 0
                )
                # 수수료: 정률(%) or 정액(원)
                platform_comm = int(order.platform_commission or 0)
                comm_type = request.POST.get(f'comm_type_{order.pk}', 'rate')
                rate_val = request.POST.get(f'rate_{order.pk}', '')
                if comm_type == 'fixed' and rate_val:
                    # 정액: 입력값 그대로 수수료
                    try:
                        settle_commission = round(Decimal(rate_val))
                    except Exception:
                        settle_commission = 0
                    comm_rate = Decimal('0')
                else:
                    # 정률: 기존 로직
                    if rate_val:
                        try:
                            comm_rate = Decimal(rate_val)
                        except Exception:
                            comm_rate = Decimal('0')
                    else:
                        comm_rate = rates_map.get(
                            order.partner_id, Decimal('0'),
                        )
                    net_revenue = max(revenue - platform_comm, 0)
                    comm_base_type = request.POST.get('commission_base', 'revenue')
                    if comm_base_type == 'profit':
                        comm_base = max(net_revenue - cost - shipping, 0)
                    else:
                        comm_base = net_revenue
                    settle_commission = round(comm_base * comm_rate / 100)
                profit = revenue - cost - shipping - platform_comm - settle_commission

                SalesSettlementOrder.objects.create(
                    settlement=settlement,
                    order=order,
                    revenue=revenue,
                    cost=cost,
                    current_cost=current_cost,
                    cost_variance=cost_variance,
                    cost_variance_rate=cost_variance_rate,
                    tax=tax,
                    shipping=shipping,
                    platform_commission=platform_comm,
                    commission_rate=comm_rate,
                    commission=settle_commission,
                    profit=profit,
                    created_by=request.user,
                )

                total_revenue += revenue
                total_cost += cost
                total_tax += tax
                total_shipping += shipping
                total_platform_commission += platform_comm
                total_commission += settle_commission
                total_profit += profit
                total_cost_variance += cost_variance

            settlement.total_revenue = total_revenue
            settlement.total_cost = total_cost
            settlement.total_tax = total_tax
            settlement.total_shipping = total_shipping
            settlement.total_platform_commission = total_platform_commission
            settlement.total_commission = total_commission
            settlement.total_profit = total_profit
            settlement.total_cost_variance = total_cost_variance
            settlement.save(update_fields=[
                'total_revenue', 'total_cost', 'total_tax',
                'total_shipping', 'total_platform_commission',
                'total_commission', 'total_profit',
                'total_cost_variance', 'updated_at',
            ])

            # 주문 정산완료 마킹
            orders.update(is_settled=True)

            # 정산 계좌 처리
            first_order = orders.first()
            partner = first_order.partner if first_order else None

            # 수수료 출금 (주문 배송완료 시그널에서 이미 출금된 건 제외)
            comm_bank_id = request.POST.get('commission_bank')
            if comm_bank_id and total_commission > 0 and partner:
                # 시그널에서 이미 처리된 수수료 출금 금액 합산
                already_paid = Payment.objects.filter(
                    partner=partner,
                    payment_type='DISBURSEMENT',
                    is_active=True,
                    reference__in=[
                        f'주문 {o.order_number} 수수료'
                        for o in orders
                    ],
                ).aggregate(total=Sum('amount'))['total'] or 0
                remaining_commission = total_commission - int(already_paid)

                comm_bank = BankAccount.objects.filter(
                    pk=comm_bank_id, is_active=True,
                ).first()
                if comm_bank and remaining_commission > 0:
                    Payment.objects.create(
                        payment_type='DISBURSEMENT',
                        partner=partner,
                        bank_account=comm_bank,
                        amount=remaining_commission,
                        payment_date=date.today(),
                        payment_method='BANK_TRANSFER',
                        reference=f'정산 {settlement.settlement_number} 수수료',
                        notes=f'매출정산 수수료 {remaining_commission:,}원',
                        created_by=request.user,
                    )
                settlement.commission_bank_account = comm_bank
                settlement.commission_paid = True
                settlement.commission_paid_date = date.today()
                settlement.commission_paid_amount = total_commission
                settlement.save()

            # 시그널에서 이미 수수료 전액 출금된 경우 자동 정산완료
            if (
                total_commission > 0
                and not settlement.commission_paid
                and partner
            ):
                already_paid = Payment.objects.filter(
                    partner=partner,
                    payment_type='DISBURSEMENT',
                    is_active=True,
                    reference__in=[
                        f'주문 {o.order_number} 수수료'
                        for o in orders
                    ],
                ).aggregate(total=Sum('amount'))['total'] or 0
                if int(already_paid) >= total_commission:
                    settlement.commission_paid = True
                    settlement.commission_paid_date = date.today()
                    settlement.commission_paid_amount = total_commission
                    settlement.save(update_fields=[
                        'commission_paid', 'commission_paid_date',
                        'commission_paid_amount', 'updated_at',
                    ])

            # 정산금 입금 (매출 - 수수료)
            settle_bank_id = request.POST.get('settlement_bank')
            if settle_bank_id and partner:
                settle_bank = BankAccount.objects.filter(
                    pk=settle_bank_id, is_active=True,
                ).first()
                if settle_bank:
                    net_amount = total_revenue + total_tax - total_commission
                    if net_amount > 0:
                        Payment.objects.create(
                            payment_type='RECEIPT',
                            partner=partner,
                            bank_account=settle_bank,
                            amount=net_amount,
                            payment_date=date.today(),
                            payment_method='BANK_TRANSFER',
                            reference=f'정산 {settlement.settlement_number} 입금',
                            notes=f'매출정산 입금 {net_amount:,}원 (매출 {total_revenue + total_tax:,} - 수수료 {total_commission:,})',
                            created_by=request.user,
                        )

        messages.success(
            request,
            f'매출정산 {settlement.settlement_number} 완료 '
            f'({orders.count()}건, 이익: {total_profit:,}원)',
        )
        return redirect(
            'accounting:sales_settlement_detail',
            slug=settlement.settlement_number,
        )


# === 은행 대사 ===
class BankReconciliationView(ManagerRequiredMixin, TemplateView):
    """은행 대사 — 시스템 잔액 vs 실제 통장 잔액 비교"""
    template_name = 'accounting/bank_reconciliation.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        accounts = BankAccount.objects.filter(is_active=True)
        recon_data = []
        total_system = 0
        total_actual = 0
        for acct in accounts:
            system_balance = int(acct.balance)
            total_system += system_balance
            # 입출금 건수
            receipt_count = Payment.objects.filter(
                bank_account=acct, payment_type='RECEIPT', is_active=True,
            ).count()
            disbursement_count = Payment.objects.filter(
                bank_account=acct, payment_type='DISBURSEMENT', is_active=True,
            ).count()
            recon_data.append({
                'pk': acct.pk,
                'name': acct.name,
                'bank_name': acct.bank,
                'account_number': acct.account_number,
                'system_balance': system_balance,
                'receipt_count': receipt_count,
                'disbursement_count': disbursement_count,
            })
        ctx['recon_data'] = recon_data
        ctx['total_system'] = total_system
        ctx['recon_date'] = date.today().isoformat()
        return ctx


# === 계정별 원장 ===
class AccountLedgerView(ManagerRequiredMixin, TemplateView):
    """계정별 원장 조회 — 특정 계정과목의 전표 라인을 시간순 표시 + 잔액 누적"""
    template_name = 'accounting/account_ledger.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        account_id = self.request.GET.get('account')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')

        ctx['accounts'] = AccountCode.objects.filter(is_active=True).order_by('code')
        ctx['selected_account'] = account_id
        ctx['date_from'] = date_from
        ctx['date_to'] = date_to

        if not account_id:
            ctx['lines'] = []
            return ctx

        account = get_object_or_404(AccountCode, pk=account_id)
        ctx['account'] = account

        qs = VoucherLine.objects.filter(
            account=account, is_active=True,
            voucher__is_active=True,
            voucher__approval_status='APPROVED',
        ).select_related('voucher').order_by('voucher__voucher_date', 'pk')

        if date_from:
            qs = qs.filter(voucher__voucher_date__gte=date_from)
        if date_to:
            qs = qs.filter(voucher__voucher_date__lte=date_to)

        # 잔액 누적 계산
        lines = []
        running_balance = 0
        total_debit = 0
        total_credit = 0
        for line in qs:
            debit = int(line.debit)
            credit = int(line.credit)
            # 자산/비용: 차변 증가, 대변 감소
            # 부채/자본/수익: 대변 증가, 차변 감소
            if account.account_type in ('ASSET', 'EXPENSE'):
                running_balance += debit - credit
            else:
                running_balance += credit - debit
            total_debit += debit
            total_credit += credit
            lines.append({
                'date': line.voucher.voucher_date,
                'voucher_number': line.voucher.voucher_number,
                'voucher_pk': line.voucher.pk,
                'voucher_type': line.voucher.get_voucher_type_display(),
                'description': line.description or line.voucher.description,
                'debit': debit,
                'credit': credit,
                'balance': running_balance,
            })

        ctx['lines'] = lines
        ctx['total_debit'] = total_debit
        ctx['total_credit'] = total_credit
        ctx['closing_balance'] = running_balance
        return ctx


# === 시산표 ===
class TrialBalanceView(ManagerRequiredMixin, TemplateView):
    """시산표 — 모든 계정과목의 차변/대변 합계 + 잔액"""
    template_name = 'accounting/trial_balance.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        ctx['date_from'] = date_from
        ctx['date_to'] = date_to

        qs = VoucherLine.objects.filter(
            is_active=True, voucher__is_active=True,
            voucher__approval_status='APPROVED',
        )
        if date_from:
            qs = qs.filter(voucher__voucher_date__gte=date_from)
        if date_to:
            qs = qs.filter(voucher__voucher_date__lte=date_to)

        account_totals = qs.values(
            'account__pk', 'account__code', 'account__name', 'account__account_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('account__code')

        rows = []
        grand_debit = 0
        grand_credit = 0
        for row in account_totals:
            debit = int(row['total_debit'] or 0)
            credit = int(row['total_credit'] or 0)
            acct_type = row['account__account_type']
            if acct_type in ('ASSET', 'EXPENSE'):
                balance = debit - credit
            else:
                balance = credit - debit
            grand_debit += debit
            grand_credit += credit
            rows.append({
                'pk': row['account__pk'],
                'code': row['account__code'],
                'name': row['account__name'],
                'account_type': acct_type,
                'debit': debit,
                'credit': credit,
                'balance': balance,
            })

        ctx['rows'] = rows
        ctx['grand_debit'] = grand_debit
        ctx['grand_credit'] = grand_credit
        ctx['is_balanced'] = grand_debit == grand_credit
        return ctx


# === 예산 관리 ===
class BudgetListView(ManagerRequiredMixin, TemplateView):
    """예산 vs 실적 대비표"""
    template_name = 'accounting/budget_list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        month = safe_int(self.request.GET.get('month'), date.today().month)

        ctx['year'] = year
        ctx['month'] = month
        ctx['years'] = range(date.today().year - 2, date.today().year + 2)
        ctx['months'] = range(1, 13)

        budgets = Budget.objects.filter(
            year=year, month=month, is_active=True,
        ).select_related('account').order_by('account__code')

        # 해당 월의 VoucherLine 실적을 한 번에 배치 조회
        period_start = date(year, month, 1)
        if month == 12:
            period_end = date(year + 1, 1, 1)
        else:
            period_end = date(year, month + 1, 1)

        actuals_qs = VoucherLine.objects.filter(
            is_active=True,
            voucher__is_active=True,
            voucher__voucher_date__gte=period_start,
            voucher__voucher_date__lt=period_end,
        ).values('account_id').annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        )
        actuals_map = {
            row['account_id']: (
                int(row['total_debit'] or 0),
                int(row['total_credit'] or 0),
            )
            for row in actuals_qs
        }

        rows = []
        total_budget = 0
        total_actual = 0
        for b in budgets:
            debit, credit = actuals_map.get(b.account_id, (0, 0))
            if b.account.account_type == 'EXPENSE':
                actual = debit
            elif b.account.account_type == 'REVENUE':
                actual = credit
            else:
                actual = debit - credit
            budget_amt = int(b.budget_amount)
            variance = budget_amt - actual
            rate = (
                round(actual / budget_amt * 100, 1)
                if budget_amt > 0 else 0
            )
            total_budget += budget_amt
            total_actual += actual
            rows.append({
                'pk': b.pk,
                'code': b.account.code,
                'name': b.account.name,
                'account_type': b.account.get_account_type_display(),
                'budget': budget_amt,
                'actual': actual,
                'variance': variance,
                'rate': rate,
            })

        ctx['rows'] = rows
        ctx['total_budget'] = total_budget
        ctx['total_actual'] = total_actual
        ctx['total_variance'] = total_budget - total_actual
        return ctx


class BudgetCreateView(ManagerRequiredMixin, View):
    """예산 등록 (POST)"""
    def get(self, request):
        ctx = {
            'accounts': AccountCode.objects.filter(
                is_active=True, account_type__in=['EXPENSE', 'REVENUE'],
            ).order_by('code'),
            'year': date.today().year,
            'month': date.today().month,
        }
        from django.shortcuts import render
        return render(request, 'accounting/budget_form.html', ctx)

    def post(self, request):
        account_id = request.POST.get('account')
        year = safe_int(request.POST.get('year'), date.today().year)
        month = safe_int(request.POST.get('month'), date.today().month)
        amount = safe_int(request.POST.get('budget_amount'), 0)
        desc = request.POST.get('description', '')

        if not account_id or amount <= 0:
            messages.error(request, '계정과목과 예산액을 입력해주세요.')
            return redirect('accounting:budget_create')

        account = get_object_or_404(AccountCode, pk=account_id)

        budget, created = Budget.objects.update_or_create(
            account=account, year=year, month=month,
            defaults={
                'budget_amount': amount,
                'description': desc,
                'created_by': request.user,
                'is_active': True,
            },
        )
        action = '등록' if created else '수정'
        messages.success(
            request,
            f'{year}년 {month}월 {account.name} 예산 {amount:,}원 {action}',
        )
        return redirect(
            f'/accounting/budget/?year={year}&month={month}',
        )


# === AR/AP Aging 분석 ===

class ARAgingView(ManagerRequiredMixin, TemplateView):
    """미수금 Aging 분석 (30/60/90/120일 이상 구간별)"""
    template_name = 'accounting/ar_aging.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()

        # 자동 연체 전환
        AccountReceivable.objects.filter(
            is_active=True,
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL'],
        ).update(status='OVERDUE')

        # 미완납 AR만 대상
        ar_qs = AccountReceivable.objects.filter(
            is_active=True,
            status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
        ).select_related('partner', 'order')

        buckets = {
            'current': {'label': '미도래', 'items': [], 'total': 0},
            'over_30': {'label': '30일 이내', 'items': [], 'total': 0},
            'over_60': {'label': '31~60일', 'items': [], 'total': 0},
            'over_90': {'label': '61~90일', 'items': [], 'total': 0},
            'over_120': {'label': '91~120일', 'items': [], 'total': 0},
            'over_120_plus': {'label': '120일 초과', 'items': [], 'total': 0},
        }

        grand_total = 0
        for ar in ar_qs:
            remaining = int(ar.amount) - int(ar.paid_amount)
            if remaining <= 0:
                continue
            days_overdue = (today - ar.due_date).days
            grand_total += remaining

            if days_overdue <= 0:
                bucket_key = 'current'
            elif days_overdue <= 30:
                bucket_key = 'over_30'
            elif days_overdue <= 60:
                bucket_key = 'over_60'
            elif days_overdue <= 90:
                bucket_key = 'over_90'
            elif days_overdue <= 120:
                bucket_key = 'over_120'
            else:
                bucket_key = 'over_120_plus'

            buckets[bucket_key]['items'].append({
                'ar': ar,
                'remaining': remaining,
                'days_overdue': max(days_overdue, 0),
            })
            buckets[bucket_key]['total'] += remaining

        ctx['buckets'] = buckets
        ctx['grand_total'] = grand_total
        ctx['today'] = today

        # 거래처별 집계
        partner_summary = {}
        for bucket_key, bucket_data in buckets.items():
            for item in bucket_data['items']:
                partner_name = item['ar'].partner.name
                if partner_name not in partner_summary:
                    partner_summary[partner_name] = {
                        'name': partner_name,
                        'current': 0, 'over_30': 0, 'over_60': 0,
                        'over_90': 0, 'over_120': 0, 'over_120_plus': 0,
                        'total': 0,
                    }
                partner_summary[partner_name][bucket_key] += item['remaining']
                partner_summary[partner_name]['total'] += item['remaining']

        ctx['partner_summary'] = sorted(
            partner_summary.values(), key=lambda x: -x['total'],
        )
        return ctx


class APAgingView(ManagerRequiredMixin, TemplateView):
    """미지급금 Aging 분석 (30/60/90/120일 이상 구간별)"""
    template_name = 'accounting/ap_aging.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()

        # 자동 연체 전환
        AccountPayable.objects.filter(
            is_active=True,
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL'],
        ).update(status='OVERDUE')

        # 미완납 AP만 대상
        ap_qs = AccountPayable.objects.filter(
            is_active=True,
            status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
        ).select_related('partner')

        buckets = {
            'current': {'label': '미도래', 'items': [], 'total': 0},
            'over_30': {'label': '30일 이내', 'items': [], 'total': 0},
            'over_60': {'label': '31~60일', 'items': [], 'total': 0},
            'over_90': {'label': '61~90일', 'items': [], 'total': 0},
            'over_120': {'label': '91~120일', 'items': [], 'total': 0},
            'over_120_plus': {'label': '120일 초과', 'items': [], 'total': 0},
        }

        grand_total = 0
        for ap in ap_qs:
            remaining = int(ap.amount) - int(ap.paid_amount)
            if remaining <= 0:
                continue
            days_overdue = (today - ap.due_date).days
            grand_total += remaining

            if days_overdue <= 0:
                bucket_key = 'current'
            elif days_overdue <= 30:
                bucket_key = 'over_30'
            elif days_overdue <= 60:
                bucket_key = 'over_60'
            elif days_overdue <= 90:
                bucket_key = 'over_90'
            elif days_overdue <= 120:
                bucket_key = 'over_120'
            else:
                bucket_key = 'over_120_plus'

            buckets[bucket_key]['items'].append({
                'ap': ap,
                'remaining': remaining,
                'days_overdue': max(days_overdue, 0),
            })
            buckets[bucket_key]['total'] += remaining

        ctx['buckets'] = buckets
        ctx['grand_total'] = grand_total
        ctx['today'] = today

        # 거래처별 집계
        partner_summary = {}
        for bucket_key, bucket_data in buckets.items():
            for item in bucket_data['items']:
                partner_name = item['ap'].partner.name
                if partner_name not in partner_summary:
                    partner_summary[partner_name] = {
                        'name': partner_name,
                        'current': 0, 'over_30': 0, 'over_60': 0,
                        'over_90': 0, 'over_120': 0, 'over_120_plus': 0,
                        'total': 0,
                    }
                partner_summary[partner_name][bucket_key] += item['remaining']
                partner_summary[partner_name]['total'] += item['remaining']

        ctx['partner_summary'] = sorted(
            partner_summary.values(), key=lambda x: -x['total'],
        )
        return ctx


# === 결산 마감 ===

class ClosingPeriodListView(ManagerRequiredMixin, TemplateView):
    """결산 마감 현황"""
    template_name = 'accounting/closing_list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        ctx['year'] = year
        ctx['years'] = range(date.today().year - 2, date.today().year + 2)

        # 1~12월 결산 현황
        existing = {
            cp.month: cp
            for cp in ClosingPeriod.objects.filter(
                year=year, is_active=True,
            ).select_related('closed_by')
        }

        periods = []
        for m in range(1, 13):
            cp = existing.get(m)
            voucher_count = Voucher.objects.filter(
                voucher_date__year=year,
                voucher_date__month=m,
                is_active=True,
            ).count()
            periods.append({
                'month': m,
                'closing': cp,
                'is_closed': cp.is_closed if cp else False,
                'closed_at': cp.closed_at if cp else None,
                'closed_by': cp.closed_by if cp else None,
                'pk': cp.pk if cp else None,
                'voucher_count': voucher_count,
            })

        ctx['periods'] = periods
        return ctx


class ClosingPeriodCloseView(AdminRequiredMixin, View):
    """결산 마감/해제 실행 (POST)"""

    def post(self, request, year, month):
        from django.utils import timezone

        if month < 1 or month > 12:
            messages.error(request, '잘못된 월입니다.')
            return redirect(f'/accounting/closing/?year={year}')

        cp, created = ClosingPeriod.objects.get_or_create(
            year=year, month=month,
            defaults={
                'created_by': request.user,
                'is_active': True,
            },
        )

        if cp.is_closed:
            # 마감 해제
            cp.is_closed = False
            cp.closed_at = None
            cp.closed_by = None
            cp.save()
            messages.success(request, f'{year}년 {month}월 결산 마감이 해제되었습니다.')
        else:
            # 마감 전 미승인 전표 검증
            pending_vouchers = Voucher.objects.filter(
                is_active=True,
                voucher_date__year=year,
                voucher_date__month=month,
                approval_status__in=['DRAFT', 'SUBMITTED'],
            ).exists()
            if pending_vouchers:
                messages.error(request, '미승인 전표가 있어 마감할 수 없습니다. 먼저 전표를 승인하거나 삭제하세요.')
                return redirect(f'/accounting/closing/?year={year}')

            # 마감 실행
            cp.is_closed = True
            cp.closed_at = timezone.now()
            cp.closed_by = request.user
            cp.save()
            messages.success(request, f'{year}년 {month}월 결산이 마감되었습니다.')

        return redirect(f'/accounting/closing/?year={year}')


# === 예산 보고서 (연도별 월간 차트) ===

class BudgetReportView(ManagerRequiredMixin, TemplateView):
    """예산 vs 실적 보고서 (연도별 월간 차트)"""
    template_name = 'accounting/budget_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        ctx['year'] = year
        ctx['years'] = range(date.today().year - 2, date.today().year + 2)

        # 월별 예산 합계 / 실적 합계
        monthly_budget = []
        monthly_actual = []
        months_label = []

        for m in range(1, 13):
            months_label.append(f'{m}월')

            # 해당 월 예산 합계
            budget_total = Budget.objects.filter(
                year=year, month=m, is_active=True,
            ).aggregate(total=Sum('budget_amount'))['total'] or 0
            monthly_budget.append(int(budget_total))

            # 해당 월 실적 합계 (전표 기반)
            period_start = date(year, m, 1)
            if m == 12:
                period_end = date(year + 1, 1, 1)
            else:
                period_end = date(year, m + 1, 1)

            actuals_qs = VoucherLine.objects.filter(
                is_active=True,
                voucher__is_active=True,
                voucher__voucher_date__gte=period_start,
                voucher__voucher_date__lt=period_end,
            ).aggregate(
                total_debit=Sum('debit'),
                total_credit=Sum('credit'),
            )
            debit = int(actuals_qs['total_debit'] or 0)
            credit = int(actuals_qs['total_credit'] or 0)
            # 비용 위주로 집행 실적 표시 (차변 합계)
            monthly_actual.append(debit)

        ctx['months_label'] = months_label
        ctx['monthly_budget'] = monthly_budget
        ctx['monthly_actual'] = monthly_actual

        # 계정별 연간 예산 vs 실적 상세
        budgets = Budget.objects.filter(
            year=year, is_active=True,
        ).select_related('account').order_by('account__code', 'month')

        # 계정별 연간 합산
        account_map = {}
        for b in budgets:
            key = b.account_id
            if key not in account_map:
                account_map[key] = {
                    'code': b.account.code,
                    'name': b.account.name,
                    'account_type': b.account.get_account_type_display(),
                    'account_type_raw': b.account.account_type,
                    'budget': 0,
                    'actual': 0,
                }
            account_map[key]['budget'] += int(b.budget_amount)

        # 연간 실적 배치 조회
        if account_map:
            period_start = date(year, 1, 1)
            period_end = date(year + 1, 1, 1)
            actuals_qs = VoucherLine.objects.filter(
                is_active=True,
                voucher__is_active=True,
                voucher__voucher_date__gte=period_start,
                voucher__voucher_date__lt=period_end,
                account_id__in=account_map.keys(),
            ).values('account_id').annotate(
                total_debit=Sum('debit'),
                total_credit=Sum('credit'),
            )
            for row in actuals_qs:
                aid = row['account_id']
                if aid in account_map:
                    d = int(row['total_debit'] or 0)
                    c = int(row['total_credit'] or 0)
                    atype = account_map[aid]['account_type_raw']
                    if atype == 'EXPENSE':
                        account_map[aid]['actual'] = d
                    elif atype == 'REVENUE':
                        account_map[aid]['actual'] = c
                    else:
                        account_map[aid]['actual'] = d - c

        rows = []
        total_budget = 0
        total_actual = 0
        for data in sorted(account_map.values(), key=lambda x: x['code']):
            variance = data['budget'] - data['actual']
            rate = round(data['actual'] / data['budget'] * 100, 1) if data['budget'] > 0 else 0
            total_budget += data['budget']
            total_actual += data['actual']
            rows.append({
                'code': data['code'],
                'name': data['name'],
                'account_type': data['account_type'],
                'budget': data['budget'],
                'actual': data['actual'],
                'variance': variance,
                'rate': rate,
            })

        ctx['rows'] = rows
        ctx['total_budget'] = total_budget
        ctx['total_actual'] = total_actual
        ctx['total_variance'] = total_budget - total_actual
        ctx['total_rate'] = round(total_actual / total_budget * 100, 1) if total_budget > 0 else 0

        return ctx


# ── 통화 관리 ──────────────────────────────────────────────


class CurrencyListView(ManagerRequiredMixin, ListView):
    model = Currency
    template_name = 'accounting/currency_list.html'
    context_object_name = 'currencies'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class CurrencyCreateView(ManagerRequiredMixin, CreateView):
    model = Currency
    form_class = CurrencyForm
    template_name = 'accounting/currency_form.html'
    success_url = reverse_lazy('accounting:currency_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CurrencyUpdateView(ManagerRequiredMixin, UpdateView):
    model = Currency
    form_class = CurrencyForm
    template_name = 'accounting/currency_form.html'
    success_url = reverse_lazy('accounting:currency_list')


# ── 환율 관리 ──────────────────────────────────────────────


class ExchangeRateListView(ManagerRequiredMixin, ListView):
    model = ExchangeRate
    template_name = 'accounting/exchange_rate_list.html'
    context_object_name = 'rates'
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('currency')
        currency_id = self.request.GET.get('currency')
        if currency_id:
            qs = qs.filter(currency_id=currency_id)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['currencies'] = Currency.objects.filter(
            is_active=True,
        )
        return ctx


class ExchangeRateCreateView(ManagerRequiredMixin, CreateView):
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'accounting/exchange_rate_form.html'
    success_url = reverse_lazy('accounting:exchangerate_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


# === 전자세금계산서 ===
class ElectronicInvoiceIssueView(ManagerRequiredMixin, View):
    """단건 전자세금계산서 발행"""

    def post(self, request, slug):
        invoice = get_object_or_404(TaxInvoice, invoice_number=slug, is_active=True)

        if invoice.electronic_status not in ('NONE', 'DRAFT', 'REJECTED'):
            messages.error(
                request,
                f'현재 상태({invoice.get_electronic_status_display()})에서는 발행할 수 없습니다.',
            )
            return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)

        from .nts_client import NTSClient, NTSAPIError

        try:
            client = NTSClient()
            result = client.issue(invoice)

            with transaction.atomic():
                invoice.electronic_status = TaxInvoice.ElectronicStatus.ISSUED
                invoice.issue_id = result.get('issue_id', '')
                invoice.nts_confirmation_number = result.get('confirmation_number', '')
                invoice.sent_at = timezone.now()
                invoice.nts_response = result.get('raw_response', {})
                invoice.save(update_fields=[
                    'electronic_status', 'issue_id', 'nts_confirmation_number',
                    'sent_at', 'nts_response', 'updated_at',
                ])

            messages.success(request, f'전자세금계산서 발행 완료 ({invoice.invoice_number})')
        except NTSAPIError as e:
            messages.error(request, f'전자세금계산서 발행 실패: {e}')

        return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)


class ElectronicInvoiceBatchIssueView(ManagerRequiredMixin, View):
    """선택 건 일괄 전자발행"""

    def post(self, request):
        invoice_ids = request.POST.getlist('invoice_ids')
        if not invoice_ids:
            messages.warning(request, '발행할 세금계산서를 선택해주세요.')
            return redirect('accounting:taxinvoice_list')

        invoices = TaxInvoice.objects.filter(
            pk__in=invoice_ids,
            is_active=True,
            electronic_status__in=['NONE', 'DRAFT', 'REJECTED'],
        )

        if not invoices.exists():
            messages.warning(request, '발행 가능한 세금계산서가 없습니다.')
            return redirect('accounting:taxinvoice_list')

        from .nts_client import NTSClient, NTSAPIError

        client = NTSClient()
        success_count = 0
        fail_count = 0

        for invoice in invoices:
            try:
                result = client.issue(invoice)
                with transaction.atomic():
                    invoice.electronic_status = TaxInvoice.ElectronicStatus.ISSUED
                    invoice.issue_id = result.get('issue_id', '')
                    invoice.nts_confirmation_number = result.get('confirmation_number', '')
                    invoice.sent_at = timezone.now()
                    invoice.nts_response = result.get('raw_response', {})
                    invoice.save(update_fields=[
                        'electronic_status', 'issue_id', 'nts_confirmation_number',
                        'sent_at', 'nts_response', 'updated_at',
                    ])
                success_count += 1
            except NTSAPIError:
                fail_count += 1

        if success_count:
            messages.success(request, f'{success_count}건 전자발행 완료')
        if fail_count:
            messages.error(request, f'{fail_count}건 전자발행 실패')

        return redirect('accounting:taxinvoice_list')


class ElectronicInvoiceStatusView(ManagerRequiredMixin, View):
    """국세청 전송상태 조회"""

    def get(self, request, slug):
        invoice = get_object_or_404(TaxInvoice, invoice_number=slug, is_active=True)

        if not invoice.issue_id:
            messages.warning(request, '전자발행되지 않은 세금계산서입니다.')
            return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)

        from .nts_client import NTSClient, NTSAPIError

        try:
            client = NTSClient()
            result = client.query(invoice)

            status_value = result.get('status', '')
            if status_value and hasattr(TaxInvoice.ElectronicStatus, status_value):
                with transaction.atomic():
                    invoice.electronic_status = status_value
                    invoice.nts_confirmation_number = result.get(
                        'confirmation_number', invoice.nts_confirmation_number,
                    )
                    invoice.nts_response = result
                    invoice.save(update_fields=[
                        'electronic_status', 'nts_confirmation_number',
                        'nts_response', 'updated_at',
                    ])

            messages.success(
                request,
                f'상태 조회 완료: {invoice.get_electronic_status_display()}',
            )
        except NTSAPIError as e:
            messages.error(request, f'상태 조회 실패: {e}')

        return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)


class ElectronicInvoiceCancelView(ManagerRequiredMixin, View):
    """전자세금계산서 취소"""

    def post(self, request, slug):
        invoice = get_object_or_404(TaxInvoice, invoice_number=slug, is_active=True)
        reason = request.POST.get('cancel_reason', '')

        if invoice.electronic_status not in ('ISSUED', 'SENT', 'ACCEPTED'):
            messages.error(
                request,
                f'현재 상태({invoice.get_electronic_status_display()})에서는 취소할 수 없습니다.',
            )
            return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)

        from .nts_client import NTSClient, NTSAPIError

        try:
            client = NTSClient()
            result = client.cancel(invoice, reason=reason)

            with transaction.atomic():
                invoice.electronic_status = TaxInvoice.ElectronicStatus.CANCELLED
                invoice.nts_response = result.get('raw_response', {})
                invoice.save(update_fields=[
                    'electronic_status', 'nts_response', 'updated_at',
                ])

            messages.success(request, f'전자세금계산서 취소 완료 ({invoice.invoice_number})')
        except NTSAPIError as e:
            messages.error(request, f'전자세금계산서 취소 실패: {e}')

        return redirect('accounting:taxinvoice_detail', slug=invoice.invoice_number)


# === 재무제표 ===

class BalanceSheetView(ManagerRequiredMixin, TemplateView):
    """대차대조표 — 기준일 기준 자산/부채/자본 잔액"""
    template_name = 'accounting/balance_sheet.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        as_of = self.request.GET.get('as_of', date.today().isoformat())
        ctx['as_of'] = as_of

        qs = VoucherLine.objects.filter(
            is_active=True, voucher__is_active=True,
            voucher__approval_status='APPROVED',
            voucher__voucher_date__lte=as_of,
        )

        account_totals = qs.values(
            'account__pk', 'account__code', 'account__name', 'account__account_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('account__code')

        assets = []
        liabilities = []
        equity = []
        total_assets = 0
        total_liabilities = 0
        total_equity = 0

        for row in account_totals:
            debit = int(row['total_debit'] or 0)
            credit = int(row['total_credit'] or 0)
            acct_type = row['account__account_type']

            if acct_type == 'ASSET':
                balance = debit - credit
                assets.append({
                    'code': row['account__code'],
                    'name': row['account__name'],
                    'balance': balance,
                })
                total_assets += balance
            elif acct_type == 'LIABILITY':
                balance = credit - debit
                liabilities.append({
                    'code': row['account__code'],
                    'name': row['account__name'],
                    'balance': balance,
                })
                total_liabilities += balance
            elif acct_type == 'EQUITY':
                balance = credit - debit
                equity.append({
                    'code': row['account__code'],
                    'name': row['account__name'],
                    'balance': balance,
                })
                total_equity += balance

        ctx['assets'] = assets
        ctx['liabilities'] = liabilities
        ctx['equity'] = equity
        ctx['total_assets'] = total_assets
        ctx['total_liabilities'] = total_liabilities
        ctx['total_equity'] = total_equity
        ctx['total_liabilities_equity'] = total_liabilities + total_equity
        ctx['is_balanced'] = total_assets == (total_liabilities + total_equity)
        return ctx


class BalanceSheetExcelView(ManagerRequiredMixin, TemplateView):
    """대차대조표 Excel 다운로드"""
    template_name = 'accounting/balance_sheet.html'

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font

        as_of = request.GET.get('as_of', date.today().isoformat())

        qs = VoucherLine.objects.filter(
            is_active=True, voucher__is_active=True,
            voucher__approval_status='APPROVED',
            voucher__voucher_date__lte=as_of,
        )

        account_totals = qs.values(
            'account__code', 'account__name', 'account__account_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('account__code')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '대차대조표'
        bold = Font(bold=True)

        ws.cell(row=1, column=1, value=f'대차대조표 (기준일: {as_of})').font = bold
        headers = ['구분', '계정코드', '계정명', '잔액']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = bold

        row_num = 4
        for row in account_totals:
            debit = int(row['total_debit'] or 0)
            credit = int(row['total_credit'] or 0)
            acct_type = row['account__account_type']
            if acct_type not in ('ASSET', 'LIABILITY', 'EQUITY'):
                continue
            if acct_type == 'ASSET':
                balance = debit - credit
                label = '자산'
            elif acct_type == 'LIABILITY':
                balance = credit - debit
                label = '부채'
            else:
                balance = credit - debit
                label = '자본'
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=row['account__code'])
            ws.cell(row=row_num, column=3, value=row['account__name'])
            ws.cell(row=row_num, column=4, value=balance)
            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=balance_sheet_{as_of}.xlsx'
        wb.save(response)
        return response


class CashFlowView(ManagerRequiredMixin, TemplateView):
    """현금흐름표 — 영업/투자/재무 활동별 현금 흐름.

    **method 파라미터:**
    - `indirect` (기본) — 간접법: 순이익 + 비현금비용 ± AR/AP 변동
    - `direct` — 직접법: 실제 Payment(RECEIPT/DISBURSEMENT) 집계

    AR/AP 변동은 기간 경계잔액 차이로 계산 (정확도 개선).
    """
    template_name = 'accounting/cash_flow.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        from_date = self.request.GET.get('from_date', date(today.year, 1, 1).isoformat())
        to_date = self.request.GET.get('to_date', today.isoformat())
        method = self.request.GET.get('method', 'indirect').lower()
        if method not in ('indirect', 'direct'):
            method = 'indirect'
        ctx['from_date'] = from_date
        ctx['to_date'] = to_date
        ctx['method'] = method

        base_qs = VoucherLine.objects.filter(
            is_active=True, voucher__is_active=True,
            voucher__approval_status='APPROVED',
            voucher__voucher_date__gte=from_date,
            voucher__voucher_date__lte=to_date,
        )

        # 영업활동: 수익(REVENUE) - 비용(EXPENSE)
        revenue_lines = base_qs.filter(account__account_type='REVENUE')
        revenue_total = int(
            (revenue_lines.aggregate(s=Sum('credit'))['s'] or 0)
            - (revenue_lines.aggregate(s=Sum('debit'))['s'] or 0)
        )

        expense_lines = base_qs.filter(account__account_type='EXPENSE')
        expense_total = int(
            (expense_lines.aggregate(s=Sum('debit'))['s'] or 0)
            - (expense_lines.aggregate(s=Sum('credit'))['s'] or 0)
        )

        # AR/AP 변동 — 기간 경계잔액 차이 (정밀화)
        # 기초: created_at < from_date 인 AR 미회수분 합계
        # 기말: created_at <= to_date 인 AR 미회수분 합계
        # 변동 = 기말 - 기초 (양수면 현금 유출, 음수면 현금 유입)
        ar_start = int(
            AccountReceivable.objects.filter(
                is_active=True,
                created_at__date__lt=from_date,
            ).aggregate(s=Sum(F('amount') - F('paid_amount')))['s'] or 0
        )
        ar_end = int(
            AccountReceivable.objects.filter(
                is_active=True,
                created_at__date__lte=to_date,
            ).aggregate(s=Sum(F('amount') - F('paid_amount')))['s'] or 0
        )
        ar_change = ar_end - ar_start

        ap_start = int(
            AccountPayable.objects.filter(
                is_active=True,
                created_at__date__lt=from_date,
            ).aggregate(s=Sum(F('amount') - F('paid_amount')))['s'] or 0
        )
        ap_end = int(
            AccountPayable.objects.filter(
                is_active=True,
                created_at__date__lte=to_date,
            ).aggregate(s=Sum(F('amount') - F('paid_amount')))['s'] or 0
        )
        ap_change = ap_end - ap_start

        if method == 'direct':
            # 직접법: 실제 Payment 집계
            pay_in = int(Payment.objects.filter(
                is_active=True,
                payment_type='RECEIPT',
                payment_date__gte=from_date,
                payment_date__lte=to_date,
            ).aggregate(s=Sum('amount'))['s'] or 0)
            pay_out = int(Payment.objects.filter(
                is_active=True,
                payment_type='DISBURSEMENT',
                payment_date__gte=from_date,
                payment_date__lte=to_date,
            ).aggregate(s=Sum('amount'))['s'] or 0)
            operating = pay_in - pay_out
            ctx['direct_inflow'] = pay_in
            ctx['direct_outflow'] = pay_out
        else:
            operating = revenue_total - expense_total - ar_change + ap_change

        # 투자활동: 자산 취득(-), 처분(+)
        from apps.asset.models import FixedAsset
        acquisitions = int(
            FixedAsset.objects.filter(
                is_active=True,
                acquisition_date__gte=from_date,
                acquisition_date__lte=to_date,
            ).aggregate(s=Sum('acquisition_cost'))['s'] or 0
        )
        disposals = int(
            FixedAsset.objects.filter(
                is_active=True,
                disposal_date__gte=from_date,
                disposal_date__lte=to_date,
            ).aggregate(s=Sum('disposal_amount'))['s'] or 0
        )
        investing = disposals - acquisitions

        # 재무활동: 자본 계정 변동
        equity_lines = base_qs.filter(account__account_type='EQUITY')
        financing = int(
            (equity_lines.aggregate(s=Sum('credit'))['s'] or 0)
            - (equity_lines.aggregate(s=Sum('debit'))['s'] or 0)
        )

        net_change = operating + investing + financing

        # 기초 현금 (BankAccount 기초잔액 합계)
        opening_cash = int(
            BankAccount.objects.filter(is_active=True).aggregate(
                s=Sum('opening_balance'),
            )['s'] or 0
        )
        closing_cash = opening_cash + net_change

        ctx.update({
            'revenue_total': revenue_total,
            'expense_total': expense_total,
            'ar_start': ar_start,
            'ar_end': ar_end,
            'ar_change': ar_change,
            'ap_start': ap_start,
            'ap_end': ap_end,
            'ap_change': ap_change,
            'operating': operating,
            'acquisitions': acquisitions,
            'disposals': disposals,
            'investing': investing,
            'financing': financing,
            'net_change': net_change,
            'opening_cash': opening_cash,
            'closing_cash': closing_cash,
        })
        return ctx


# === 카드 관리 ===
class CreditCardListView(ManagerRequiredMixin, ListView):
    model = CreditCard
    template_name = 'accounting/card_list.html'
    context_object_name = 'cards'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related(
            'employee', 'payment_account', 'account_code',
        )
        card_type = self.request.GET.get('card_type')
        if card_type:
            qs = qs.filter(card_type=card_type)
        card_issuer = self.request.GET.get('card_issuer')
        if card_issuer:
            qs = qs.filter(card_issuer=card_issuer)
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(cardholder__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['type_choices'] = CreditCard.CardType.choices
        ctx['issuer_choices'] = CreditCard.CardIssuer.choices
        return ctx


class CreditCardCreateView(ManagerRequiredMixin, CreateView):
    model = CreditCard
    form_class = CreditCardForm
    template_name = 'accounting/card_form.html'
    success_url = reverse_lazy('accounting:card_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CreditCardUpdateView(ManagerRequiredMixin, UpdateView):
    model = CreditCard
    form_class = CreditCardForm
    template_name = 'accounting/card_form.html'
    success_url = reverse_lazy('accounting:card_list')


class CreditCardDetailView(ManagerRequiredMixin, DetailView):
    model = CreditCard
    template_name = 'accounting/card_detail.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['recent_transactions'] = CardTransaction.objects.filter(
            card=self.object, is_active=True,
        ).select_related('partner').order_by('-transaction_date', '-pk')[:10]
        ctx['recent_billings'] = CardBilling.objects.filter(
            card=self.object, is_active=True,
        ).order_by('-billing_month')[:5]
        return ctx


# === 카드 거래 ===
class CardTransactionListView(ManagerRequiredMixin, ListView):
    model = CardTransaction
    template_name = 'accounting/card_transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('card', 'partner')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)
        card_id = self.request.GET.get('card_id')
        if card_id:
            qs = qs.filter(card_id=card_id)
        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(category=category)
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(merchant_name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cards'] = CreditCard.objects.filter(is_active=True)
        ctx['category_choices'] = CardTransaction.Category.choices
        return ctx


class CardTransactionCreateView(ManagerRequiredMixin, CreateView):
    model = CardTransaction
    form_class = CardTransactionForm
    template_name = 'accounting/card_transaction_form.html'
    success_url = reverse_lazy('accounting:card_transaction_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CardTransactionUpdateView(ManagerRequiredMixin, UpdateView):
    model = CardTransaction
    form_class = CardTransactionForm
    template_name = 'accounting/card_transaction_form.html'
    success_url = reverse_lazy('accounting:card_transaction_list')


# === 카드 청구 ===
class CardBillingListView(ManagerRequiredMixin, ListView):
    model = CardBilling
    template_name = 'accounting/card_billing_list.html'
    context_object_name = 'billings'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('card')
        card_id = self.request.GET.get('card_id')
        if card_id:
            qs = qs.filter(card_id=card_id)
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cards'] = CreditCard.objects.filter(is_active=True)
        ctx['status_choices'] = CardBilling.Status.choices
        return ctx


class CardBillingDetailView(ManagerRequiredMixin, DetailView):
    model = CardBilling
    template_name = 'accounting/card_billing_detail.html'

    def get_queryset(self):
        return super().get_queryset().select_related('card', 'payment')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['billing_transactions'] = CardTransaction.objects.filter(
            billing=self.object, is_active=True,
        ).select_related('card', 'partner').order_by('-transaction_date')
        return ctx


class CardBillingPayView(ManagerRequiredMixin, View):
    """카드 청구 결제 처리 (POST only)"""

    def post(self, request, pk):
        with transaction.atomic():
            billing = CardBilling.objects.select_for_update().get(
                pk=pk, is_active=True,
            )

            if billing.status == 'PAID':
                messages.error(request, '이미 결제 완료된 청구서입니다.')
                return redirect('accounting:card_billing_detail', pk=pk)

            card = billing.card
            if not card.payment_account:
                messages.error(request, '카드에 결제계좌가 설정되지 않았습니다.')
                return redirect('accounting:card_billing_detail', pk=pk)

            pay_amount = billing.remaining_amount
            if pay_amount <= 0:
                messages.error(request, '결제할 금액이 없습니다.')
                return redirect('accounting:card_billing_detail', pk=pk)

            # Payment(DISBURSEMENT) 생성 — 거래처는 카드 거래내역에서 추출
            from apps.core.utils import generate_document_number
            from apps.sales.models import Partner
            # 카드 청구 결제용 거래처: 청구서 내 거래의 partner 또는 기본 거래처
            billing_partner = (
                CardTransaction.objects.filter(
                    billing=billing, partner__isnull=False, is_active=True,
                ).values_list('partner', flat=True).first()
            )
            if billing_partner:
                pay_partner = Partner.objects.get(pk=billing_partner)
            else:
                pay_partner, _ = Partner.objects.get_or_create(
                    code='CARD-ISSUER',
                    defaults={
                        'name': '카드사결제',
                        'partner_type': 'SUPPLIER',
                    },
                )
            payment = Payment.objects.create(
                payment_number=generate_document_number(Payment, 'payment_number', 'PM'),
                payment_type='DISBURSEMENT',
                partner=pay_partner,
                bank_account=card.payment_account,
                amount=pay_amount,
                payment_date=date.today(),
                payment_method='CARD',
                reference=f'카드청구결제: {card.name} {billing.billing_month.strftime("%Y-%m")}',
                created_by=request.user,
            )

            # CardBilling 갱신 (F() 원자적 업데이트)
            CardBilling.objects.filter(pk=billing.pk).update(
                paid_amount=F('paid_amount') + pay_amount,
                payment=payment,
            )
            billing.refresh_from_db()

            if billing.paid_amount >= billing.total_amount:
                new_status = 'PAID'
            elif billing.paid_amount > 0:
                new_status = 'PARTIAL'
            else:
                new_status = billing.status
            if new_status != billing.status:
                CardBilling.objects.filter(pk=billing.pk).update(status=new_status)

        messages.success(request, f'{pay_amount:,}원 카드 청구 결제 처리 완료')
        return redirect('accounting:card_billing_detail', pk=pk)


# === 환차손익 ===

class ExchangeGainLossView(ManagerRequiredMixin, TemplateView):
    """환율 변동에 따른 외화 미수/미지급 환차손익 계산"""
    template_name = 'accounting/exchange_gain_loss.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # 외화 주문 기반 AR (KRW가 아닌 통화)
        from apps.sales.models import Order
        from apps.purchase.models import PurchaseOrder

        ar_rows = []
        for ar in AccountReceivable.objects.filter(
            is_active=True, status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
            order__isnull=False, order__currency__isnull=False,
        ).select_related('partner', 'order__currency'):
            order = ar.order
            if not order.currency or order.currency.is_base:
                continue
            # 잔액 (외화 기준)
            remaining_krw = ar.remaining_amount
            original_rate = order.exchange_rate or 1
            if original_rate <= 0:
                continue
            foreign_remaining = remaining_krw / original_rate

            # 현재 환율 조회
            current_rate_obj = ExchangeRate.objects.filter(
                currency=order.currency, is_active=True,
            ).order_by('-rate_date').first()
            current_rate = current_rate_obj.rate if current_rate_obj else original_rate

            revalued_krw = int(foreign_remaining * current_rate)
            gain_loss = revalued_krw - int(remaining_krw)

            ar_rows.append({
                'partner': ar.partner.name,
                'currency': order.currency.code,
                'foreign_amount': round(foreign_remaining, 2),
                'original_rate': original_rate,
                'original_krw': int(remaining_krw),
                'current_rate': current_rate,
                'revalued_krw': revalued_krw,
                'gain_loss': gain_loss,
                'type': 'AR',
            })

        ap_rows = []
        for ap in AccountPayable.objects.filter(
            is_active=True, status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
            purchase_order__isnull=False,
            purchase_order__currency__isnull=False,
        ).select_related('partner', 'purchase_order__currency'):
            po = ap.purchase_order
            if not po.currency or po.currency.is_base:
                continue
            remaining_krw = ap.remaining_amount
            original_rate = po.exchange_rate or 1
            if original_rate <= 0:
                continue
            foreign_remaining = remaining_krw / original_rate

            current_rate_obj = ExchangeRate.objects.filter(
                currency=po.currency, is_active=True,
            ).order_by('-rate_date').first()
            current_rate = current_rate_obj.rate if current_rate_obj else original_rate

            revalued_krw = int(foreign_remaining * current_rate)
            gain_loss = revalued_krw - int(remaining_krw)

            ap_rows.append({
                'partner': ap.partner.name,
                'currency': po.currency.code,
                'foreign_amount': round(foreign_remaining, 2),
                'original_rate': original_rate,
                'original_krw': int(remaining_krw),
                'current_rate': current_rate,
                'revalued_krw': revalued_krw,
                'gain_loss': gain_loss,
                'type': 'AP',
            })

        all_rows = ar_rows + ap_rows
        total_gain = sum(r['gain_loss'] for r in all_rows if r['gain_loss'] > 0)
        total_loss = sum(r['gain_loss'] for r in all_rows if r['gain_loss'] < 0)
        net_gain_loss = total_gain + total_loss

        ctx['ar_rows'] = ar_rows
        ctx['ap_rows'] = ap_rows
        ctx['total_gain'] = total_gain
        ctx['total_loss'] = total_loss
        ctx['net_gain_loss'] = net_gain_loss
        return ctx


# ── 부가세 신고서 ────────────────────────────────────────

class VATReturnView(ManagerRequiredMixin, TemplateView):
    """부가가치세 신고서 — 분기별 매출/매입 + 4구분(과세/영세율/면세/기타) 집계

    매출 구성 (TaxInvoice.tax_type + TaxInvoice.issuer_type 고려):
      1) 과세 자사발행 — 세금계산서 발행, 매출세액 대상
      2) 과세 플랫폼대행 — 네이버/쿠팡 등 플랫폼 대행 발행, 자사 매출세액에서 제외
      3) 영세율 — 공급가액만, 부가세 0원
      4) 면세 — 공급가액만, 부가세 0원
      5) 현금영수증 매출 — 세금계산서 미발행건 보조 집계

    매입:
      1) 과세 매입 (매입세액공제 대상)
      2) 면세/불공제 매입

    납부세액 = 과세매출세액(자사발행분) − 과세매입세액
    """
    template_name = 'accounting/vat_return.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        quarter = safe_int(self.request.GET.get('quarter'), (date.today().month - 1) // 3 + 1)
        quarter = max(1, min(4, quarter))

        ctx['year'] = year
        ctx['quarter'] = quarter
        ctx['years'] = list(range(date.today().year, date.today().year - 5, -1))
        ctx['quarters'] = [1, 2, 3, 4]

        import calendar as _cal
        m_start = (quarter - 1) * 3 + 1
        m_end = m_start + 2
        period_start = date(year, m_start, 1)
        period_end = date(year, m_end, _cal.monthrange(year, m_end)[1])

        ctx['period_start'] = period_start
        ctx['period_end'] = period_end

        # ── 매출 4구분 집계 ──
        base_sales_qs = TaxInvoice.objects.filter(
            is_active=True,
            invoice_type=TaxInvoice.InvoiceType.SALES,
            issue_date__gte=period_start,
            issue_date__lte=period_end,
        )

        def _sum(qs, **filters):
            agg = qs.filter(**filters).aggregate(
                supply=Sum('supply_amount'),
                tax=Sum('tax_amount'),
                cnt=Count('pk'),
            )
            return {
                'supply': agg['supply'] or 0,
                'tax': agg['tax'] or 0,
                'count': agg['cnt'] or 0,
            }

        sales_taxable_self = _sum(
            base_sales_qs,
            tax_type=TaxInvoice.TaxType.TAXABLE,
            issuer_type=TaxInvoice.IssuerType.SELF,
        )
        sales_taxable_platform = _sum(
            base_sales_qs,
            tax_type=TaxInvoice.TaxType.TAXABLE,
        )
        # 위는 자사+플랫폼+세무사 통합이므로 "플랫폼/세무사분"만 별도 계산
        sales_taxable_outsourced_agg = base_sales_qs.filter(
            tax_type=TaxInvoice.TaxType.TAXABLE,
        ).exclude(
            issuer_type=TaxInvoice.IssuerType.SELF,
        ).aggregate(
            supply=Sum('supply_amount'),
            tax=Sum('tax_amount'),
            cnt=Count('pk'),
        )
        sales_taxable_outsourced = {
            'supply': sales_taxable_outsourced_agg['supply'] or 0,
            'tax': sales_taxable_outsourced_agg['tax'] or 0,
            'count': sales_taxable_outsourced_agg['cnt'] or 0,
        }

        sales_zero_rate = _sum(base_sales_qs, tax_type=TaxInvoice.TaxType.ZERO_RATE)
        sales_exempt = _sum(base_sales_qs, tax_type=TaxInvoice.TaxType.EXEMPT)

        # ── 현금영수증 매출 (보조 집계) ──
        from .models import CashReceipt
        cash_receipt_agg = CashReceipt.objects.filter(
            is_active=True,
            status=CashReceipt.Status.ISSUED,
            issued_at__date__gte=period_start,
            issued_at__date__lte=period_end,
        ).aggregate(
            supply=Sum('supply_amount'),
            tax=Sum('vat'),
            cnt=Count('pk'),
        )
        cash_receipts_total = {
            'supply': cash_receipt_agg['supply'] or 0,
            'tax': cash_receipt_agg['tax'] or 0,
            'count': cash_receipt_agg['cnt'] or 0,
        }

        # ── 매입 4구분 ──
        base_purchase_qs = TaxInvoice.objects.filter(
            is_active=True,
            invoice_type=TaxInvoice.InvoiceType.PURCHASE,
            issue_date__gte=period_start,
            issue_date__lte=period_end,
        )
        # 일반매입(공제) — 과세 + DEDUCTIBLE
        purchase_taxable = _sum(
            base_purchase_qs,
            tax_type=TaxInvoice.TaxType.TAXABLE,
            vat_deduction_type=TaxInvoice.VatDeductionType.DEDUCTIBLE,
        )
        # 의제매입세액 — 면세 농축수산물 등 공급가액 일부를 세액공제
        purchase_deemed = _sum(
            base_purchase_qs,
            vat_deduction_type=TaxInvoice.VatDeductionType.DEEMED,
        )
        # 공제받지못할매입(접대비·사업무관) — 집계만, 세액공제 제외
        purchase_non_deductible = _sum(
            base_purchase_qs,
            vat_deduction_type=TaxInvoice.VatDeductionType.NON_DEDUCTIBLE,
        )
        # 면세 매입 (세액 없음)
        purchase_exempt_agg = base_purchase_qs.exclude(
            tax_type=TaxInvoice.TaxType.TAXABLE,
        ).aggregate(
            supply=Sum('supply_amount'),
            tax=Sum('tax_amount'),
            cnt=Count('pk'),
        )
        purchase_exempt = {
            'supply': purchase_exempt_agg['supply'] or 0,
            'tax': purchase_exempt_agg['tax'] or 0,
            'count': purchase_exempt_agg['cnt'] or 0,
        }

        # ── 매출세액(자사발행분만, 플랫폼대행은 플랫폼이 납부) ──
        sales_output_tax = sales_taxable_self['tax']
        # 매입세액공제 = 일반매입 + 의제매입세액 (공제받지못할매입은 제외)
        purchase_input_tax = purchase_taxable['tax'] + purchase_deemed['tax']
        payable_tax = sales_output_tax - purchase_input_tax

        # 총 매출(공급가액 기준) — 부가세신고서 상단 공급가액 합계
        total_sales_supply = (
            sales_taxable_self['supply']
            + sales_taxable_outsourced['supply']
            + sales_zero_rate['supply']
            + sales_exempt['supply']
        )
        total_sales_tax = (
            sales_taxable_self['tax']
            + sales_taxable_outsourced['tax']
        )

        total_purchase_supply = (
            purchase_taxable['supply'] + purchase_exempt['supply']
        )

        ctx.update({
            'sales_taxable_self': sales_taxable_self,
            'sales_taxable_outsourced': sales_taxable_outsourced,
            'sales_zero_rate': sales_zero_rate,
            'sales_exempt': sales_exempt,
            'cash_receipts_total': cash_receipts_total,
            'purchase_taxable': purchase_taxable,
            'purchase_deemed': purchase_deemed,
            'purchase_non_deductible': purchase_non_deductible,
            'purchase_exempt': purchase_exempt,
            'sales_output_tax': sales_output_tax,
            'purchase_input_tax': purchase_input_tax,
            'payable_tax': payable_tax,
            'total_sales_supply': total_sales_supply,
            'total_sales_tax': total_sales_tax,
            'total_purchase_supply': total_purchase_supply,
            # 구버전 호환
            'sales_supply': total_sales_supply,
            'sales_tax': total_sales_tax,
            'purchase_supply': total_purchase_supply,
            'purchase_tax': purchase_input_tax,
        })
        return ctx


class VATReturnExcelView(VATReturnView):
    """부가세 신고서 Excel 내보내기 — VATReturnView 의 집계 결과를 재사용."""

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment

        ctx = self.get_context_data(**kwargs)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '부가세신고서'
        bold = Font(bold=True)
        center = Alignment(horizontal='center')

        ws.cell(row=1, column=1, value=(
            f'{ctx["year"]}년 {ctx["quarter"]}분기 부가가치세 신고서'
        )).font = bold
        ws.cell(row=2, column=1, value=(
            f'기간: {ctx["period_start"]} ~ {ctx["period_end"]}'
        ))

        # 매출 섹션
        ws.cell(row=4, column=1, value='【매출 4구분】').font = bold
        row = 5
        for label, data in [
            ('① 과세 자사발행', ctx['sales_taxable_self']),
            ('② 과세 플랫폼/세무사대행', ctx['sales_taxable_outsourced']),
            ('③ 영세율(수출)', ctx['sales_zero_rate']),
            ('④ 면세', ctx['sales_exempt']),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=int(data['supply']))
            ws.cell(row=row, column=3, value=int(data['tax']))
            ws.cell(row=row, column=4, value=int(data['count']))
            row += 1
        ws.cell(row=row, column=1, value='(보조) 현금영수증 매출')
        cr = ctx['cash_receipts_total']
        ws.cell(row=row, column=2, value=int(cr['supply']))
        ws.cell(row=row, column=3, value=int(cr['tax']))
        ws.cell(row=row, column=4, value=int(cr['count']))
        row += 2

        # 매입 섹션
        ws.cell(row=row, column=1, value='【매입 4구분】').font = bold
        row += 1
        for label, data in [
            ('① 일반매입(공제)', ctx['purchase_taxable']),
            ('② 의제매입세액', ctx['purchase_deemed']),
            ('③ 공제받지못할매입', ctx['purchase_non_deductible']),
            ('④ 면세 매입', ctx['purchase_exempt']),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=int(data['supply']))
            ws.cell(row=row, column=3, value=int(data['tax']))
            ws.cell(row=row, column=4, value=int(data['count']))
            row += 1
        row += 1

        # 납부세액
        ws.cell(row=row, column=1, value='【납부세액】').font = bold
        row += 1
        ws.cell(row=row, column=1, value='매출세액(자사발행 과세분)')
        ws.cell(row=row, column=3, value=int(ctx['sales_output_tax']))
        row += 1
        ws.cell(row=row, column=1, value='매입세액공제(일반 + 의제)')
        ws.cell(row=row, column=3, value=int(ctx['purchase_input_tax']))
        row += 1
        ws.cell(row=row, column=1, value='납부할 세액').font = bold
        c = ws.cell(row=row, column=3, value=int(ctx['payable_tax']))
        c.font = bold
        c.alignment = center

        ws.cell(row=4, column=1).alignment = center
        for col in [1, 2, 3, 4]:
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 28

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename=vat_return_{ctx["year"]}Q{ctx["quarter"]}.xlsx'
        )
        wb.save(response)
        return response


class VATReturnPDFView(VATReturnView):
    """부가세 신고서 PDF 내보내기 — ReportLab A4 단면."""

    def get(self, request, *args, **kwargs):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from apps.core.pdf import _get_font

        ctx = self.get_context_data(**kwargs)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename=vat_return_{ctx["year"]}Q{ctx["quarter"]}.pdf'
        )

        doc = SimpleDocTemplate(response, pagesize=A4)
        font = _get_font()
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.fontName = font
        body_style = styles['BodyText']
        body_style.fontName = font

        flowables = []
        flowables.append(Paragraph(
            f'{ctx["year"]}년 {ctx["quarter"]}분기 부가가치세 신고서',
            title_style,
        ))
        flowables.append(Paragraph(
            f'기간: {ctx["period_start"]} ~ {ctx["period_end"]}',
            body_style,
        ))
        flowables.append(Spacer(1, 12))

        def _tbl(title, rows):
            flowables.append(Paragraph(f'<b>{title}</b>', body_style))
            data = [['구분', '공급가액', '세액', '건수']]
            for r in rows:
                data.append([r[0], f'{int(r[1]["supply"]):,}', f'{int(r[1]["tax"]):,}', r[1]['count']])
            t = Table(data, colWidths=[160, 110, 110, 60])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 0), (-1, -1), font),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]))
            flowables.append(t)
            flowables.append(Spacer(1, 12))

        _tbl('매출 4구분', [
            ('① 과세 자사발행', ctx['sales_taxable_self']),
            ('② 과세 플랫폼/세무사대행', ctx['sales_taxable_outsourced']),
            ('③ 영세율(수출)', ctx['sales_zero_rate']),
            ('④ 면세', ctx['sales_exempt']),
        ])
        _tbl('매입 4구분', [
            ('① 일반매입(공제)', ctx['purchase_taxable']),
            ('② 의제매입세액', ctx['purchase_deemed']),
            ('③ 공제받지못할매입', ctx['purchase_non_deductible']),
            ('④ 면세 매입', ctx['purchase_exempt']),
        ])

        summary = [
            ['매출세액(자사발행 과세분)', f'{int(ctx["sales_output_tax"]):,}'],
            ['매입세액공제(일반+의제)', f'{int(ctx["purchase_input_tax"]):,}'],
            ['납부할 세액', f'{int(ctx["payable_tax"]):,}'],
        ]
        t = Table(summary, colWidths=[260, 180])
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
        ]))
        flowables.append(t)

        doc.build(flowables)
        return response


# ── 원가센터 ────────────────────────────────────────

class CostCenterListView(ManagerRequiredMixin, ListView):
    """원가센터 목록 (트리 구조)"""
    model = CostCenter
    template_name = 'accounting/cost_center_list.html'
    context_object_name = 'cost_centers'
    paginate_by = 20

    def get_queryset(self):
        return CostCenter.objects.filter(
            is_active=True,
        ).select_related('parent', 'department', 'manager')


class CostCenterCreateView(ManagerRequiredMixin, CreateView):
    model = CostCenter
    form_class = CostCenterForm
    template_name = 'accounting/cost_center_form.html'
    success_url = reverse_lazy('accounting:cost_center_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CostCenterUpdateView(ManagerRequiredMixin, UpdateView):
    model = CostCenter
    form_class = CostCenterForm
    template_name = 'accounting/cost_center_form.html'
    success_url = reverse_lazy('accounting:cost_center_list')


class CostCenterReportView(ManagerRequiredMixin, TemplateView):
    """원가센터별 차변/대변 집계 리포트"""
    template_name = 'accounting/cost_center_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        center_type = self.request.GET.get('center_type', '')
        ctx['date_from'] = date_from
        ctx['date_to'] = date_to
        ctx['center_type'] = center_type

        qs = VoucherLine.objects.filter(
            is_active=True,
            voucher__is_active=True,
            voucher__approval_status='APPROVED',
            cost_center__isnull=False,
        )
        if date_from:
            qs = qs.filter(voucher__voucher_date__gte=date_from)
        if date_to:
            qs = qs.filter(voucher__voucher_date__lte=date_to)
        if center_type:
            qs = qs.filter(cost_center__center_type=center_type)

        totals = qs.values(
            'cost_center__pk', 'cost_center__code', 'cost_center__name',
            'cost_center__center_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('cost_center__code')

        rows = []
        grand_debit = 0
        grand_credit = 0
        for row in totals:
            d = int(row['total_debit'] or 0)
            c = int(row['total_credit'] or 0)
            grand_debit += d
            grand_credit += c
            rows.append({
                'pk': row['cost_center__pk'],
                'code': row['cost_center__code'],
                'name': row['cost_center__name'],
                'center_type': row['cost_center__center_type'],
                'debit': d,
                'credit': c,
                'balance': d - c,
            })

        ctx['rows'] = rows
        ctx['grand_debit'] = grand_debit
        ctx['grand_credit'] = grand_credit
        return ctx


class ProfitCenterReportView(ManagerRequiredMixin, TemplateView):
    """이익센터별 수익/비용/이익 분석"""
    template_name = 'accounting/profit_center_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        ctx['date_from'] = date_from
        ctx['date_to'] = date_to

        # PROFIT/INVESTMENT 센터만 대상
        qs = VoucherLine.objects.filter(
            is_active=True,
            voucher__is_active=True,
            voucher__approval_status='APPROVED',
            cost_center__isnull=False,
            cost_center__center_type__in=['PROFIT', 'INVESTMENT'],
        )
        if date_from:
            qs = qs.filter(voucher__voucher_date__gte=date_from)
        if date_to:
            qs = qs.filter(voucher__voucher_date__lte=date_to)

        # 센터별 + 계정유형별 집계
        raw = qs.values(
            'cost_center__pk', 'cost_center__code', 'cost_center__name',
            'cost_center__center_type', 'account__account_type',
        ).annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
        ).order_by('cost_center__code', 'account__account_type')

        centers = {}
        for row in raw:
            pk = row['cost_center__pk']
            if pk not in centers:
                centers[pk] = {
                    'pk': pk,
                    'code': row['cost_center__code'],
                    'name': row['cost_center__name'],
                    'center_type': row['cost_center__center_type'],
                    'revenue': 0,
                    'expense': 0,
                }
            d = int(row['total_debit'] or 0)
            c = int(row['total_credit'] or 0)
            acct_type = row['account__account_type']
            if acct_type == 'REVENUE':
                centers[pk]['revenue'] += c - d
            elif acct_type == 'EXPENSE':
                centers[pk]['expense'] += d - c

        rows = []
        total_revenue = 0
        total_expense = 0
        for center in sorted(centers.values(), key=lambda x: x['code']):
            profit = center['revenue'] - center['expense']
            margin = round(profit / center['revenue'] * 100, 1) if center['revenue'] > 0 else 0
            center['profit'] = profit
            center['margin'] = margin
            total_revenue += center['revenue']
            total_expense += center['expense']
            rows.append(center)

        total_profit = total_revenue - total_expense
        ctx['rows'] = rows
        ctx['total_revenue'] = total_revenue
        ctx['total_expense'] = total_expense
        ctx['total_profit'] = total_profit
        ctx['total_margin'] = round(total_profit / total_revenue * 100, 1) if total_revenue > 0 else 0
        return ctx


class AgedTrialBalanceView(ManagerRequiredMixin, TemplateView):
    """매출채권/매입채무 경과기간별 시산표"""
    template_name = 'accounting/aged_trial_balance.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        report_type = self.request.GET.get('type', 'AR')
        ctx['report_type'] = report_type
        ctx['today'] = today

        if report_type == 'AP':
            qs = AccountPayable.objects.filter(
                is_active=True,
                status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
            ).select_related('partner')
        else:
            qs = AccountReceivable.objects.filter(
                is_active=True,
                status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
            ).select_related('partner')

        partner_data = {}
        grand_totals = {
            'current': 0, 'b30': 0, 'b60': 0,
            'b90': 0, 'b120': 0, 'b120_plus': 0, 'total': 0,
        }

        for item in qs:
            remaining = int(item.amount) - int(item.paid_amount)
            if remaining <= 0:
                continue
            days = (today - item.due_date).days
            partner_name = item.partner.name
            partner_id = item.partner_id

            if partner_id not in partner_data:
                partner_data[partner_id] = {
                    'name': partner_name,
                    'current': 0, 'b30': 0, 'b60': 0,
                    'b90': 0, 'b120': 0, 'b120_plus': 0, 'total': 0,
                }

            if days <= 0:
                bucket = 'current'
            elif days <= 30:
                bucket = 'b30'
            elif days <= 60:
                bucket = 'b60'
            elif days <= 90:
                bucket = 'b90'
            elif days <= 120:
                bucket = 'b120'
            else:
                bucket = 'b120_plus'

            partner_data[partner_id][bucket] += remaining
            partner_data[partner_id]['total'] += remaining
            grand_totals[bucket] += remaining
            grand_totals['total'] += remaining

        ctx['partners'] = sorted(
            partner_data.values(), key=lambda x: -x['total'],
        )
        ctx['grand_totals'] = grand_totals
        return ctx


class AdvancedReportView(ManagerRequiredMixin, TemplateView):
    """고급 재무 리포트 — YoY/MoM, 추이, 거래처별, 제품별"""
    template_name = 'accounting/advanced_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        year = safe_int(self.request.GET.get('year'), today.year)
        month = safe_int(self.request.GET.get('month'), today.month)
        top_n = safe_int(self.request.GET.get('top_n'), 10)
        ctx['year'] = year
        ctx['month'] = month
        ctx['top_n'] = top_n
        ctx['years'] = list(range(today.year, today.year - 5, -1))

        # --- 월별 매출/매입 12개월 추이 ---
        monthly_data = []
        for m_offset in range(11, -1, -1):
            y = year
            m = month - m_offset
            while m <= 0:
                m += 12
                y -= 1

            if m == 12:
                next_y, next_m = y + 1, 1
            else:
                next_y, next_m = y, m + 1

            sales_total = TaxInvoice.objects.filter(
                is_active=True, invoice_type='SALES',
                issue_date__gte=date(y, m, 1),
                issue_date__lt=date(next_y, next_m, 1),
            ).aggregate(t=Sum('supply_amount'))['t'] or 0

            purchase_total = TaxInvoice.objects.filter(
                is_active=True, invoice_type='PURCHASE',
                issue_date__gte=date(y, m, 1),
                issue_date__lt=date(next_y, next_m, 1),
            ).aggregate(t=Sum('supply_amount'))['t'] or 0

            monthly_data.append({
                'label': f'{y}-{m:02d}',
                'sales': int(sales_total),
                'purchase': int(purchase_total),
                'profit': int(sales_total) - int(purchase_total),
            })

        ctx['monthly_data'] = monthly_data
        ctx['monthly_json'] = json.dumps(monthly_data)

        # --- YoY / MoM 비교 ---
        def _period_totals(y, m):
            if m == 12:
                ny, nm = y + 1, 1
            else:
                ny, nm = y, m + 1
            s = TaxInvoice.objects.filter(
                is_active=True, invoice_type='SALES',
                issue_date__gte=date(y, m, 1),
                issue_date__lt=date(ny, nm, 1),
            ).aggregate(t=Sum('supply_amount'))['t'] or 0
            p = TaxInvoice.objects.filter(
                is_active=True, invoice_type='PURCHASE',
                issue_date__gte=date(y, m, 1),
                issue_date__lt=date(ny, nm, 1),
            ).aggregate(t=Sum('supply_amount'))['t'] or 0
            return int(s), int(p)

        cur_sales, cur_purchase = _period_totals(year, month)

        # 전월
        prev_m_year = year if month > 1 else year - 1
        prev_m_month = month - 1 if month > 1 else 12
        prev_m_sales, prev_m_purchase = _period_totals(prev_m_year, prev_m_month)

        # 전년동월
        prev_y_sales, prev_y_purchase = _period_totals(year - 1, month)

        def _change_rate(cur, prev):
            if prev == 0:
                return None
            return round((cur - prev) / prev * 100, 1)

        ctx['current'] = {
            'sales': cur_sales, 'purchase': cur_purchase,
            'profit': cur_sales - cur_purchase,
        }
        ctx['mom'] = {
            'sales': prev_m_sales, 'purchase': prev_m_purchase,
            'profit': prev_m_sales - prev_m_purchase,
            'sales_rate': _change_rate(cur_sales, prev_m_sales),
            'purchase_rate': _change_rate(cur_purchase, prev_m_purchase),
        }
        ctx['yoy'] = {
            'sales': prev_y_sales, 'purchase': prev_y_purchase,
            'profit': prev_y_sales - prev_y_purchase,
            'sales_rate': _change_rate(cur_sales, prev_y_sales),
            'purchase_rate': _change_rate(cur_purchase, prev_y_purchase),
        }

        # --- 거래처별 상위 N 매출 ---
        top_partners = TaxInvoice.objects.filter(
            is_active=True, invoice_type='SALES',
            issue_date__year=year,
        ).values(
            'partner__name',
        ).annotate(
            total=Sum('supply_amount'),
        ).order_by('-total')[:top_n]
        ctx['top_partners'] = list(top_partners)

        # --- 제품별 수익성 (주문 기반) ---
        product_rows = OrderItem.objects.filter(
            is_active=True,
            order__is_active=True,
            order__status='CONFIRMED',
            order__order_date__year=year,
        ).values(
            'product__name',
        ).annotate(
            total_revenue=Sum(F('quantity') * F('unit_price')),
            total_cost=Sum(F('quantity') * F('product__cost_price')),
        ).order_by('-total_revenue')[:top_n]

        product_data = []
        for row in product_rows:
            revenue = int(row['total_revenue'] or 0)
            cost = int(row['total_cost'] or 0)
            profit = revenue - cost
            margin = round(profit / revenue * 100, 1) if revenue > 0 else 0
            product_data.append({
                'name': row['product__name'],
                'revenue': revenue,
                'cost': cost,
                'profit': profit,
                'margin': margin,
            })
        ctx['product_data'] = product_data
        return ctx


# ──── Phase 15: 다중법인/연결회계 ────

class CompanyListView(ManagerRequiredMixin, ListView):
    model = None  # set below
    template_name = 'accounting/company_list.html'
    context_object_name = 'companies'
    paginate_by = 20

    def get_queryset(self):
        from .models import Company
        return Company.objects.filter(is_active=True).select_related('parent', 'currency')

    def get(self, request, *args, **kwargs):
        from .models import Company
        self.model = Company
        return super().get(request, *args, **kwargs)


class CompanyCreateView(ManagerRequiredMixin, CreateView):
    template_name = 'accounting/company_form.html'
    success_url = reverse_lazy('accounting:company_list')

    def get_form_class(self):
        from .forms import CompanyForm
        return CompanyForm

    def get_model(self):
        from .models import Company
        return Company

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '법인이 등록되었습니다.')
        return super().form_valid(form)


class CompanyUpdateView(ManagerRequiredMixin, UpdateView):
    template_name = 'accounting/company_form.html'
    success_url = reverse_lazy('accounting:company_list')

    def get_form_class(self):
        from .forms import CompanyForm
        return CompanyForm

    def get_queryset(self):
        from .models import Company
        return Company.objects.filter(is_active=True)

    def form_valid(self, form):
        messages.success(self.request, '법인 정보가 수정되었습니다.')
        return super().form_valid(form)


class InterCompanyTransactionListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/ic_transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 20

    def get_queryset(self):
        from .models import InterCompanyTransaction
        return InterCompanyTransaction.objects.filter(is_active=True).select_related(
            'from_company', 'to_company',
        )


class InterCompanyTransactionCreateView(ManagerRequiredMixin, CreateView):
    template_name = 'accounting/ic_transaction_form.html'
    success_url = reverse_lazy('accounting:ic_transaction_list')

    def get_form_class(self):
        from .forms import InterCompanyTransactionForm
        return InterCompanyTransactionForm

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '내부거래가 등록되었습니다.')
        return super().form_valid(form)


class InterCompanyTransactionConfirmView(ManagerRequiredMixin, View):
    def post(self, request, pk):
        from .models import InterCompanyTransaction
        txn = get_object_or_404(InterCompanyTransaction, pk=pk, is_active=True)
        txn.status = InterCompanyTransaction.Status.CONFIRMED
        txn.save(update_fields=['status', 'updated_at'])
        messages.success(request, '내부거래가 확정되었습니다.')
        return redirect('accounting:ic_transaction_list')


class ConsolidationPeriodListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/consolidation_list.html'
    context_object_name = 'periods'
    paginate_by = 20

    def get_queryset(self):
        from .models import ConsolidationPeriod
        return ConsolidationPeriod.objects.filter(is_active=True).prefetch_related('companies')


class ConsolidationPeriodCreateView(ManagerRequiredMixin, CreateView):
    template_name = 'accounting/consolidation_form.html'
    success_url = reverse_lazy('accounting:consolidation_list')

    def get_form_class(self):
        from .forms import ConsolidationPeriodForm
        return ConsolidationPeriodForm

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '연결결산 기간이 생성되었습니다.')
        return super().form_valid(form)


class ConsolidationPeriodCloseView(ManagerRequiredMixin, View):
    def post(self, request, pk):
        from .models import ConsolidationPeriod, InterCompanyTransaction
        period = get_object_or_404(ConsolidationPeriod, pk=pk, is_active=True)

        # Collect elimination entries from confirmed IC transactions
        from datetime import date as date_cls
        start = date_cls(period.year, period.month, 1)
        if period.month == 12:
            end = date_cls(period.year + 1, 1, 1)
        else:
            end = date_cls(period.year, period.month + 1, 1)

        ic_txns = InterCompanyTransaction.objects.filter(
            is_active=True,
            status=InterCompanyTransaction.Status.CONFIRMED,
            transaction_date__gte=start,
            transaction_date__lt=end,
        )
        eliminations = []
        for txn in ic_txns:
            eliminations.append({
                'from': txn.from_company.code,
                'to': txn.to_company.code,
                'amount': str(txn.amount),
                'description': txn.description,
            })
            txn.status = InterCompanyTransaction.Status.ELIMINATED
            txn.save(update_fields=['status', 'updated_at'])

        period.elimination_entries = eliminations
        period.status = ConsolidationPeriod.Status.CLOSED
        period.consolidated_at = timezone.now()
        period.save(update_fields=['elimination_entries', 'status', 'consolidated_at', 'updated_at'])
        messages.success(request, '연결결산이 마감되었습니다.')
        return redirect('accounting:consolidation_list')


class ConsolidatedReportView(ManagerRequiredMixin, DetailView):
    template_name = 'accounting/consolidated_report.html'

    def get_object(self):
        from .models import ConsolidationPeriod
        return get_object_or_404(ConsolidationPeriod, pk=self.kwargs['pk'], is_active=True)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        period = self.object
        ctx['period'] = period
        ctx['reports'] = period.reports.filter(is_active=True)
        ctx['elimination_entries'] = period.elimination_entries
        return ctx


# ──── Phase 15: 오픈뱅킹 연동 ────

class BankConnectionListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/bank_connection_list.html'
    context_object_name = 'connections'
    paginate_by = 20

    def get_queryset(self):
        from .models import BankConnection
        return BankConnection.objects.filter(is_active=True).select_related('company')


class BankConnectionCreateView(ManagerRequiredMixin, CreateView):
    template_name = 'accounting/bank_connection_form.html'
    success_url = reverse_lazy('accounting:bank_connection_list')

    def get_form_class(self):
        from .forms import BankConnectionForm
        return BankConnectionForm

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '은행 연결이 등록되었습니다.')
        return super().form_valid(form)


class BankConnectionUpdateView(ManagerRequiredMixin, UpdateView):
    template_name = 'accounting/bank_connection_form.html'
    success_url = reverse_lazy('accounting:bank_connection_list')

    def get_form_class(self):
        from .forms import BankConnectionForm
        return BankConnectionForm

    def get_queryset(self):
        from .models import BankConnection
        return BankConnection.objects.filter(is_active=True)

    def form_valid(self, form):
        messages.success(self.request, '은행 연결이 수정되었습니다.')
        return super().form_valid(form)


class BankConnectionSyncView(ManagerRequiredMixin, View):
    def post(self, request, pk):
        from .models import BankConnection
        conn = get_object_or_404(BankConnection, pk=pk, is_active=True)
        conn.last_sync = timezone.now()
        conn.save(update_fields=['last_sync', 'updated_at'])
        messages.success(request, f'{conn.bank_name} 동기화가 완료되었습니다.')
        return redirect('accounting:bank_connection_list')


class BankStatementListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/bank_statement_list.html'
    context_object_name = 'statements'
    paginate_by = 20

    def get_queryset(self):
        from .models import BankStatement
        qs = BankStatement.objects.filter(is_active=True).select_related('connection')
        connection = self.request.GET.get('connection')
        if connection:
            qs = qs.filter(connection_id=connection)
        return qs


class BankStatementDetailView(ManagerRequiredMixin, DetailView):
    template_name = 'accounting/bank_statement_detail.html'

    def get_queryset(self):
        from .models import BankStatement
        return BankStatement.objects.filter(is_active=True).select_related(
            'connection',
        ).prefetch_related('transactions')


class BankTransactionListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/bank_transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 20

    def get_queryset(self):
        from .models import BankTransaction
        qs = BankTransaction.objects.filter(is_active=True).select_related(
            'statement__connection', 'matched_voucher', 'matched_payment',
        )
        match_status = self.request.GET.get('match_status')
        if match_status:
            qs = qs.filter(match_status=match_status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import BankTransaction
        ctx['match_status_choices'] = BankTransaction.MatchStatus.choices
        return ctx


class BankTransactionMatchView(ManagerRequiredMixin, View):
    def post(self, request, pk):
        from .models import BankTransaction
        txn = get_object_or_404(BankTransaction, pk=pk, is_active=True)
        voucher_id = request.POST.get('voucher_id')
        payment_id = request.POST.get('payment_id')
        if voucher_id:
            from .models import Voucher
            txn.matched_voucher = get_object_or_404(Voucher, pk=voucher_id, is_active=True)
            txn.match_status = BankTransaction.MatchStatus.MANUAL_MATCHED
        elif payment_id:
            from .models import Payment
            txn.matched_payment = get_object_or_404(Payment, pk=payment_id, is_active=True)
            txn.match_status = BankTransaction.MatchStatus.MANUAL_MATCHED
        txn.save(update_fields=['matched_voucher', 'matched_payment', 'match_status', 'updated_at'])
        messages.success(request, '거래가 매칭되었습니다.')
        return redirect('accounting:bank_transaction_list')


class BankAutoReconcileView(ManagerRequiredMixin, View):
    """자동 대사 실행 — 금액+날짜 매칭"""
    def post(self, request):
        from .models import BankTransaction, Payment
        unmatched = BankTransaction.objects.filter(
            is_active=True, match_status=BankTransaction.MatchStatus.UNMATCHED,
        )
        matched_count = 0
        for txn in unmatched:
            # Try matching by amount and date
            payment = Payment.objects.filter(
                is_active=True,
                amount=abs(txn.amount),
                payment_date=txn.transaction_date,
            ).first()
            if payment:
                txn.matched_payment = payment
                txn.match_status = BankTransaction.MatchStatus.AUTO_MATCHED
                txn.save(update_fields=['matched_payment', 'match_status', 'updated_at'])
                matched_count += 1
        messages.success(request, f'자동 대사 완료: {matched_count}건 매칭')
        return redirect('accounting:bank_transaction_list')


class BankReconciliationDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/bank_reconciliation_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import BankTransaction, BankConnection
        ctx['connections'] = BankConnection.objects.filter(is_active=True)
        ctx['total_transactions'] = BankTransaction.objects.filter(is_active=True).count()
        ctx['unmatched_count'] = BankTransaction.objects.filter(
            is_active=True, match_status=BankTransaction.MatchStatus.UNMATCHED,
        ).count()
        ctx['auto_matched_count'] = BankTransaction.objects.filter(
            is_active=True, match_status=BankTransaction.MatchStatus.AUTO_MATCHED,
        ).count()
        ctx['manual_matched_count'] = BankTransaction.objects.filter(
            is_active=True, match_status=BankTransaction.MatchStatus.MANUAL_MATCHED,
        ).count()
        return ctx


# ==========================================================================
# 현금영수증 (CashReceipt)
# ==========================================================================
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.urls import reverse


class CashReceiptListView(LoginRequiredMixin, ListView):
    model = CashReceipt
    template_name = 'accounting/cash_receipt_list.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_queryset(self):
        qs = CashReceipt.objects.filter(is_active=True).select_related('partner')
        purpose = self.request.GET.get('purpose')
        status = self.request.GET.get('status')
        partner_id = self.request.GET.get('partner')
        start = self.request.GET.get('start')
        end = self.request.GET.get('end')
        if purpose in (CashReceipt.Purpose.INDIVIDUAL, CashReceipt.Purpose.BUSINESS):
            qs = qs.filter(purpose=purpose)
        if status in (CashReceipt.Status.ISSUED, CashReceipt.Status.CANCELLED):
            qs = qs.filter(status=status)
        if partner_id:
            qs = qs.filter(partner_id=partner_id)
        if start:
            qs = qs.filter(issued_at__date__gte=start)
        if end:
            qs = qs.filter(issued_at__date__lte=end)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['current_purpose'] = self.request.GET.get('purpose', '')
        ctx['current_status'] = self.request.GET.get('status', '')
        ctx['start'] = self.request.GET.get('start', '')
        ctx['end'] = self.request.GET.get('end', '')
        ctx['purpose_choices'] = CashReceipt.Purpose.choices
        ctx['status_choices'] = CashReceipt.Status.choices
        return ctx


class CashReceiptDetailView(LoginRequiredMixin, DetailView):
    model = CashReceipt
    template_name = 'accounting/cash_receipt_detail.html'
    context_object_name = 'receipt'

    def get_queryset(self):
        return CashReceipt.objects.select_related('partner', 'content_type')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        receipt = self.object
        # source_orders M2M 우선, 없으면 라인의 source_order_item 역추적 (레거시)
        m2m_orders = list(
            receipt.source_orders.select_related('partner').order_by('order_number')
        )
        if m2m_orders:
            ctx['linked_orders'] = m2m_orders
        else:
            linked_order_ids = list(
                receipt.items.filter(is_active=True)
                .exclude(source_order_item__isnull=True)
                .values_list('source_order_item__order_id', flat=True)
                .distinct()
            )
            if linked_order_ids:
                ctx['linked_orders'] = list(
                    Order.objects.filter(pk__in=linked_order_ids)
                    .select_related('partner')
                    .order_by('order_number')
                )
            else:
                ctx['linked_orders'] = []
        return ctx


class CashReceiptCreateView(ManagerRequiredMixin, CreateView):
    model = CashReceipt
    form_class = CashReceiptForm
    template_name = 'accounting/cash_receipt_form.html'

    def _parse_order_ids(self):
        """GET source_order (multi-value + comma-separated) + 레거시 order_id 파싱 → sorted unique int list"""
        raw_values = []
        raw_values.extend(self.request.GET.getlist('source_order'))
        legacy = self.request.GET.get('order_id')
        if legacy:
            raw_values.append(legacy)
        ids = []
        seen = set()
        for raw in raw_values:
            if not raw:
                continue
            for token in str(raw).split(','):
                token = token.strip()
                if not token:
                    continue
                try:
                    oid = int(token)
                except ValueError:
                    continue
                if oid in seen:
                    continue
                seen.add(oid)
                ids.append(oid)
        return ids

    def _resolve_source_orders(self):
        """주문 ID 리스트 → Order QuerySet (조회 실패 ID는 조용히 skip)."""
        ids = self._parse_order_ids()
        if not ids:
            return Order.objects.none()
        qs = (
            Order.objects.filter(is_active=True, pk__in=ids)
            .select_related('partner')
            .prefetch_related('items__product')
        )
        # Preserve user-supplied ID order for deterministic UX
        order_map = {o.pk: o for o in qs}
        ordered = [order_map[i] for i in ids if i in order_map]
        return ordered

    def _partner_mismatch(self, orders):
        """거래처가 2종류 이상이면 True (None 은 무시)."""
        partner_ids = {o.partner_id for o in orders if o.partner_id}
        return len(partner_ids) > 1

    def get_initial(self):
        initial = super().get_initial()
        initial['issued_at'] = timezone.localtime().strftime('%Y-%m-%dT%H:%M')
        orders = self._resolve_source_orders()
        if not orders:
            return initial

        supply_sum = Decimal('0')
        vat_sum = Decimal('0')
        for o in orders:
            supply_sum += Decimal(o.total_amount or 0)
            vat_sum += Decimal(o.tax_total or 0)
        initial['supply_amount'] = supply_sum
        initial['vat'] = vat_sum

        # 세법상 수취자 단일 원칙 — 불일치 시 경고는 get_context_data/form_valid 에서 처리
        first_partner = next((o.partner_id for o in orders if o.partner_id), None)
        if first_partner:
            initial['partner'] = first_partner

        order_dates = [o.order_date for o in orders if o.order_date]
        if order_dates:
            # 가장 늦은 주문일 00:00 로 설정 (날짜 정보만 있어 시간은 의미 없음)
            latest = max(order_dates)
            initial['issued_at'] = latest.strftime('%Y-%m-%dT00:00')

        # sales_channel 기준 issuer 추론 — 플랫폼 대행 채널이 하나라도 있으면 PLATFORM 기본값
        platform_name = None
        for o in orders:
            channel = getattr(o, 'sales_channel', '')
            if not channel:
                continue
            config = PlatformFinancialConfig.objects.filter(
                code=channel, is_enabled=True, is_active=True,
            ).first()
            if config and config.cash_receipt_issuer == PlatformFinancialConfig.IssuerType.PLATFORM:
                platform_name = config.name
                break
        if platform_name:
            initial['issuer_type'] = CashReceipt.IssuerType.PLATFORM
            initial['platform_name'] = platform_name
        return initial

    def get_context_data(self, **kwargs):
        from django.forms.models import inlineformset_factory
        from .forms import CashReceiptItemForm
        ctx = super().get_context_data(**kwargs)
        orders = self._resolve_source_orders()
        is_multi = len(orders) > 1

        if self.request.POST:
            ctx['items_formset'] = CashReceiptItemFormSet(self.request.POST, prefix='items')
        else:
            from collections import OrderedDict
            from decimal import ROUND_HALF_UP
            # 단가는 VAT 포함 기준으로 통일 — 주문의 vat_included 여부에 따라 정규화
            def to_inc(v, vat_included):
                v = Decimal(v or 0)
                return v if vat_included else (v * Decimal('1.1')).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

            aggregated = OrderedDict()
            for order in orders:
                items = list(order.items.filter(is_active=True))
                vat_inc = bool(order.vat_included)
                ship_inc = to_inc(order.shipping_charged, vat_inc)
                total_qty = sum(Decimal(it.quantity or 0) for it in items)
                ship_per_unit = Decimal('0')
                if ship_inc and total_qty:
                    ship_per_unit = (ship_inc / total_qty).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                # 주문 전체에 걸쳐 (품목, 단가) 동일 라인은 합산 — 주문↔영수증 연결은
                # CashReceipt.source_orders M2M 으로 별도 저장하므로 라인 병합 가능
                for it in items:
                    name = it.product.name if it.product_id else ''
                    unit_inc = to_inc(it.unit_price, vat_inc)
                    net_unit = int(unit_inc - ship_per_unit)
                    key = (name, net_unit)
                    qty = Decimal(it.quantity or 0)
                    if key in aggregated:
                        aggregated[key]['quantity'] += qty
                        aggregated[key]['source_order_item'] = None
                    else:
                        aggregated[key] = {
                            'name': name,
                            'quantity': qty,
                            'unit_price': net_unit,
                            'source_order_item': it.pk,
                        }
                if ship_inc:
                    key = ('배송비', int(ship_inc))
                    if key in aggregated:
                        aggregated[key]['quantity'] += Decimal('1')
                    else:
                        aggregated[key] = {
                            'name': '배송비',
                            'quantity': Decimal('1'),
                            'unit_price': int(ship_inc),
                            'source_order_item': None,
                        }
            initial_items = []
            for entry in aggregated.values():
                qty = entry['quantity']
                total_inc = Decimal(entry['unit_price']) * qty
                supply = (total_inc / Decimal('1.1')).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                vat = total_inc - supply
                row = {
                    'name': entry['name'],
                    'quantity': qty,
                    'unit_price': entry['unit_price'],
                    'supply_amount': int(supply),
                    'vat': int(vat),
                }
                if entry['source_order_item'] is not None:
                    row['source_order_item'] = entry['source_order_item']
                initial_items.append(row)
            DynamicFS = inlineformset_factory(
                CashReceipt, CashReceiptItem,
                form=CashReceiptItemForm,
                extra=max(len(initial_items), 1),
                can_delete=True,
            )
            ctx['items_formset'] = DynamicFS(
                prefix='items',
                initial=initial_items,
                queryset=CashReceiptItem.objects.none(),
            )

        # 하위 호환 — 단일 주문일 때만 source_order 컨텍스트 유지
        ctx['source_order'] = orders[0] if len(orders) == 1 else None
        ctx['selected_orders'] = orders
        ctx['is_multi_order'] = is_multi
        if is_multi and self._partner_mismatch(orders):
            ctx['partner_mismatch_warning'] = (
                '선택된 주문의 거래처가 서로 다릅니다. 현금영수증은 세법상 단일 거래처에만 발행할 수 있으니 '
                '하나의 거래처로 묶인 주문만 선택하세요.'
            )
        return ctx

    def form_valid(self, form):
        items_formset = CashReceiptItemFormSet(self.request.POST, prefix='items')
        if not items_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, items_formset=items_formset))

        orders = self._resolve_source_orders()
        if len(orders) > 1 and self._partner_mismatch(orders):
            form.add_error(
                None,
                '선택된 주문의 거래처가 일치하지 않습니다. 동일 거래처 주문만 묶어서 발행하세요.',
            )
            return self.render_to_response(self.get_context_data(form=form, items_formset=items_formset))

        # 플랫폼 중복발행 차단 — sales_channel이 PLATFORM 발행 설정이면 자사 발행 차단
        declared_issuer = form.cleaned_data.get('issuer_type')
        if orders and declared_issuer != CashReceipt.IssuerType.PLATFORM:
            blocking_channels = set()
            blocking_name = None
            for o in orders:
                channel = getattr(o, 'sales_channel', '')
                if not channel:
                    continue
                config = PlatformFinancialConfig.objects.filter(
                    code=channel, is_enabled=True, is_active=True,
                ).first()
                if config and config.cash_receipt_issuer == PlatformFinancialConfig.IssuerType.PLATFORM:
                    blocking_channels.add(channel)
                    blocking_name = config.name
            if blocking_channels:
                form.add_error(
                    None,
                    f'{blocking_name or ", ".join(blocking_channels)} 채널은 현금영수증을 '
                    '플랫폼이 대행 발행하는 설정입니다. 중복발행을 차단합니다. '
                    '발행주체를 "플랫폼대행"으로 변경하거나 해당 주문을 제외하세요.',
                )
                return self.render_to_response(self.get_context_data(form=form, items_formset=items_formset))

        legacy_payment_id = self.request.GET.get('payment_id')
        legacy_voucher_id = self.request.GET.get('voucher_id')

        with transaction.atomic():
            form.instance.created_by = self.request.user
            if len(orders) == 1:
                form.instance.content_type = ContentType.objects.get(app_label='sales', model='order')
                form.instance.object_id = orders[0].pk
            elif len(orders) > 1:
                # 다중 주문 — content_type/object_id 는 비워두고 line 단위 source_order_item 으로 추적
                form.instance.content_type = None
                form.instance.object_id = None
            elif legacy_payment_id:
                form.instance.content_type = ContentType.objects.get(app_label='accounting', model='payment')
                form.instance.object_id = int(legacy_payment_id)
            elif legacy_voucher_id:
                form.instance.content_type = ContentType.objects.get(app_label='accounting', model='voucher')
                form.instance.object_id = int(legacy_voucher_id)

            self.object = form.save()

            # 연결된 주문을 M2M 에 저장 (단일/다중 모두)
            if orders:
                self.object.source_orders.set(orders)

            items_formset.instance = self.object
            saved_items = items_formset.save(commit=False)
            for item in saved_items:
                if item.created_by_id is None:
                    item.created_by = self.request.user
                item.save()
            for obj in items_formset.deleted_objects:
                obj.delete()

            if self.object.items.filter(is_active=True).exists():
                self.object.recalculate_totals()

            # TaxInvoice 중복 경고는 전자세금계산서 연동(홈택스) 완성 후 재도입

            if len(orders) > 1:
                messages.success(
                    self.request,
                    f'현금영수증({self.object.receipt_number})이 주문 {len(orders)}건에 대해 발행되었습니다.',
                )
            else:
                messages.success(
                    self.request,
                    f'현금영수증({self.object.receipt_number})이 발행되었습니다.',
                )
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse('accounting:cash_receipt_detail', kwargs={'pk': self.object.pk})


class CashReceiptOrderLookupView(ManagerRequiredMixin, View):
    """AJAX: Order 검색 → 거래처/공급가/부가세/라인 반환. `order_ids=1,2,3` 로 벌크 조회 지원."""
    def get(self, request):
        import re
        from decimal import Decimal
        from django.http import JsonResponse
        oid = request.GET.get('order_id')
        order_ids_param = request.GET.get('order_ids')
        raw = (request.GET.get('search') or request.GET.get('q') or '').strip()

        # 다중 주문번호 검색 모드: 쉼표/세미콜론/공백으로 구분된 2개 이상 토큰
        tokens = [t.strip() for t in re.split(r'[,;\s]+', raw) if t.strip()]
        seen = set()
        unique_tokens = []
        for t in tokens:
            if t not in seen:
                seen.add(t)
                unique_tokens.append(t)

        if len(unique_tokens) >= 2:
            qs = (
                Order.objects.filter(is_active=True, order_number__in=unique_tokens)
                .select_related('partner')
            )
            order_map = {o.order_number: o for o in qs}
            ordered = [order_map[tok] for tok in unique_tokens if tok in order_map]
            not_found = [tok for tok in unique_tokens if tok not in order_map]
            supply_total = sum(Decimal(o.total_amount or 0) for o in ordered)
            vat_total = sum(Decimal(o.tax_total or 0) for o in ordered)
            partner_ids = {o.partner_id for o in ordered if o.partner_id}
            return JsonResponse({
                'orders': [
                    {
                        'id': o.pk,
                        'order_number': o.order_number,
                        'partner_id': o.partner_id,
                        'partner_name': o.partner.name if o.partner_id else '',
                        'supply_amount': int(o.total_amount or 0),
                        'vat_amount': int(o.tax_total or 0),
                        'total_amount': int(o.grand_total or 0),
                        'order_date': o.order_date.isoformat() if o.order_date else '',
                        'status': o.status,
                    }
                    for o in ordered
                ],
                'supply_total': str(supply_total),
                'vat_total': str(vat_total),
                'grand_total': str(supply_total + vat_total),
                'partner_mismatch': len(partner_ids) > 1,
                'not_found': not_found,
            })

        q = raw  # 단일 토큰이면 기존 icontains 로직 재사용

        if order_ids_param:
            ids = []
            for token in order_ids_param.split(','):
                token = token.strip()
                if not token:
                    continue
                try:
                    ids.append(int(token))
                except ValueError:
                    continue
            if not ids:
                return JsonResponse({'orders': [], 'supply_total': 0, 'vat_total': 0, 'grand_total': 0})
            qs = (
                Order.objects.filter(is_active=True, pk__in=ids)
                .select_related('partner')
            )
            order_map = {o.pk: o for o in qs}
            ordered = [order_map[i] for i in ids if i in order_map]
            supply_total = sum(int(o.total_amount or 0) for o in ordered)
            vat_total = sum(int(o.tax_total or 0) for o in ordered)
            partner_ids = {o.partner_id for o in ordered if o.partner_id}
            return JsonResponse({
                'orders': [
                    {
                        'id': o.pk,
                        'order_number': o.order_number,
                        'partner_id': o.partner_id,
                        'partner_name': o.partner.name if o.partner_id else '',
                        'order_date': o.order_date.isoformat() if o.order_date else '',
                        'total_amount': int(o.total_amount or 0),
                        'tax_total': int(o.tax_total or 0),
                    }
                    for o in ordered
                ],
                'supply_total': supply_total,
                'vat_total': vat_total,
                'grand_total': supply_total + vat_total,
                'partner_mismatch': len(partner_ids) > 1,
            })

        if oid:
            try:
                order = Order.objects.filter(is_active=True).select_related('partner').prefetch_related('items__product').get(pk=int(oid))
            except (ValueError, Order.DoesNotExist):
                return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)
            return JsonResponse({
                'id': order.pk,
                'order_number': order.order_number,
                'partner_id': order.partner_id,
                'partner_name': order.partner.name if order.partner_id else '',
                'total_amount': int(order.total_amount or 0),
                'tax_total': int(order.tax_total or 0),
                'items': [
                    {
                        'source_order_item': it.pk,
                        'name': it.product.name if it.product_id else '',
                        'quantity': int(it.quantity or 0),
                        'unit_price': int(it.unit_price or 0),
                        'supply_amount': int(it.amount or 0),
                        'vat': int(it.tax_amount or 0),
                    }
                    for it in order.items.filter(is_active=True)
                ],
            })
        qs = Order.objects.filter(is_active=True).select_related('partner')
        if q:
            qs = qs.filter(Q(order_number__icontains=q) | Q(partner__name__icontains=q))
        qs = qs.order_by('-order_date', '-pk')[:20]
        return JsonResponse({
            'results': [
                {
                    'id': o.pk,
                    'order_number': o.order_number,
                    'partner_name': o.partner.name if o.partner_id else '',
                    'order_date': o.order_date.isoformat() if o.order_date else '',
                    'total_amount': int(o.total_amount or 0),
                }
                for o in qs
            ],
        })


class CashReceiptCancelView(ManagerRequiredMixin, View):
    def get(self, request, pk):
        receipt = get_object_or_404(CashReceipt, pk=pk, is_active=True)
        if receipt.status == CashReceipt.Status.CANCELLED:
            messages.error(request, '이미 취소된 현금영수증입니다.')
            return redirect('accounting:cash_receipt_detail', pk=pk)
        form = CashReceiptCancelForm()
        return self._render(request, receipt, form)

    def post(self, request, pk):
        receipt = get_object_or_404(CashReceipt, pk=pk, is_active=True)
        if receipt.status == CashReceipt.Status.CANCELLED:
            messages.error(request, '이미 취소된 현금영수증입니다.')
            return redirect('accounting:cash_receipt_detail', pk=pk)
        form = CashReceiptCancelForm(request.POST)
        if not form.is_valid():
            return self._render(request, receipt, form)
        with transaction.atomic():
            receipt.status = CashReceipt.Status.CANCELLED
            receipt.cancelled_at = timezone.now()
            receipt.cancel_reason = form.cleaned_data['cancel_reason']
            receipt.is_active = False
            receipt.save(update_fields=['status', 'cancelled_at', 'cancel_reason', 'is_active', 'updated_at'])
        messages.success(request, f'현금영수증({receipt.receipt_number})이 취소되었습니다.')
        return redirect('accounting:cash_receipt_list')

    def _render(self, request, receipt, form):
        from django.shortcuts import render
        return render(request, 'accounting/cash_receipt_cancel.html', {
            'receipt': receipt, 'form': form,
        })


class CashReceiptMonthlyReportView(ManagerRequiredMixin, TemplateView):
    template_name = 'accounting/cash_receipt_monthly_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = safe_int(self.request.GET.get('year'), date.today().year)
        ctx['year'] = year
        ctx['years'] = list(range(date.today().year, date.today().year - 5, -1))

        months = []
        for m in range(1, 13):
            qs = CashReceipt.objects.filter(
                issued_at__year=year, issued_at__month=m,
                status=CashReceipt.Status.ISSUED,
            )
            individual = qs.filter(purpose=CashReceipt.Purpose.INDIVIDUAL).aggregate(
                cnt=Count('id'), total=Sum('total_amount'),
            )
            business = qs.filter(purpose=CashReceipt.Purpose.BUSINESS).aggregate(
                cnt=Count('id'), total=Sum('total_amount'),
            )
            months.append({
                'month': m,
                'individual_count': individual['cnt'] or 0,
                'individual_total': individual['total'] or 0,
                'business_count': business['cnt'] or 0,
                'business_total': business['total'] or 0,
            })
        ctx['months'] = months
        ctx['annual_individual'] = sum(m['individual_total'] for m in months)
        ctx['annual_business'] = sum(m['business_total'] for m in months)
        ctx['annual_total'] = ctx['annual_individual'] + ctx['annual_business']
        return ctx


# ── 플랫폼 재무설정 ────────────────────────────────────────────


class PlatformFinancialConfigListView(ManagerRequiredMixin, ListView):
    model = PlatformFinancialConfig
    template_name = 'accounting/platform_config_list.html'
    context_object_name = 'configs'
    paginate_by = 30

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).order_by('code')


class PlatformFinancialConfigCreateView(ManagerRequiredMixin, CreateView):
    model = PlatformFinancialConfig
    form_class = PlatformFinancialConfigForm
    template_name = 'accounting/platform_config_form.html'
    success_url = reverse_lazy('accounting:platform_config_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class PlatformFinancialConfigUpdateView(ManagerRequiredMixin, UpdateView):
    model = PlatformFinancialConfig
    form_class = PlatformFinancialConfigForm
    template_name = 'accounting/platform_config_form.html'
    success_url = reverse_lazy('accounting:platform_config_list')


class PlatformFinancialConfigDetailView(ManagerRequiredMixin, DetailView):
    model = PlatformFinancialConfig
    template_name = 'accounting/platform_config_detail.html'
    context_object_name = 'config'


class PlatformFinancialConfigDeleteView(ManagerRequiredMixin, View):
    """소프트 삭제 (is_active=False)."""
    template_name = 'accounting/platform_config_confirm_delete.html'

    def get(self, request, pk):
        config = get_object_or_404(PlatformFinancialConfig, pk=pk, is_active=True)
        from django.shortcuts import render
        return render(request, self.template_name, {'config': config})

    def post(self, request, pk):
        config = get_object_or_404(PlatformFinancialConfig, pk=pk, is_active=True)
        config.soft_delete()
        messages.success(request, f'플랫폼 설정 "{config.name}"이(가) 삭제되었습니다.')
        return redirect('accounting:platform_config_list')


class AdvanceReceivedListView(ManagerRequiredMixin, ListView):
    model = AdvanceReceived
    template_name = 'accounting/advance_received_list.html'
    context_object_name = 'advances'
    paginate_by = 20

    def get_queryset(self):
        qs = AdvanceReceived.objects.filter(is_active=True).select_related(
            'partner', 'customer', 'applied_to_order',
        )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models_advance import AdvanceStatus
        ctx['status_choices'] = AdvanceStatus.choices
        ctx['current_status'] = self.request.GET.get('status', '')
        return ctx


class AdvanceReceivedDetailView(ManagerRequiredMixin, DetailView):
    model = AdvanceReceived
    template_name = 'accounting/advance_received_detail.html'
    context_object_name = 'advance'

    def get_queryset(self):
        return AdvanceReceived.objects.filter(is_active=True).select_related(
            'partner', 'customer', 'received_voucher', 'applied_to_order',
        )


class AdvancePaidListView(ManagerRequiredMixin, ListView):
    model = AdvancePaid
    template_name = 'accounting/advance_paid_list.html'
    context_object_name = 'advances'
    paginate_by = 20

    def get_queryset(self):
        qs = AdvancePaid.objects.filter(is_active=True).select_related(
            'partner', 'applied_to_po',
        )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models_advance import AdvanceStatus
        ctx['status_choices'] = AdvanceStatus.choices
        ctx['current_status'] = self.request.GET.get('status', '')
        return ctx


class AdvancePaidDetailView(ManagerRequiredMixin, DetailView):
    model = AdvancePaid
    template_name = 'accounting/advance_paid_detail.html'
    context_object_name = 'advance'

    def get_queryset(self):
        return AdvancePaid.objects.filter(is_active=True).select_related(
            'partner', 'paid_voucher', 'applied_to_po',
        )


# ── 대손충당금 ────────────────────────────────────────────


class BadDebtAllowanceListView(ManagerRequiredMixin, ListView):
    model = BadDebtAllowance
    template_name = 'accounting/bad_debt_list.html'
    context_object_name = 'allowances'
    paginate_by = 30

    def get_queryset(self):
        qs = BadDebtAllowance.objects.filter(is_active=True).select_related(
            'receivable__partner', 'voucher',
        )
        bucket = self.request.GET.get('bucket', '')
        if bucket:
            qs = qs.filter(aging_bucket=bucket)
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(receivable__partner__name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['bucket_choices'] = AgingBucket.choices
        ctx['current_bucket'] = self.request.GET.get('bucket', '')
        from django.db.models import Sum
        ctx['total_allowance'] = (
            BadDebtAllowance.objects.filter(is_active=True)
            .aggregate(total=Sum('allowance_amount'))['total'] or 0
        )
        return ctx


# ── 카드매출전표 ────────────────────────────────────────


class CardSalesSlipListView(ManagerRequiredMixin, ListView):
    template_name = 'accounting/card_sales_slip_list.html'
    context_object_name = 'slips'
    paginate_by = 30

    def get_queryset(self):
        from .models_cardslip import CardSalesSlip
        qs = CardSalesSlip.objects.filter(is_active=True).select_related(
            'order', 'partner', 'card_transaction',
        )
        status = self.request.GET.get('status', '')
        if status:
            qs = qs.filter(status=status)
        q = self.request.GET.get('q', '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(slip_number__icontains=q)
                | Q(approval_code__icontains=q)
                | Q(merchant_number__icontains=q)
            )
        return qs.order_by('-approved_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models_cardslip import CardSalesSlip
        ctx['status_choices'] = CardSalesSlip.Status.choices
        ctx['current_status'] = self.request.GET.get('status', '')
        ctx['q'] = self.request.GET.get('q', '')
        return ctx


class CardSalesSlipDetailView(ManagerRequiredMixin, DetailView):
    template_name = 'accounting/card_sales_slip_detail.html'
    context_object_name = 'slip'

    def get_queryset(self):
        from .models_cardslip import CardSalesSlip
        return CardSalesSlip.objects.filter(is_active=True).select_related(
            'order', 'partner', 'card_transaction',
        )


class CardSalesSlipCreateView(ManagerRequiredMixin, CreateView):
    template_name = 'accounting/card_sales_slip_form.html'

    def get_form_class(self):
        from django import forms
        from .models_cardslip import CardSalesSlip
        from apps.core.forms import BaseForm

        class CardSalesSlipForm(BaseForm):
            class Meta:
                model = CardSalesSlip
                fields = [
                    'approved_at', 'approval_code', 'card_brand',
                    'card_number_masked', 'merchant_number',
                    'supply_amount', 'vat',
                    'order', 'partner', 'card_transaction',
                ]
                widgets = {
                    'approved_at': forms.DateTimeInput(
                        attrs={'type': 'datetime-local', 'class': 'form-input'},
                    ),
                }

        return CardSalesSlipForm

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        from django.contrib import messages
        messages.success(self.request, '카드매출전표가 등록되었습니다.')
        return super().form_valid(form)

    def get_success_url(self):
        from django.urls import reverse
        return reverse('accounting:cardslip_detail', kwargs={'pk': self.object.pk})
