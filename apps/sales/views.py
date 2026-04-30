import tablib
import logging
from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Q, F, ExpressionWrapper, DecimalField, Case, When, Value, Exists, OuterRef
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView, TemplateView

from apps.core.mixins import ManagerRequiredMixin
from apps.inventory.models import Product
from .models import (
    Partner, Customer, Order, OrderItem,
    Quotation, QuotationItem, Shipment, ShipmentItem,
    ShippingCarrier, ShipmentTracking, PriceRule,
    CustomerTier, SalesTarget,
    SalesLead, LeadActivity, CustomerSatisfaction,
)
from .forms import (
    PartnerForm, CustomerForm, CustomerPurchaseFormSet,
    OrderForm, OrderItemFormSet, ReturnOrderForm,
    QuotationForm, QuotationItemFormSet,
    ShippingCarrierForm, PriceRuleForm,
    CustomerTierForm, SalesTargetForm,
    SalesLeadForm, LeadActivityForm, CustomerSatisfactionForm,
)


def _product_units_json():
    return {str(p.pk): p.unit or '' for p in Product.objects.filter(is_active=True)}


def _product_costs_json(stage='QUOTATION'):
    """제품별 원가 JSON — CostBasisConfig 설정에 따라 단계별 원가 반환.

    설정이 없으면 기존 로직(생산원가 → BOM → 제품원가)으로 폴백.
    """
    from apps.inventory.models import CostBasisConfig

    products = list(Product.objects.filter(is_active=True))
    product_ids = [p.pk for p in products]

    # CostBasisConfig 설정이 있으면 설정 기반 조회
    config = CostBasisConfig.objects.filter(stage=stage, is_active=True).first()
    if config:
        return CostBasisConfig.get_costs_bulk(product_ids, stage)

    # 설정 없으면 기존 로직 (하위 호환)
    from django.db.models import Max
    from apps.production.models import BOM, ProductionRecord

    # 1순위: 제품별 최신 생산실적의 unit_cost (배치 조회)
    latest_costs = dict(
        ProductionRecord.objects.filter(
            work_order__production_plan__product_id__in=product_ids,
            is_active=True,
            unit_cost__gt=0,
        ).values(
            'work_order__production_plan__product_id',
        ).annotate(
            latest_cost=Max('unit_cost'),
        ).values_list(
            'work_order__production_plan__product_id', 'latest_cost',
        )
    )

    # 2순위: BOM 원가 (배치 조회 - 기본BOM 우선)
    bom_costs = {}
    boms = BOM.objects.filter(
        product_id__in=product_ids, is_active=True,
    ).prefetch_related('items__material').order_by(
        'product_id', '-is_default', 'pk',
    )
    for bom in boms:
        pid = bom.product_id
        if pid not in bom_costs:
            cost = bom.total_material_cost
            if cost:
                bom_costs[pid] = int(cost)

    # 결과 조합
    costs = {}
    for p in products:
        if p.pk in latest_costs:
            costs[str(p.pk)] = int(latest_costs[p.pk])
        elif p.pk in bom_costs:
            costs[str(p.pk)] = bom_costs[p.pk]
        else:
            costs[str(p.pk)] = int(p.cost_price or 0)
    return costs
from .commission import CommissionRate, CommissionRecord
from .commission_forms import CommissionRateForm, CommissionRecordForm
from .resources import PartnerResource, CustomerResource, CommissionRateResource


class PartnerListView(LoginRequiredMixin, ListView):
    model = Partner
    template_name = 'sales/partner_list.html'
    context_object_name = 'partners'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).annotate(
            order_count=Count('orders'),
        ).order_by('name')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class PartnerCreateView(ManagerRequiredMixin, CreateView):
    model = Partner
    form_class = PartnerForm
    template_name = 'sales/partner_form.html'
    success_url = reverse_lazy('sales:partner_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .commission_forms import CommissionRateInlineFormSet
        if self.request.POST:
            ctx['commission_formset'] = CommissionRateInlineFormSet(
                self.request.POST, prefix='comm',
            )
        else:
            ctx['commission_formset'] = CommissionRateInlineFormSet(
                prefix='comm',
            )
        ctx['next_codes'] = {
            'CUSTOMER': Partner.generate_next_code('CUSTOMER'),
            'SUPPLIER': Partner.generate_next_code('SUPPLIER'),
            'BOTH': Partner.generate_next_code('BOTH'),
        }
        return ctx

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        ctx = self.get_context_data()
        formset = ctx['commission_formset']
        if formset.is_valid():
            resp = super().form_valid(form)
            formset.instance = self.object
            formset.save()
            return resp
        return self.form_invalid(form)


class PartnerUpdateView(ManagerRequiredMixin, UpdateView):
    model = Partner
    form_class = PartnerForm
    template_name = 'sales/partner_form.html'
    success_url = reverse_lazy('sales:partner_list')
    slug_field = 'code'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .commission_forms import CommissionRateInlineFormSet
        if self.request.POST:
            ctx['commission_formset'] = CommissionRateInlineFormSet(
                self.request.POST, instance=self.object,
                prefix='comm',
            )
        else:
            ctx['commission_formset'] = CommissionRateInlineFormSet(
                instance=self.object, prefix='comm',
            )
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['commission_formset']
        if formset.is_valid():
            resp = super().form_valid(form)
            formset.save()
            return resp
        return self.form_invalid(form)


class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'sales/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).prefetch_related('purchases__product')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(name__icontains=q) | Q(phone__icontains=q)
                | Q(purchases__serial_number__icontains=q)
            ).distinct()
        return qs


class CustomerCreateView(ManagerRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'sales/customer_form.html'
    success_url = reverse_lazy('sales:customer_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = CustomerPurchaseFormSet(self.request.POST)
        else:
            ctx['formset'] = CustomerPurchaseFormSet()
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            formset.instance = self.object
            formset.save()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'sales/customer_detail.html'
    slug_field = 'code'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['purchases'] = self.object.purchases.filter(
            is_active=True,
        ).select_related('product')
        ctx['orders'] = self.object.orders.filter(
            is_active=True,
        ).select_related(
            'partner', 'customer',
        )
        ctx['service_requests'] = (
            self.object.service_requests
            .filter(is_active=True)
            .select_related('product')
        )
        from apps.warranty.models import ProductRegistration
        ctx['registrations'] = ProductRegistration.objects.filter(
            customer=self.object, is_active=True,
        ).select_related('product')
        return ctx


class CustomerUpdateView(ManagerRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'sales/customer_form.html'
    success_url = reverse_lazy('sales:customer_list')
    slug_field = 'code'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = CustomerPurchaseFormSet(
                self.request.POST, instance=self.object,
            )
        else:
            ctx['formset'] = CustomerPurchaseFormSet(instance=self.object)
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class OrderListView(LoginRequiredMixin, ListView):
    model = Order
    template_name = 'sales/order_list.html'
    context_object_name = 'orders'
    paginate_by = 20

    def get_queryset(self):
        from django.contrib.contenttypes.models import ContentType
        from apps.accounting.models import CashReceipt, Payment, PlatformFinancialConfig
        from django.db.models import Subquery
        ct = ContentType.objects.get_for_model(Order)
        # 현금영수증 3경로 — 직접 FK / source_orders M2M / 라인의 source_order_item
        direct_sq = CashReceipt.objects.filter(
            content_type=ct, object_id=OuterRef('pk'),
            is_active=True, status='ISSUED',
        )
        m2m_sq = CashReceipt.objects.filter(
            source_orders=OuterRef('pk'),
            is_active=True, status='ISSUED',
        )
        line_sq = CashReceipt.objects.filter(
            items__source_order_item__order=OuterRef('pk'),
            items__is_active=True,
            is_active=True, status='ISSUED',
        )
        # 카드결제 — Payment(method=CARD) → AR → Order
        card_sq = Payment.objects.filter(
            is_active=True,
            payment_method=Payment.PaymentMethod.CARD,
            receivable__order=OuterRef('pk'),
        )
        # 채널 PlatformFinancialConfig 발행주체 어노테이션 — PLATFORM 자동발행 표시용
        platform_cr_sq = PlatformFinancialConfig.objects.filter(
            code=OuterRef('sales_channel'),
            is_enabled=True, is_active=True,
        ).values('cash_receipt_issuer')[:1]
        platform_card_sq = PlatformFinancialConfig.objects.filter(
            code=OuterRef('sales_channel'),
            is_enabled=True, is_active=True,
        ).values('card_receipt_issuer')[:1]
        platform_name_sq = PlatformFinancialConfig.objects.filter(
            code=OuterRef('sales_channel'),
            is_enabled=True, is_active=True,
        ).values('name')[:1]
        qs = super().get_queryset().filter(is_active=True).select_related(
            'partner', 'customer',
        ).prefetch_related('items', 'source_quotation').annotate(
            has_cash_receipt_ann=Exists(direct_sq) | Exists(m2m_sq) | Exists(line_sq),
            has_card_payment_ann=Exists(card_sq),
            channel_cr_issuer=Subquery(platform_cr_sq),
            channel_card_issuer=Subquery(platform_card_sq),
            channel_name=Subquery(platform_name_sq),
        )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        order_type = self.request.GET.get('order_type')
        if order_type:
            qs = qs.filter(order_type=order_type)
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(order_number__icontains=q)
                | Q(partner__name__icontains=q)
                | Q(customer__name__icontains=q)
                | Q(marketplace_orders__store_order_id__icontains=q)
            ).distinct()
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        return context


class OrderCreateView(ManagerRequiredMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'sales/order_form.html'
    success_url = reverse_lazy('sales:order_list')

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['order_number'] = generate_document_number(
            Order, 'order_number', 'ORD',
        )
        from apps.accounting.models import BankAccount
        default_bank = BankAccount.objects.filter(
            is_active=True, is_default=True,
        ).first()
        if default_bank:
            initial['bank_account'] = default_bank.pk
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = OrderItemFormSet(self.request.POST)
        else:
            ctx['formset'] = OrderItemFormSet()
        ctx['product_units_json'] = _product_units_json()
        ctx['product_costs_json'] = _product_costs_json(stage='ORDER')
        ctx['quote_notes'] = ''
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            # 입금계좌 미설정 + 거래처 기본계좌 있으면 자동 설정
            if not self.object.bank_account_id and self.object.partner_id:
                partner = self.object.partner
                if partner.default_bank_account_id:
                    self.object.bank_account_id = partner.default_bank_account_id
            self.object.save()
            formset.instance = self.object
            formset.save()
            self.object.update_total()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class OrderDetailView(LoginRequiredMixin, DetailView):
    model = Order
    template_name = 'sales/order_detail.html'
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'partner', 'customer', 'bank_account',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        items = self.object.items.select_related('product').all()
        ctx['items'] = items
        ctx['status_actions'] = self._build_status_actions()
        total_cost = sum(item.total_cost for item in items)
        total_amount = int(self.object.total_amount)
        shipping_cost = int(self.object.shipping_cost or 0)
        total_profit = total_amount - int(total_cost) - shipping_cost
        ctx['total_cost'] = total_cost
        ctx['shipping_cost'] = shipping_cost
        ctx['total_profit'] = total_profit
        ctx['total_profit_rate'] = (
            round(total_profit / total_amount * 100, 1)
            if total_amount else 0
        )
        # 연결된 현금영수증 — source_orders M2M(신규) / 직접 FK(단일) / 라인 FK(레거시)
        from django.contrib.contenttypes.models import ContentType
        from apps.accounting.models import CashReceipt
        ct = ContentType.objects.get_for_model(Order)
        cash_receipts = CashReceipt.objects.filter(is_active=True).filter(
            Q(source_orders=self.object)
            | Q(content_type=ct, object_id=self.object.pk)
            | Q(items__source_order_item__order=self.object, items__is_active=True)
        ).distinct().order_by('-issued_at')
        ctx['cash_receipts'] = cash_receipts
        ctx['active_cash_receipt'] = cash_receipts.filter(status='ISSUED').first()
        ctx['can_issue_cash_receipt'] = (
            self.object.status not in ('CANCELLED',)
            and self.object.order_type not in ('RETURN',)
        )
        # 증빙 발행현황 통합 — TaxInvoice / CardSalesSlip
        from apps.accounting.models import (
            TaxInvoice, CardSalesSlip, PlatformFinancialConfig,
        )
        ctx['active_tax_invoice'] = TaxInvoice.objects.filter(
            order=self.object, is_active=True,
            invoice_type=TaxInvoice.InvoiceType.SALES,
        ).order_by('-issue_date', '-pk').first()
        ctx['active_card_slip'] = CardSalesSlip.objects.filter(
            order=self.object, is_active=True,
        ).order_by('-approved_at', '-pk').first()
        # 채널 PlatformFinancialConfig — PLATFORM 대행분 자동발행 표시용
        channel_code = (self.object.sales_channel or '').strip()
        ctx['platform_config'] = PlatformFinancialConfig.objects.filter(
            code=channel_code, is_enabled=True, is_active=True,
        ).first() if channel_code else None
        # 연결된 반품/교환 주문
        ctx['return_orders'] = Order.objects.filter(
            original_order=self.object, is_active=True,
        ).order_by('-created_at')
        # 원본 주문 (반품/교환인 경우)
        if self.object.original_order:
            ctx['original_order'] = self.object.original_order
        # 거래처 수수료 (입금 시 차감용)
        if not self.object.is_paid and self.object.partner:
            partner = self.object.partner
            supply = int(self.object.total_amount or 0)
            grand = int(self.object.grand_total or 0)
            comm = partner.calculate_commission(supply, total_amount=grand)
            if comm > 0:
                rate = partner.total_commission_rate
                fixed = partner.total_fixed_commission
                ctx['partner_commission_rate'] = float(rate)
                ctx['partner_fixed_commission'] = fixed
                ctx['partner_commission'] = comm
                ctx['auto_deposit'] = grand - comm
        return ctx

    def _build_status_actions(self):
        order = self.object
        allowed = Order.STATUS_TRANSITIONS.get(order.status, [])
        action_map = {
            'CONFIRMED': {
                'label': '주문 확정',
                'css': 'btn-primary',
                'icon': '<svg class="w-4 h-4 inline mr-1 -mt-0.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/></svg>',
                'confirm': '주문을 확정하시겠습니까?',
            },
            'SHIPPED': {
                'label': '출고 완료',
                'css': 'btn-primary',
                'icon': '<svg class="w-4 h-4 inline mr-1 -mt-0.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 8h14M5 8a2 2 0 110-4h14a2 2 0 110 4M5 8v10a2 2 0 002 2h10a2 2 0 002-2V8"/></svg>',
                'confirm': '출고 처리하시겠습니까?',
            },
            'DELIVERED': {
                'label': '배송 완료',
                'css': 'btn btn-primary bg-green-600 hover:bg-green-700',
                'icon': '<svg class="w-4 h-4 inline mr-1 -mt-0.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z"/></svg>',
                'confirm': '배송 완료 처리하시겠습니까?',
            },
            'CLOSED': {
                'label': '주문 종결',
                'css': 'btn btn-sm bg-gray-600 hover:bg-gray-700 text-white',
                'icon': '<svg class="w-4 h-4 inline mr-1 -mt-0.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z"/></svg>',
                'confirm': '주문을 종결하시겠습니까?',
            },
            'CANCELLED': {
                'label': '주문 취소',
                'css': 'btn btn-secondary text-red-600 hover:bg-red-50',
                'icon': '<svg class="w-4 h-4 inline mr-1 -mt-0.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M6 18L18 6M6 6l12 12"/></svg>',
                'confirm': '주문을 취소하시겠습니까? 이 작업은 되돌릴 수 없습니다.',
            },
        }
        actions = []
        for status in allowed:
            if status in action_map:
                actions.append({**action_map[status], 'value': status})
        return actions


class OrderUpdateView(ManagerRequiredMixin, UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'sales/order_form.html'
    success_url = reverse_lazy('sales:order_list')
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = OrderItemFormSet(self.request.POST, instance=self.object)
        else:
            ctx['formset'] = OrderItemFormSet(instance=self.object)
        ctx['product_units_json'] = _product_units_json()
        ctx['product_costs_json'] = _product_costs_json(stage='ORDER')
        ctx['is_settled'] = self.object.is_settled
        # 견적서 비고 (있으면)
        quote = self.object.source_quotation.first()
        ctx['quote_notes'] = quote.notes if quote else ''
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            # vat_included 변경 시 모든 항목 VAT 재계산
            for item in self.object.items.all():
                item.save()
            self.object.update_total()
            return super().form_valid(form)
        return self.form_invalid(form)


class OrderStatusChangeView(ManagerRequiredMixin, View):
    """주문 상태 전환 (POST only)"""

    def post(self, request, slug):
        order = get_object_or_404(Order, order_number=slug, is_active=True)
        new_status = request.POST.get('status')

        allowed = Order.STATUS_TRANSITIONS.get(order.status, [])
        if new_status not in allowed:
            messages.error(request, '허용되지 않는 상태 전환입니다.')
            return redirect('sales:order_detail', slug=order.order_number)

        order.status = new_status
        try:
            order.save(update_fields=['status', 'updated_at'])
        except Exception as e:
            messages.error(request, str(e))
            return redirect('sales:order_detail', slug=order.order_number)
        messages.success(request, f'주문 상태가 "{order.get_status_display()}"(으)로 변경되었습니다.')
        return redirect('sales:order_detail', slug=order.order_number)


class OrderPaymentView(ManagerRequiredMixin, View):
    """주문 입금 처리 — 전액 입금 + 수수료 자동 출금"""

    def post(self, request, slug):
        order = get_object_or_404(Order, order_number=slug, is_active=True)

        if order.is_paid:
            messages.warning(request, '이미 입금 처리된 주문입니다.')
            return redirect('sales:order_detail', slug=order.order_number)

        from apps.sales.signals import _auto_create_payment
        try:
            # _auto_create_payment 내부에서 commission + close 처리
            _auto_create_payment(order)
            messages.success(
                request,
                f'{order.order_number} 입금 처리 완료'
                f' ({int(order.grand_total):,}원)',
            )
        except Exception as e:
            messages.error(request, f'입금 처리 실패: {e}')

        return redirect('sales:order_detail', slug=order.order_number)


class ReturnOrderCreateView(ManagerRequiredMixin, CreateView):
    """원본 주문 기반 반품 주문 생성"""
    model = Order
    form_class = ReturnOrderForm
    template_name = 'sales/order_return_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.original_order = get_object_or_404(
            Order, order_number=kwargs['slug'], is_active=True,
        )
        if self.original_order.order_type in ('RETURN', 'EXCHANGE'):
            messages.error(request, '반품/교환 주문에서는 반품을 생성할 수 없습니다.')
            return redirect('sales:order_detail', slug=self.original_order.order_number)
        if self.original_order.status not in ('SHIPPED', 'DELIVERED', 'CLOSED'):
            messages.error(request, '출고 이후 상태의 주문만 반품할 수 있습니다.')
            return redirect('sales:order_detail', slug=self.original_order.order_number)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['original_order'] = self.original_order
        ctx['order_type'] = 'RETURN'
        return ctx

    def form_valid(self, form):
        from apps.core.utils import generate_document_number
        with transaction.atomic():
            order = form.save(commit=False)
            order.order_type = Order.OrderType.RETURN
            order.original_order = self.original_order
            order.partner = self.original_order.partner
            order.customer = self.original_order.customer
            order.order_date = date.today()
            order.order_number = generate_document_number(Order, 'order_number', 'ORD')
            order.vat_included = self.original_order.vat_included
            order.created_by = self.request.user
            order.save()

            for item in self.original_order.items.filter(is_active=True):
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    quantity=item.quantity,
                    cost_price=item.cost_price,
                    unit_price=item.unit_price,
                    discount_rate=item.discount_rate,
                    discount_amount=item.discount_amount,
                    created_by=self.request.user,
                )
            order.update_total()
        messages.success(self.request, f'반품주문 {order.order_number}이(가) 생성되었습니다.')
        return redirect('sales:order_detail', slug=order.order_number)


class ExchangeOrderCreateView(ManagerRequiredMixin, CreateView):
    """원본 주문 기반 교환 주문 생성"""
    model = Order
    form_class = ReturnOrderForm
    template_name = 'sales/order_return_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.original_order = get_object_or_404(
            Order, order_number=kwargs['slug'], is_active=True,
        )
        if self.original_order.order_type in ('RETURN', 'EXCHANGE'):
            messages.error(request, '반품/교환 주문에서는 교환을 생성할 수 없습니다.')
            return redirect('sales:order_detail', slug=self.original_order.order_number)
        if self.original_order.status not in ('SHIPPED', 'DELIVERED', 'CLOSED'):
            messages.error(request, '출고 이후 상태의 주문만 교환할 수 있습니다.')
            return redirect('sales:order_detail', slug=self.original_order.order_number)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['original_order'] = self.original_order
        ctx['order_type'] = 'EXCHANGE'
        return ctx

    def form_valid(self, form):
        from apps.core.utils import generate_document_number
        with transaction.atomic():
            order = form.save(commit=False)
            order.order_type = Order.OrderType.EXCHANGE
            order.original_order = self.original_order
            order.partner = self.original_order.partner
            order.customer = self.original_order.customer
            order.order_date = date.today()
            order.order_number = generate_document_number(Order, 'order_number', 'ORD')
            order.vat_included = self.original_order.vat_included
            order.created_by = self.request.user
            order.save()

            for item in self.original_order.items.filter(is_active=True):
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    quantity=item.quantity,
                    cost_price=item.cost_price,
                    unit_price=item.unit_price,
                    discount_rate=item.discount_rate,
                    discount_amount=item.discount_amount,
                    created_by=self.request.user,
                )
            order.update_total()
        messages.success(self.request, f'교환주문 {order.order_number}이(가) 생성되었습니다.')
        return redirect('sales:order_detail', slug=order.order_number)


class OrderModifyView(ManagerRequiredMixin, UpdateView):
    """CONFIRMED 주문의 항목 수량/가격 수정"""
    model = Order
    form_class = OrderForm
    template_name = 'sales/order_form.html'
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def dispatch(self, request, *args, **kwargs):
        order = get_object_or_404(
            Order, order_number=kwargs['slug'], is_active=True,
        )
        if order.status != 'CONFIRMED':
            messages.error(request, '확정(CONFIRMED) 상태의 주문만 수정할 수 있습니다.')
            return redirect('sales:order_detail', slug=order.order_number)
        # 출고 시작된 항목이 있으면 수정 불가
        if order.items.filter(shipped_quantity__gt=0).exists():
            messages.error(request, '출고가 시작된 항목이 있어 수정할 수 없습니다.')
            return redirect('sales:order_detail', slug=order.order_number)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = OrderItemFormSet(self.request.POST, instance=self.object)
        else:
            ctx['formset'] = OrderItemFormSet(instance=self.object)
        ctx['product_units_json'] = _product_units_json()
        ctx['product_costs_json'] = _product_costs_json(stage='ORDER')
        ctx['is_modify'] = True
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if not formset.is_valid():
            return self.form_invalid(form)

        from apps.accounting.models import AccountReceivable, TaxInvoice

        with transaction.atomic():
            order = self.object

            # 1. 기존 reserved_stock 해제
            for item in order.items.select_related('product').all():
                if item.product.is_stockable:
                    product = Product.objects.get(pk=item.product_id)
                    actual_release = min(item.quantity, product.reserved_stock)
                    if actual_release > 0:
                        Product.objects.filter(pk=item.product_id).update(
                            reserved_stock=F('reserved_stock') - actual_release,
                        )

            # 2. OrderItem 수정 저장
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            # VAT 재계산
            for item in self.object.items.filter(is_active=True):
                item.save()
            self.object.update_total()
            order.refresh_from_db()

            # 3. 새 reserved_stock 예약
            for item in order.items.select_related('product').filter(is_active=True):
                if item.product.is_stockable:
                    Product.objects.filter(pk=item.product_id).update(
                        reserved_stock=F('reserved_stock') + item.quantity,
                    )

            # 4. AR 재계산
            grand_total = int(order.grand_total) if order.grand_total else 0
            ar = AccountReceivable.objects.filter(
                order=order, is_active=True,
            ).first()
            if ar:
                ar.amount = grand_total
                ar.save(update_fields=['amount', 'updated_at'])
            elif grand_total > 0 and order.partner:
                from datetime import timedelta
                AccountReceivable.objects.create(
                    partner=order.partner,
                    order=order,
                    amount=grand_total,
                    due_date=order.delivery_date or (date.today() + timedelta(days=30)),
                    status='PENDING',
                    created_by=self.request.user,
                )

            # 5. 세금계산서 재발행 (기존 soft delete + 신규 생성)
            old_invoices = TaxInvoice.objects.filter(order=order, is_active=True)
            for inv in old_invoices:
                inv.is_active = False
                inv.save(update_fields=['is_active', 'updated_at'])

            supply_amount = int(order.total_amount) if order.total_amount else 0
            tax_amount = int(order.tax_total) if order.tax_total else 0
            if supply_amount > 0 and order.partner:
                from apps.accounting.models_platform import PlatformFinancialConfig
                issuer_type = TaxInvoice.IssuerType.SELF
                platform_name = ''
                channel = getattr(order, 'sales_channel', '')
                config = None
                if channel:
                    config = PlatformFinancialConfig.objects.filter(
                        code=channel, is_enabled=True, is_active=True,
                    ).first()
                if config and config.tax_invoice_issuer == PlatformFinancialConfig.IssuerType.PLATFORM:
                    # 플랫폼 대행 채널은 재발행도 skip
                    pass
                else:
                    TaxInvoice.objects.create(
                        invoice_type='SALES',
                        partner=order.partner,
                        order=order,
                        issue_date=date.today(),
                        supply_amount=supply_amount,
                        tax_amount=tax_amount,
                        total_amount=supply_amount + tax_amount,
                        tax_type=getattr(order, 'tax_type', TaxInvoice.TaxType.TAXABLE),
                        issuer_type=issuer_type,
                        platform_name=platform_name,
                        description=f'주문 {order.order_number} 수정 세금계산서',
                        created_by=self.request.user,
                    )

        messages.success(self.request, f'주문 {order.order_number}이(가) 수정되었습니다.')
        return redirect('sales:order_detail', slug=order.order_number)


# === Soft Delete Views ===
class _SoftDeleteView(ManagerRequiredMixin, DeleteView):
    """공용 soft delete 뷰 (비활성화)"""
    template_name = 'sales/confirm_delete.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['model_name'] = self.model._meta.verbose_name
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.soft_delete()
        messages.success(request, f'{self.model._meta.verbose_name}이(가) 삭제되었습니다.')
        return HttpResponseRedirect(self.get_success_url())


class OrderDeleteView(_SoftDeleteView):
    model = Order
    success_url = reverse_lazy('sales:order_list')
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order = self.object

        # 상태 전이 검증: CANCELLED가 허용 전이가 아닌 상태면 삭제 차단
        allowed = Order.STATUS_TRANSITIONS.get(order.status, [])
        if order.status != Order.Status.CANCELLED and 'CANCELLED' not in allowed:
            messages.error(
                request,
                f'현재 상태({order.get_status_display()})에서는 '
                f'주문을 삭제할 수 없습니다.',
            )
            return HttpResponseRedirect(
                reverse_lazy('sales:order_detail', kwargs={'slug': order.order_number}),
            )

        with transaction.atomic():
            # 취소 시그널 트리거를 위해 CANCELLED 상태로 먼저 변경
            if order.status != Order.Status.CANCELLED:
                order.status = Order.Status.CANCELLED
                order.save(update_fields=['status', 'updated_at'])

            order.soft_delete()

        messages.success(request, f'{self.model._meta.verbose_name}이(가) 삭제되었습니다.')
        return HttpResponseRedirect(self.get_success_url())


class QuotationDeleteView(ManagerRequiredMixin, DeleteView):
    model = Quotation
    template_name = 'sales/quote_confirm_delete.html'
    success_url = reverse_lazy('sales:quote_list')
    slug_field = 'quote_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order = self.object.converted_order
        blocked = False
        if order and order.is_active and order.status in (
            Order.Status.SHIPPED, Order.Status.DELIVERED, Order.Status.CLOSED
        ):
            blocked = True
        ctx['converted_order'] = order if (order and order.is_active) else None
        ctx['order_blocked'] = blocked
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order = self.object.converted_order

        # 출고/배송 완료 주문이 있으면 삭제 차단
        if order and order.is_active and order.status in (
            Order.Status.SHIPPED, Order.Status.DELIVERED, Order.Status.CLOSED
        ):
            messages.error(
                request,
                f'출고 완료된 주문({order.order_number})이 연결되어 있어 삭제할 수 없습니다.',
            )
            return HttpResponseRedirect(request.path)

        with transaction.atomic():
            delete_order = request.POST.get('delete_order') == '1'
            if delete_order and order and order.is_active:
                # STATUS_TRANSITIONS 검증: CANCELLED가 허용 전이인지 확인
                allowed = Order.STATUS_TRANSITIONS.get(order.status, [])
                if 'CANCELLED' not in allowed:
                    messages.error(
                        request,
                        f'연결된 주문({order.order_number})이 '
                        f'{order.get_status_display()} 상태여서 취소할 수 없습니다.',
                    )
                    return HttpResponseRedirect(request.path)
                order.status = Order.Status.CANCELLED
                order.save(update_fields=['status', 'updated_at'])
                order.soft_delete()

            self.object.soft_delete()

        messages.success(request, '견적서가 삭제되었습니다.')
        return HttpResponseRedirect(self.get_success_url())


class CustomerDeleteView(_SoftDeleteView):
    model = Customer
    success_url = reverse_lazy('sales:customer_list')
    slug_field = 'code'
    slug_url_kwarg = 'slug'


class PartnerDeleteView(_SoftDeleteView):
    model = Partner
    success_url = reverse_lazy('sales:partner_list')
    slug_field = 'code'
    slug_url_kwarg = 'slug'


class ShipmentDeleteView(_SoftDeleteView):
    model = Shipment
    success_url = reverse_lazy('sales:shipment_list')
    slug_field = 'shipment_number'
    slug_url_kwarg = 'slug'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        shipment = self.object

        with transaction.atomic():
            # ShipmentItem 연쇄 처리: StockMovement soft delete + shipped_quantity 롤백
            for si in shipment.items.filter(is_active=True):
                # 관련 OUT StockMovement soft delete + 실제매출원가 합산(T10)
                from apps.inventory.models import StockMovement
                movements = StockMovement.objects.filter(
                    movement_number__contains=f'SH{shipment.pk}-{si.pk}',
                    movement_type='OUT',
                    is_active=True,
                )
                cogs_to_reverse = Decimal('0')
                for mv in movements:
                    cogs_to_reverse += mv.cogs_amount or Decimal('0')
                    mv.is_active = False
                    mv.save(update_fields=['is_active', 'updated_at'])

                # shipped_quantity 롤백
                OrderItem.objects.filter(pk=si.order_item_id).update(
                    shipped_quantity=F('shipped_quantity') - si.quantity,
                )

                # T10 — 실제매출원가 차감 (출고가 취소되었으므로)
                if cogs_to_reverse > 0:
                    OrderItem.objects.filter(pk=si.order_item_id).update(
                        actual_cogs=F('actual_cogs') - cogs_to_reverse,
                    )

                # 예약재고 복원 (출고 취소이므로 다시 예약)
                Product.objects.filter(pk=si.order_item.product_id).update(
                    reserved_stock=F('reserved_stock') + si.quantity,
                )

                si.soft_delete()

            shipment.soft_delete()

        messages.success(request, f'{self.model._meta.verbose_name}이(가) 삭제되었습니다.')
        return HttpResponseRedirect(self.get_success_url())


# === 수수료율 ===
class CommissionRateListView(ManagerRequiredMixin, ListView):
    model = CommissionRate
    template_name = 'sales/commission_rate_list.html'
    context_object_name = 'commission_rates'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).select_related(
            'partner', 'product',
        ).order_by('-pk')


class CommissionRateCreateView(ManagerRequiredMixin, CreateView):
    model = CommissionRate
    form_class = CommissionRateForm
    template_name = 'sales/commission_rate_form.html'
    success_url = reverse_lazy('sales:commission_rate_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CommissionRateUpdateView(ManagerRequiredMixin, UpdateView):
    model = CommissionRate
    form_class = CommissionRateForm
    template_name = 'sales/commission_rate_form.html'
    success_url = reverse_lazy('sales:commission_rate_list')


# === 수수료내역 ===
class CommissionRecordListView(ManagerRequiredMixin, ListView):
    model = CommissionRecord
    template_name = 'sales/commission_list.html'
    context_object_name = 'commissions'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related(
            'partner', 'order',
        )
        status = self.request.GET.get('status')
        partner = self.request.GET.get('partner')
        if status:
            qs = qs.filter(status=status)
        if partner:
            qs = qs.filter(partner_id=partner)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['partners'] = Partner.objects.filter(is_active=True)
        ctx['status_choices'] = CommissionRecord.Status.choices
        # 수수료 구성 설명 추가
        for commission in ctx.get('object_list', []):
            parts = []
            rates = CommissionRate.objects.filter(
                partner=commission.partner, is_active=True,
            )
            for r in rates:
                if r.calc_type == 'PERCENT':
                    base_label = '판매가' if r.base_type == 'TOTAL' else '공급가'
                    parts.append(f'{r.rate}%({base_label})')
                else:
                    parts.append(f'{int(r.fixed_amount):,}원')
            commission.commission_desc = ' + '.join(parts) if parts else '-'
        return ctx


class CommissionRecordCreateView(ManagerRequiredMixin, CreateView):
    model = CommissionRecord
    form_class = CommissionRecordForm
    template_name = 'sales/commission_form.html'
    success_url = reverse_lazy('sales:commission_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # 주문별 금액 + 서버 계산 수수료
        order_amounts = {}
        for o in Order.objects.filter(
            is_active=True, status__in=['DELIVERED', 'CLOSED'],
        ).select_related('partner').prefetch_related('items__product'):
            supply = int(o.total_amount)
            grand = int(o.grand_total or 0)
            comm = 0
            if o.partner:
                for item in o.items.filter(is_active=True):
                    s = int(item.amount) if item.amount else 0
                    t = s + (int(item.tax_amount) if item.tax_amount else 0)
                    comm += o.partner.calculate_commission(
                        s, total_amount=t, product=item.product,
                    )
            order_amounts[str(o.pk)] = {
                'amount': supply,
                'grand_total': grand,
                'partner': o.partner_id,
                'commission': comm,
            }
        ctx['order_amounts_json'] = order_amounts
        # 거래처별 수수료 구성 설명 (참고용)
        rate_map = {}
        for r in CommissionRate.objects.filter(is_active=True):
            pid = str(r.partner_id)
            if pid not in rate_map:
                rate_map[pid] = {'rate': 0, 'fixed': 0}
            if r.calc_type == 'PERCENT':
                rate_map[pid]['rate'] += float(r.rate)
            else:
                rate_map[pid]['fixed'] += int(r.fixed_amount)
        ctx['commission_rates_json'] = rate_map
        return ctx

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


logger = logging.getLogger(__name__)


class CommissionRecordSettleView(ManagerRequiredMixin, View):
    def post(self, request, pk):
        record = get_object_or_404(CommissionRecord, pk=pk)
        if record.status == 'PENDING':
            with transaction.atomic():
                from apps.sales.signals import _auto_create_commission_disbursement
                record.status = 'SETTLED'
                record.settled_date = date.today()
                record.save(update_fields=['status', 'settled_date', 'updated_at'])
                _auto_create_commission_disbursement(record)

            messages.success(
                request,
                f'{record.partner.name} 수수료 {record.commission_amount:,}원 정산완료 처리',
            )
        return redirect('sales:commission_list')


class CommissionSummaryView(ManagerRequiredMixin, TemplateView):
    template_name = 'sales/commission_summary.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        summary = CommissionRecord.objects.filter(
            is_active=True,
        ).values(
            'partner__id', 'partner__name',
        ).annotate(
            total_orders=Count('id'),
            total_commission=Sum('commission_amount'),
            settled=Sum('commission_amount', filter=Q(status='SETTLED')),
            unsettled=Sum('commission_amount', filter=Q(status='PENDING')),
        ).order_by('partner__name')
        ctx['summary'] = summary
        return ctx


# === PDF 다운로드 ===
class OrderQuotePDFView(LoginRequiredMixin, DetailView):
    """견적서 PDF 다운로드"""
    model = Order
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def get(self, request, *args, **kwargs):
        order = self.get_object()
        from apps.core.pdf import generate_quotation_pdf
        return generate_quotation_pdf(order)


class OrderPurchaseOrderPDFView(LoginRequiredMixin, DetailView):
    """발주서 PDF 다운로드"""
    model = Order
    slug_field = 'order_number'
    slug_url_kwarg = 'slug'

    def get(self, request, *args, **kwargs):
        order = self.get_object()
        from apps.core.pdf import generate_purchase_order_pdf
        return generate_purchase_order_pdf(order)


# === Excel 다운로드 ===
class OrderExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import io
        from collections import defaultdict
        from datetime import date as _date
        from decimal import Decimal

        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        orders = Order.objects.filter(is_active=True).select_related(
            'partner', 'customer', 'assigned_to',
        ).prefetch_related('source_quotation', 'items__product')
        order_list = list(orders)

        # --- 공통 스타일 ---
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill('solid', fgColor='1F4E79')
        hdr_font = Font('맑은 고딕', 11, bold=True, color='FFFFFF')
        item_hdr_fill = PatternFill('solid', fgColor='4472C4')
        item_hdr_font = Font('맑은 고딕', 10, bold=True, color='FFFFFF')
        body = Font('맑은 고딕', 10)
        body_bold = Font('맑은 고딕', 10, bold=True)
        title_font = Font('맑은 고딕', 16, bold=True, color='1F4E79')
        section_font = Font('맑은 고딕', 12, bold=True, color='1F4E79')
        subtitle_font = Font('맑은 고딕', 9, color='666666')
        kpi_val_font = Font('맑은 고딕', 20, bold=True, color='1F4E79')
        kpi_lbl_font = Font('맑은 고딕', 9, color='666666')
        money_fmt = '#,##0'
        pct_fmt = '0.0%'
        stripe_fill = PatternFill('solid', fgColor='F2F7FB')
        subtotal_fill = PatternFill('solid', fgColor='E2EFDA')
        grand_fill = PatternFill('solid', fgColor='D6E4F0')
        kpi_fill = PatternFill('solid', fgColor='F8F9FA')
        center = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(vertical='center', wrap_text=True)

        status_colors = {
            'DRAFT': 'F2F2F2', 'CONFIRMED': 'DAEEF3',
            'SHIPPED': 'E2EFDA', 'DELIVERED': 'C6EFCE',
            'CANCELLED': 'FFC7CE', 'CLOSED': 'B4C6E7',
            'PARTIAL_SHIPPED': 'FCE4D6',
        }
        status_labels = {
            'DRAFT': '작성중', 'CONFIRMED': '확정',
            'SHIPPED': '출고완료', 'DELIVERED': '배송완료',
            'CANCELLED': '취소', 'CLOSED': '종결',
            'PARTIAL_SHIPPED': '부분출고',
        }

        # --- 사전 집계 ---
        total_count = len(order_list)
        active_orders = [o for o in order_list if o.status != 'CANCELLED']
        cancelled_orders = [o for o in order_list if o.status == 'CANCELLED']
        total_grand = sum(int(o.grand_total) for o in active_orders)
        total_supply = sum(int(o.total_amount) for o in active_orders)
        total_tax = sum(int(o.tax_total) for o in active_orders)
        avg_order = total_grand // len(active_orders) if active_orders else 0

        # 상태별
        status_agg = defaultdict(lambda: {'count': 0, 'amount': 0})
        for o in order_list:
            status_agg[o.status]['count'] += 1
            status_agg[o.status]['amount'] += int(o.grand_total)

        # 월별
        monthly_agg = defaultdict(lambda: {'count': 0, 'amount': 0})
        for o in active_orders:
            if o.order_date:
                key = o.order_date.strftime('%Y-%m')
                monthly_agg[key]['count'] += 1
                monthly_agg[key]['amount'] += int(o.grand_total)

        # 거래처 TOP 10
        partner_agg = defaultdict(lambda: {'count': 0, 'amount': 0})
        for o in active_orders:
            name = (o.partner.name if o.partner else '') or (o.customer.name if o.customer else '') or '(미지정)'
            partner_agg[name]['count'] += 1
            partner_agg[name]['amount'] += int(o.grand_total)
        top_partners = sorted(partner_agg.items(), key=lambda x: x[1]['amount'], reverse=True)[:10]

        # 제품 TOP 10
        product_agg = defaultdict(lambda: {'qty': 0, 'amount': 0})
        for o in active_orders:
            for item in o.items.all():
                if item.is_active and item.product:
                    product_agg[item.product.name]['qty'] += item.quantity
                    product_agg[item.product.name]['amount'] += int(item.amount)
        top_products = sorted(product_agg.items(), key=lambda x: x[1]['amount'], reverse=True)[:10]

        # 담당자별
        staff_agg = defaultdict(lambda: {'count': 0, 'amount': 0})
        for o in active_orders:
            name = o.assigned_to.get_full_name() if o.assigned_to else '(미배정)'
            staff_agg[name]['count'] += 1
            staff_agg[name]['amount'] += int(o.grand_total)

        # 입금 현황
        paid_count = sum(1 for o in active_orders if o.is_paid)
        unpaid_count = len(active_orders) - paid_count
        paid_amount = sum(int(o.grand_total) for o in active_orders if o.is_paid)
        unpaid_amount = total_grand - paid_amount

        # 견적 전환율
        quoted_count = sum(1 for o in order_list if o.source_quotation.first())
        conversion_rate = quoted_count / total_count if total_count else 0

        # --- 헬퍼 ---
        def _write_header(ws, row, headers_list, fill=None, font=None):
            fill = fill or hdr_fill
            font = font or hdr_font
            for ci, (name, w) in enumerate(headers_list, 1):
                c = ws.cell(row, ci, name)
                c.font = font
                c.fill = fill
                c.alignment = center
                c.border = border
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[row].height = 28

        def _write_row(ws, row_num, values, money_cols=(), date_cols=(),
                       status_col=None, status_val=None):
            for ci, v in enumerate(values, 1):
                c = ws.cell(row_num, ci, v)
                c.font = body
                c.border = border
                if ci in money_cols:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci in date_cols:
                    c.number_format = 'YYYY-MM-DD'
                    c.alignment = center
                elif status_col and ci == status_col:
                    c.alignment = center
                    fc = status_colors.get(status_val)
                    if fc:
                        c.fill = PatternFill('solid', fgColor=fc)
                elif isinstance(v, (int, float, Decimal)):
                    c.alignment = right_align
                else:
                    c.alignment = left_align
            if row_num % 2 == 1:
                for ci in range(1, len(values) + 1):
                    cell = ws.cell(row_num, ci)
                    if not (status_col and ci == status_col):
                        cell.fill = stripe_fill

        wb = Workbook()

        # ========================================================
        # 시트 1: 경영 분석 대시보드
        # ========================================================
        ws = wb.active
        ws.title = '경영 분석'
        ws.sheet_properties.tabColor = '1F4E79'

        ws.merge_cells('A1:H1')
        c = ws.cell(1, 1, '주문 경영 분석 보고서')
        c.font = title_font
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 40

        ws.merge_cells('A2:H2')
        c = ws.cell(2, 1, f'출력일: {_date.today().strftime("%Y-%m-%d")}  |  데이터 기준: 활성 주문 전체')
        c.font = subtitle_font
        c.alignment = Alignment(horizontal='right')

        # --- KPI 카드 ---
        kpi_data = [
            ('총 주문건수', f'{total_count:,}건'),
            ('유효 주문금액', f'{total_grand:,}원'),
            ('평균 주문단가', f'{avg_order:,}원'),
            ('견적 전환율', f'{conversion_rate:.1%}'),
            ('입금완료', f'{paid_count:,}건'),
            ('미입금', f'{unpaid_count:,}건'),
        ]
        for ci, (label, val) in enumerate(kpi_data):
            col = ci * 2 + 1
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
            vc = ws.cell(4, col, val)
            vc.font = kpi_val_font
            vc.alignment = center
            lc = ws.cell(5, col, label)
            lc.font = kpi_lbl_font
            lc.alignment = center
            for r in (4, 5):
                for cc in (col, col + 1):
                    ws.cell(r, cc).fill = kpi_fill
                    ws.cell(r, cc).border = border
        ws.row_dimensions[4].height = 36
        ws.row_dimensions[5].height = 22
        for ci in range(1, 13):
            ws.column_dimensions[get_column_letter(ci)].width = 14

        # --- 상태별 현황 ---
        r = 7
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '상태별 주문 현황').font = section_font
        r += 1
        for ci, (name, w) in enumerate([('상태', 14), ('건수', 10), ('금액', 16), ('비중', 10)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
        chart_data_start = r + 1
        r += 1
        for status_code in ['DRAFT', 'CONFIRMED', 'SHIPPED', 'DELIVERED', 'CANCELLED']:
            d = status_agg.get(status_code, {'count': 0, 'amount': 0})
            ratio = d['count'] / total_count if total_count else 0
            vals = [status_labels.get(status_code, status_code), d['count'], d['amount'], ratio]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci, v)
                c.font = body
                c.border = border
                if ci == 3:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci == 4:
                    c.number_format = pct_fmt
                    c.alignment = center
                elif ci == 2:
                    c.alignment = center
                else:
                    c.alignment = left_align
                    fc = status_colors.get(status_code)
                    if fc:
                        c.fill = PatternFill('solid', fgColor=fc)
            r += 1
        chart_data_end = r - 1

        # 상태 파이 차트
        pie = PieChart()
        pie.title = '상태별 건수 비율'
        pie.style = 10
        pie.width = 14
        pie.height = 10
        labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
        data = Reference(ws, min_col=2, min_row=chart_data_start - 1, max_row=chart_data_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, 'F8')

        # --- 월별 추이 ---
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '월별 주문 추이').font = section_font
        r += 1
        for ci, (name, w) in enumerate([('월', 14), ('건수', 10), ('금액', 16), ('평균단가', 14)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = item_hdr_fill
            c.alignment = center
            c.border = border
        monthly_chart_start = r + 1
        r += 1
        for month_key in sorted(monthly_agg.keys()):
            d = monthly_agg[month_key]
            avg = d['amount'] // d['count'] if d['count'] else 0
            for ci, v in enumerate([month_key, d['count'], d['amount'], avg], 1):
                c = ws.cell(r, ci, v)
                c.font = body
                c.border = border
                if ci in (3, 4):
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci == 2:
                    c.alignment = center
                else:
                    c.alignment = center
            r += 1
        monthly_chart_end = r - 1

        if monthly_chart_end >= monthly_chart_start:
            bar = BarChart()
            bar.type = 'col'
            bar.title = '월별 매출 추이'
            bar.style = 10
            bar.width = 14
            bar.height = 10
            bar.y_axis.title = '금액(원)'
            cats = Reference(ws, min_col=1, min_row=monthly_chart_start, max_row=monthly_chart_end)
            vals = Reference(ws, min_col=3, min_row=monthly_chart_start - 1, max_row=monthly_chart_end)
            bar.add_data(vals, titles_from_data=True)
            bar.set_categories(cats)
            bar.shape = 4
            ws.add_chart(bar, f'F{monthly_chart_start - 1}')

        # --- 거래처/고객 TOP 10 ---
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '거래처/고객 매출 TOP 10').font = section_font
        r += 1
        for ci, (name, _) in enumerate([('순위', 6), ('거래처/고객', 20), ('건수', 10), ('금액', 16)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
        r += 1
        for rank, (name, d) in enumerate(top_partners, 1):
            for ci, v in enumerate([rank, name, d['count'], d['amount']], 1):
                c = ws.cell(r, ci, v)
                c.font = body_bold if rank <= 3 else body
                c.border = border
                if ci == 4:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci in (1, 3):
                    c.alignment = center
                else:
                    c.alignment = left_align
            if rank <= 3:
                for ci in range(1, 5):
                    ws.cell(r, ci).fill = PatternFill('solid', fgColor='FFF2CC')
            r += 1

        # --- 제품 TOP 10 ---
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '제품별 매출 TOP 10').font = section_font
        r += 1
        for ci, (name, _) in enumerate([('순위', 6), ('제품명', 28), ('판매수량', 10), ('매출액', 16)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = item_hdr_fill
            c.alignment = center
            c.border = border
        r += 1
        for rank, (name, d) in enumerate(top_products, 1):
            for ci, v in enumerate([rank, name, d['qty'], d['amount']], 1):
                c = ws.cell(r, ci, v)
                c.font = body_bold if rank <= 3 else body
                c.border = border
                if ci == 4:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci in (1, 3):
                    c.alignment = center
                else:
                    c.alignment = left_align
            if rank <= 3:
                for ci in range(1, 5):
                    ws.cell(r, ci).fill = PatternFill('solid', fgColor='FFF2CC')
            r += 1

        # --- 담당자별 실적 ---
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '담당자별 실적').font = section_font
        r += 1
        for ci, (name, _) in enumerate([('담당자', 14), ('건수', 10), ('금액', 16), ('비중', 10)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
        r += 1
        for name, d in sorted(staff_agg.items(), key=lambda x: x[1]['amount'], reverse=True):
            ratio = d['amount'] / total_grand if total_grand else 0
            for ci, v in enumerate([name, d['count'], d['amount'], ratio], 1):
                c = ws.cell(r, ci, v)
                c.font = body
                c.border = border
                if ci == 3:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci == 4:
                    c.number_format = pct_fmt
                    c.alignment = center
                elif ci == 2:
                    c.alignment = center
                else:
                    c.alignment = left_align
            r += 1

        # --- 입금 현황 ---
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, '입금 현황').font = section_font
        r += 1
        for ci, (name, _) in enumerate([('구분', 14), ('건수', 10), ('금액', 16), ('비중', 10)], 1):
            c = ws.cell(r, ci, name)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
        r += 1
        for label, cnt, amt in [
            ('입금완료', paid_count, paid_amount),
            ('미입금', unpaid_count, unpaid_amount),
        ]:
            ratio = cnt / len(active_orders) if active_orders else 0
            for ci, v in enumerate([label, cnt, amt, ratio], 1):
                c = ws.cell(r, ci, v)
                c.font = body
                c.border = border
                if ci == 3:
                    c.number_format = money_fmt
                    c.alignment = right_align
                elif ci == 4:
                    c.number_format = pct_fmt
                    c.alignment = center
                elif ci == 2:
                    c.alignment = center
                else:
                    c.alignment = left_align
                    if label == '미입금':
                        c.fill = PatternFill('solid', fgColor='FFC7CE')
                    else:
                        c.fill = PatternFill('solid', fgColor='C6EFCE')
            r += 1

        ws.freeze_panes = 'A4'

        # ========================================================
        # 시트 2: 주문 목록
        # ========================================================
        ws2 = wb.create_sheet('주문 목록')
        ws2.sheet_properties.tabColor = '2E75B6'

        ws2.merge_cells('A1:N1')
        ws2.cell(1, 1, '주문 전체 목록').font = title_font
        ws2.row_dimensions[1].height = 36
        ws2.merge_cells('A2:N2')
        c = ws2.cell(2, 1, f'총 {total_count}건  |  유효 매출 {total_grand:,}원  |  취소 {len(cancelled_orders)}건')
        c.font = subtitle_font
        c.alignment = Alignment(horizontal='right')

        list_headers = [
            ('주문번호', 16), ('유형', 8), ('거래처', 18), ('고객', 14),
            ('고객 연락처', 16), ('고객 주소', 30),
            ('담당자', 10), ('주문일', 12), ('주문 전환일', 14), ('납기일', 12),
            ('상태', 10), ('공급가액', 14), ('부가세', 12), ('총합계(VAT)', 14),
        ]
        _write_header(ws2, 4, list_headers)

        row = 5
        sum_supply = sum_tax = sum_grand_all = 0
        for o in order_list:
            sq = o.source_quotation.first()
            vals = [
                o.order_number,
                o.get_order_type_display() if hasattr(o, 'get_order_type_display') else '',
                o.partner.name if o.partner else '',
                o.customer.name if o.customer else '',
                o.customer.phone if o.customer else '',
                o.customer.address if o.customer else '',
                o.assigned_to.get_full_name() if o.assigned_to else '',
                sq.quote_date if sq and sq.quote_date else None,
                o.order_date,
                o.delivery_date,
                o.get_status_display(),
                int(o.total_amount),
                int(o.tax_total),
                int(o.grand_total),
            ]
            _write_row(ws2, row, vals,
                       money_cols=(12, 13, 14), date_cols=(8, 9, 10),
                       status_col=11, status_val=o.status)
            sum_supply += int(o.total_amount)
            sum_tax += int(o.tax_total)
            sum_grand_all += int(o.grand_total)
            row += 1

        # 합계
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws2.cell(row, 1, '합 계')
        c.font = Font('맑은 고딕', 11, bold=True, color='1F4E79')
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.fill = grand_fill
        c.border = border
        for ci in range(2, 12):
            ws2.cell(row, ci).fill = grand_fill
            ws2.cell(row, ci).border = border
        for ci, val in [(12, sum_supply), (13, sum_tax), (14, sum_grand_all)]:
            c = ws2.cell(row, ci, val)
            c.font = Font('맑은 고딕', 11, bold=True)
            c.number_format = money_fmt
            c.alignment = right_align
            c.fill = grand_fill
            c.border = border

        ws2.auto_filter.ref = f'A4:{get_column_letter(len(list_headers))}{row - 1}'
        ws2.freeze_panes = 'A5'

        # ========================================================
        # 시트 3: 품목 상세
        # ========================================================
        ws3 = wb.create_sheet('품목 상세')
        ws3.sheet_properties.tabColor = '4472C4'

        ws3.merge_cells('A1:K1')
        ws3.cell(1, 1, '주문별 품목 상세').font = title_font
        ws3.row_dimensions[1].height = 36

        detail_headers = [
            ('주문번호', 16), ('거래처/고객', 18), ('주문일', 12), ('상태', 10),
            ('제품코드', 14), ('제품명', 28), ('수량', 8), ('출고수량', 10),
            ('단가', 14), ('공급가액', 14), ('부가세', 12),
        ]
        _write_header(ws3, 3, detail_headers, fill=item_hdr_fill, font=item_hdr_font)

        row3 = 4
        for o in order_list:
            items = [i for i in o.items.all() if i.is_active]
            if not items:
                continue
            client = (o.partner.name if o.partner else '') or (o.customer.name if o.customer else '')
            for item in items:
                vals = [
                    o.order_number, client, o.order_date, o.get_status_display(),
                    item.product.code if item.product else '',
                    item.product.name if item.product else '',
                    item.quantity, item.shipped_quantity,
                    int(item.unit_price), int(item.amount), int(item.tax_amount),
                ]
                _write_row(ws3, row3, vals,
                           money_cols=(9, 10, 11), date_cols=(3,),
                           status_col=4, status_val=o.status)
                # 미출고 강조
                if item.quantity > item.shipped_quantity and o.status in ('CONFIRMED', 'SHIPPED'):
                    ws3.cell(row3, 8).fill = PatternFill('solid', fgColor='FFC7CE')
                row3 += 1

            # 소계
            ws3.merge_cells(start_row=row3, start_column=1, end_row=row3, end_column=9)
            c = ws3.cell(row3, 1, f'  {o.order_number} 소계')
            c.font = body_bold
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.fill = subtotal_fill
            c.border = border
            for ci in range(2, 10):
                ws3.cell(row3, ci).fill = subtotal_fill
                ws3.cell(row3, ci).border = border
            for ci, val in [(10, int(o.total_amount)), (11, int(o.tax_total))]:
                c = ws3.cell(row3, ci, val)
                c.font = body_bold
                c.number_format = money_fmt
                c.alignment = right_align
                c.fill = subtotal_fill
                c.border = border
            row3 += 1

        ws3.auto_filter.ref = f'A3:{get_column_letter(len(detail_headers))}{row3 - 1}'
        ws3.freeze_panes = 'A4'

        # ========================================================
        # 응답
        # ========================================================
        filename = f'주문관리보고서_{_date.today().strftime("%Y%m%d")}.xlsx'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# === Excel 일괄 가져오기 공통 헬퍼 ===
def _build_preview(result, data):
    headers = list(data.headers) if data.headers else []
    rows = []
    for i, row_result in enumerate(result.rows):
        values = list(data[i]) if i < len(data) else []
        if row_result.errors:
            status = 'error'
        elif row_result.import_type == 'new':
            status = 'new'
        elif row_result.import_type == 'update':
            status = 'update'
        else:
            status = 'skip'
        rows.append({'status': status, 'values': values})

    totals = {
        'new': result.totals.get('new', 0),
        'update': result.totals.get('update', 0),
        'skip': result.totals.get('skip', 0),
        'error': result.totals.get('error', 0) + result.totals.get('invalid', 0),
    }
    return {
        'headers': headers,
        'rows': rows,
        'totals': totals,
        'has_valid_rows': totals['new'] > 0 or totals['update'] > 0,
        'file_name': '',
    }


def _collect_errors(result):
    errors = []
    if result.base_errors:
        for err in result.base_errors:
            errors.append(str(err.error))
    for i, row in enumerate(result.rows):
        if row.errors:
            for err in row.errors:
                errors.append(f'{i + 1}행: {err.error}')
        if hasattr(row, 'validation_error') and row.validation_error:
            errors.append(f'{i + 1}행: {row.validation_error}')
    return errors


def _parse_import_file(request, import_file):
    """업로드된 파일을 tablib Dataset으로 파싱"""
    from apps.core.import_views import parse_import_file
    return parse_import_file(import_file)


class PartnerImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import PartnerResource
            return export_resource_data(PartnerResource(), '거래처_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '거래처 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:partner_list')
        ctx['sample_url'] = reverse_lazy('sales:partner_import_sample')
        ctx['field_hints'] = [
            '거래처코드(code)가 동일하면 기존 거래처가 수정됩니다.',
            '유형(partner_type): CUSTOMER(고객), SUPPLIER(공급처), BOTH(고객/공급처)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')

        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = PartnerResource()

        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv)')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = result.totals.get('new', 0) + result.totals.get('update', 0)
            messages.success(request, f'거래처 {total}건이 성공적으로 가져오기 되었습니다.')
            return HttpResponseRedirect(reverse_lazy('sales:partner_list'))


class PartnerImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('code', 15), ('name', 25), ('partner_type', 12),
            ('business_number', 15), ('representative', 12),
            ('contact_name', 12), ('phone', 15), ('email', 20),
            ('address_road', 30), ('address_detail', 20),
        ]
        rows = [
            ['PTN-001', '샘플 거래처', 'CUSTOMER', '123-45-67890', '홍길동',
             '김담당', '02-1234-5678', 'sample@example.com',
             '서울시 강남구 테헤란로 123', '4층 401호'],
            ['PTN-002', '샘플 공급처', 'SUPPLIER', '987-65-43210', '이대표',
             '박담당', '031-9876-5432', 'supplier@example.com',
             '경기도 성남시 분당구 판교로 456', '2층'],
        ]
        return export_to_excel(
            '거래처_가져오기_양식', headers, rows,
            filename='거래처_가져오기_양식.xlsx',
            required_columns=[0, 1, 2],  # code, name, partner_type
        )


class CustomerImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import CustomerResource
            return export_resource_data(CustomerResource(), '고객_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '고객 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:customer_list')
        ctx['sample_url'] = reverse_lazy('sales:customer_import_sample')
        ctx['field_hints'] = [
            '고객명(name)과 연락처(phone)가 동일하면 기존 고객이 수정됩니다.',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')

        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = CustomerResource()

        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv)')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = result.totals.get('new', 0) + result.totals.get('update', 0)
            messages.success(request, f'고객 {total}건이 성공적으로 가져오기 되었습니다.')
            return HttpResponseRedirect(reverse_lazy('sales:customer_list'))


class CustomerImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('name', 20), ('phone', 15), ('email', 25),
            ('address_road', 30), ('address_detail', 20),
        ]
        rows = [
            ['홍길동', '010-1234-5678', 'hong@example.com',
             '서울시 강남구 테헤란로 123', '4층 401호'],
            ['김철수', '010-9876-5432', 'kim@example.com',
             '경기도 성남시 분당구 판교로 456', '2층'],
        ]
        # name, phone 필수 (import_id_fields)
        return export_to_excel(
            '고객_가져오기_양식', headers, rows,
            filename='고객_가져오기_양식.xlsx',
            required_columns=[0, 1],  # name, phone
        )


# === 수수료율 일괄 가져오기 ===

class CommissionRateImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import CommissionRateResource
            return export_resource_data(CommissionRateResource(), '수수료율_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '수수료율 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:commission_rate_list')
        ctx['sample_url'] = reverse_lazy(
            'sales:commission_rate_import_sample',
        )
        ctx['field_hints'] = [
            '거래처코드(partner_code) + 제품코드(product_code) '
            '조합이 동일하면 기존 수수료율이 수정됩니다.',
            'product_code를 비워두면 해당 거래처의 기본 수수료율로 등록됩니다.',
            '수수료율(rate)은 소수점 2자리까지 (예: 3.50)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')

        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = CommissionRateResource()

        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(
                    request,
                    '지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv)',
                )
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(
                request,
                f'{total}건의 수수료율이 가져오기 되었습니다.',
            )
            return HttpResponseRedirect(
                str(reverse_lazy('sales:commission_rate_list')),
            )


class CommissionRateImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('partner_code', 15),
            ('product_code', 15),
            ('rate', 12),
        ]
        rows = [
            ['PTN-001', '', '3.00'],
            ['PTN-001', 'PRD-001', '5.00'],
            ['PTN-002', '', '2.50'],
            ['PTN-002', 'PRD-003', '4.00'],
        ]
        return export_to_excel(
            '수수료율_가져오기_양식', headers, rows,
            filename='수수료율_가져오기_양식.xlsx',
            required_columns=[0, 2],  # partner_code, rate
        )


class CommissionRateExcelView(ManagerRequiredMixin, View):
    """현재 수수료율 전체를 Excel로 내보내기 (템플릿 대용)"""
    def get(self, request):
        from apps.core.excel import export_to_excel
        qs = CommissionRate.objects.filter(
            is_active=True,
        ).select_related('partner', 'product').order_by(
            'partner__code', 'product__code',
        )
        headers = [
            ('partner_code', 15),
            ('product_code', 15),
            ('rate', 12),
        ]
        rows = [
            [
                cr.partner.code,
                cr.product.code if cr.product else '',
                float(cr.rate),
            ]
            for cr in qs
        ]
        return export_to_excel(
            '수수료율_현황', headers, rows,
            filename='수수료율_현황.xlsx',
        )


# === 주문 일괄 가져오기 ===

class OrderImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import OrderResource
            return export_resource_data(OrderResource(), '주문_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '주문 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:order_list')
        ctx['sample_url'] = reverse_lazy('sales:order_import_sample')
        ctx['field_hints'] = [
            '주문번호(order_number)가 동일하면 기존 주문이 수정됩니다.',
            'partner_code: 거래처코드, customer_name: 고객명',
            'status: PENDING(대기), CONFIRMED(확정), '
            'SHIPPED(출하), DELIVERED(배송완료), CANCELLED(취소)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from .resources import OrderResource
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)
        resource = OrderResource()
        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(
                    request, '지원하지 않는 파일 형식입니다.',
                )
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)
        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(request, f'{total}건 가져오기 완료.')
            return HttpResponseRedirect(
                str(reverse_lazy('sales:order_list')),
            )


class OrderImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('order_number', 15), ('partner_code', 15),
            ('customer_code', 15), ('order_date', 12),
            ('status', 12), ('shipping_address_road', 30),
            ('shipping_address_detail', 20), ('notes', 20),
        ]
        rows = [
            ['ORD-2026-001', 'PTN-001', 'CST-0001',
             '2026-03-01', 'PENDING', '서울시 강남구 테헤란로 123', '4층', ''],
            ['ORD-2026-002', 'PTN-002', 'CST-0002',
             '2026-03-05', 'CONFIRMED', '경기도 성남시 분당구 판교로 456', '2층', ''],
        ]
        return export_to_excel(
            '주문_가져오기_양식', headers, rows,
            filename='주문_가져오기_양식.xlsx',
            required_columns=[0, 3],  # order_number, order_date
        )


# === 배송 일괄 가져오기 ===

class ShipmentImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import ShipmentResource
            return export_resource_data(ShipmentResource(), '배송_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '배송 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:shipment_list')
        ctx['sample_url'] = reverse_lazy(
            'sales:shipment_import_sample',
        )
        ctx['field_hints'] = [
            '배송번호(shipment_number)가 동일하면 수정됩니다.',
            'order_number: 연결할 주문번호',
            'carrier: 택배사명 (CJ, HANJIN, LOTTE 등)',
            'status: READY(준비), SHIPPED(출하), '
            'IN_TRANSIT(배송중), DELIVERED(배송완료)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from .resources import ShipmentResource
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)
        resource = ShipmentResource()
        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(
                    request, '지원하지 않는 파일 형식입니다.',
                )
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)
        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(request, f'{total}건 가져오기 완료.')
            return HttpResponseRedirect(
                str(reverse_lazy('sales:shipment_list')),
            )


class ShipmentImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('shipment_number', 18), ('order_number', 15),
            ('carrier', 10), ('tracking_number', 18),
            ('status', 12), ('shipped_date', 12),
            ('delivered_date', 12), ('receiver_name', 12),
            ('receiver_phone', 15),
            ('receiver_address_road', 30), ('receiver_address_detail', 20),
        ]
        rows = [
            ['SHP-2026-001', 'ORD-2026-001', 'CJ',
             '123456789012', 'SHIPPED', '2026-03-02',
             '', '홍길동', '010-1234-5678',
             '서울시 강남구 테헤란로 123', '4층 401호'],
        ]
        return export_to_excel(
            '배송_가져오기_양식', headers, rows,
            filename='배송_가져오기_양식.xlsx',
            required_columns=[0, 1],  # shipment_number, order_number
        )


# === 판매제품 통합 관리 ===

def _build_sold_product_queryset(request):
    """판매제품 목록 조회를 위한 공통 쿼리셋 빌더."""
    qs = OrderItem.objects.filter(
        is_active=True,
        order__is_active=True,
        order__status__in=[Order.Status.SHIPPED, Order.Status.DELIVERED],
    ).select_related(
        'order', 'order__customer', 'order__partner', 'product',
    ).order_by('-order__order_date', '-order__pk')

    # 검색
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(product__name__icontains=q)
            | Q(product__code__icontains=q)
            | Q(order__customer__name__icontains=q)
            | Q(order__partner__name__icontains=q)
            | Q(order__customer__purchases__serial_number__icontains=q)
            | Q(order__order_number__icontains=q)
        ).distinct()

    # 날짜 범위 필터
    sold_from = request.GET.get('sold_from', '').strip()
    sold_to = request.GET.get('sold_to', '').strip()
    if sold_from:
        qs = qs.filter(order__order_date__gte=sold_from)
    if sold_to:
        qs = qs.filter(order__order_date__lte=sold_to)

    # 제품유형 필터
    product_type = request.GET.get('product_type', '').strip()
    if product_type:
        qs = qs.filter(product__product_type=product_type)

    return qs


def _get_warranty_map(order_items):
    """OrderItem 목록에 대해 구매내역 맵 + ProductRegistration 맵 생성."""
    from apps.warranty.models import ProductRegistration
    from .models import CustomerPurchase

    customer_ids = set()
    for item in order_items:
        if item.order.customer_id:
            customer_ids.add(item.order.customer_id)

    # CustomerPurchase 맵: (customer_id, product_id) -> purchase
    purchase_map = {}
    if customer_ids:
        purchases = CustomerPurchase.objects.filter(
            customer_id__in=customer_ids, is_active=True,
        ).select_related('product')
        for p in purchases:
            purchase_map[(p.customer_id, p.product_id)] = p

    # ProductRegistration 맵: (product_id, serial_number) -> registration
    serial_numbers = {p.serial_number for p in purchase_map.values() if p.serial_number}
    reg_map = {}
    if serial_numbers:
        regs = ProductRegistration.objects.filter(
            serial_number__in=serial_numbers,
        ).select_related('product')
        for reg in regs:
            reg_map[(reg.product_id, reg.serial_number)] = reg

    return purchase_map, reg_map


def _get_service_count_map(order_items):
    """OrderItem 목록에 대해 (customer_id, product_id) -> AS건수 맵 생성."""
    from apps.service.models import ServiceRequest

    customer_product_pairs = set()
    for item in order_items:
        if item.order.customer_id:
            customer_product_pairs.add((item.order.customer_id, item.product_id))

    if not customer_product_pairs:
        return {}

    customer_ids = {cp[0] for cp in customer_product_pairs}
    product_ids = {cp[1] for cp in customer_product_pairs}

    counts = (
        ServiceRequest.objects.filter(
            customer_id__in=customer_ids,
            product_id__in=product_ids,
        )
        .values('customer_id', 'product_id')
        .annotate(cnt=Count('id'))
    )

    count_map = {}
    for row in counts:
        count_map[(row['customer_id'], row['product_id'])] = row['cnt']

    return count_map


def _enrich_sold_items(items):
    """OrderItem 리스트에 보증/AS 정보를 추가한 dict 리스트 반환."""
    items_list = list(items)
    if not items_list:
        return []

    purchase_map, reg_map = _get_warranty_map(items_list)
    svc_map = _get_service_count_map(items_list)
    today = date.today()

    result = []
    for item in items_list:
        customer = item.order.customer
        partner = item.order.partner

        customer_name = ''
        phone = ''
        serial_number = ''
        warranty_end = None
        warranty_status = 'none'

        if customer:
            customer_name = customer.name
            phone = customer.phone or ''

            # CustomerPurchase에서 해당 제품의 구매내역 조회
            purchase = purchase_map.get((customer.pk, item.product_id))
            if purchase:
                serial_number = purchase.serial_number or ''
                warranty_end = purchase.warranty_end

                # ProductRegistration으로 보증정보 보완
                if serial_number:
                    reg = reg_map.get((item.product_id, serial_number))
                    if reg:
                        warranty_end = reg.warranty_end

            if warranty_end:
                warranty_status = 'valid' if warranty_end >= today else 'expired'
        elif partner:
            customer_name = partner.name
            phone = partner.phone or ''

        service_count = svc_map.get((item.order.customer_id, item.product_id), 0) if customer else 0

        result.append({
            'item': item,
            'order': item.order,
            'product': item.product,
            'customer_name': customer_name,
            'phone': phone,
            'serial_number': serial_number,
            'warranty_end': warranty_end,
            'warranty_status': warranty_status,
            'service_count': service_count,
        })

    return result


class SoldProductListView(ManagerRequiredMixin, TemplateView):
    template_name = 'sales/sold_product_list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.inventory.models import Product

        qs = _build_sold_product_queryset(self.request)

        # 보증상태 필터 (post-query)
        warranty_filter = self.request.GET.get('warranty_status', '').strip()

        # 페이지네이션
        paginator = Paginator(qs, 20)
        page_number = self.request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        enriched = _enrich_sold_items(page_obj.object_list)

        # 보증상태 필터 적용 (post-processing)
        if warranty_filter == 'valid':
            enriched = [r for r in enriched if r['warranty_status'] == 'valid']
        elif warranty_filter == 'expired':
            enriched = [r for r in enriched if r['warranty_status'] == 'expired']

        ctx['sold_items'] = enriched
        ctx['page_obj'] = page_obj
        ctx['is_paginated'] = page_obj.has_other_pages()
        ctx['product_type_choices'] = Product.ProductType.choices
        return ctx


class SoldProductDetailView(ManagerRequiredMixin, DetailView):
    model = OrderItem
    template_name = 'sales/sold_product_detail.html'
    context_object_name = 'order_item'

    def get_queryset(self):
        return OrderItem.objects.filter(
            order__status__in=[Order.Status.SHIPPED, Order.Status.DELIVERED],
        ).select_related(
            'order', 'order__customer', 'order__partner', 'product',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.warranty.models import ProductRegistration
        from apps.service.models import ServiceRequest
        from .models import CustomerPurchase

        item = self.object
        customer = item.order.customer
        today = date.today()

        # 보증 정보
        warranty_info = {
            'serial_number': '',
            'warranty_end': None,
            'is_valid': False,
            'source': None,
        }
        if customer:
            # CustomerPurchase에서 해당 제품 구매내역 조회
            purchase = CustomerPurchase.objects.filter(
                customer=customer, product=item.product, is_active=True,
            ).first()
            if purchase:
                warranty_info['serial_number'] = purchase.serial_number or ''
                warranty_info['warranty_end'] = purchase.warranty_end
                if purchase.warranty_end:
                    warranty_info['is_valid'] = purchase.warranty_end >= today
                    warranty_info['source'] = 'customer'

                # ProductRegistration 보완
                if purchase.serial_number:
                    try:
                        reg = ProductRegistration.objects.get(
                            serial_number=purchase.serial_number,
                            product=item.product,
                        )
                        warranty_info['warranty_end'] = reg.warranty_end
                        warranty_info['is_valid'] = reg.warranty_end >= today
                        warranty_info['source'] = 'registration'
                        warranty_info['registration'] = reg
                    except ProductRegistration.DoesNotExist:
                        pass

        ctx['warranty_info'] = warranty_info

        # AS 이력
        service_requests = []
        if customer:
            service_requests = ServiceRequest.objects.filter(
                customer=customer,
                product=item.product,
            ).prefetch_related('repairs').order_by('-received_date')

        ctx['service_requests'] = service_requests
        return ctx


class SoldProductExcelView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel

        qs = _build_sold_product_queryset(request)
        enriched = _enrich_sold_items(qs[:5000])  # 최대 5000건

        # 보증상태 필터
        warranty_filter = request.GET.get('warranty_status', '').strip()
        if warranty_filter == 'valid':
            enriched = [r for r in enriched if r['warranty_status'] == 'valid']
        elif warranty_filter == 'expired':
            enriched = [r for r in enriched if r['warranty_status'] == 'expired']

        headers = [
            ('주문번호', 18), ('주문일', 12), ('제품코드', 15), ('제품명', 25),
            ('수량', 8), ('단가', 15), ('금액', 15), ('고객명', 15),
            ('연락처', 15), ('시리얼번호', 20), ('보증만료일', 12),
            ('보증상태', 10), ('AS건수', 8),
        ]

        warranty_label = {'valid': '유효', 'expired': '만료', 'none': '미등록'}

        rows = []
        for r in enriched:
            rows.append([
                r['order'].order_number,
                r['order'].order_date.strftime('%Y-%m-%d') if r['order'].order_date else '',
                r['product'].code,
                r['product'].name,
                r['item'].quantity,
                int(r['item'].unit_price),
                int(r['item'].amount),
                r['customer_name'],
                r['phone'],
                r['serial_number'],
                r['warranty_end'].strftime('%Y-%m-%d') if r['warranty_end'] else '',
                warranty_label.get(r['warranty_status'], '미등록'),
                r['service_count'],
            ])

        return export_to_excel(
            '판매제품목록', headers, rows,
            money_columns=[5, 6],
        )


# ── 견적서 ────────────────────────────────────────────────

class QuotationListView(LoginRequiredMixin, ListView):
    model = Quotation
    template_name = 'sales/quote_list.html'
    context_object_name = 'quotes'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related('partner', 'customer')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(quote_number__icontains=q)
                | Q(partner__name__icontains=q)
                | Q(customer__name__icontains=q)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs


class QuotationCreateView(ManagerRequiredMixin, CreateView):
    model = Quotation
    form_class = QuotationForm
    template_name = 'sales/quote_form.html'
    success_url = reverse_lazy('sales:quote_list')

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['quote_number'] = generate_document_number(
            Quotation, 'quote_number', 'QT',
        )
        from apps.accounting.models import BankAccount
        default_bank = BankAccount.objects.filter(
            is_active=True, is_default=True,
        ).first()
        if default_bank:
            initial['bank_account'] = default_bank.pk
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = QuotationItemFormSet(self.request.POST)
        else:
            ctx['formset'] = QuotationItemFormSet()
        ctx['product_units_json'] = _product_units_json()
        ctx['product_costs_json'] = _product_costs_json()
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            formset.instance = self.object
            formset.save()
            self.object.update_total()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class QuotationDetailView(LoginRequiredMixin, DetailView):
    model = Quotation
    template_name = 'sales/quote_detail.html'
    context_object_name = 'quote'
    slug_field = 'quote_number'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'partner', 'customer', 'bank_account',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        items = self.object.quote_items.select_related('product')
        ctx['items'] = items
        ctx['total_cost'] = sum(i.total_cost for i in items)
        ctx['total_profit'] = sum(i.profit for i in items)
        total_amount = int(self.object.total_amount)
        ctx['total_profit_rate'] = (
            round(ctx['total_profit'] / total_amount * 100, 1)
            if total_amount else 0
        )
        return ctx


class QuotationUpdateView(ManagerRequiredMixin, UpdateView):
    model = Quotation
    form_class = QuotationForm
    template_name = 'sales/quote_form.html'
    success_url = reverse_lazy('sales:quote_list')
    slug_field = 'quote_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['formset'] = QuotationItemFormSet(
                self.request.POST, instance=self.object,
            )
        else:
            ctx['formset'] = QuotationItemFormSet(instance=self.object)
        ctx['product_units_json'] = _product_units_json()
        ctx['product_costs_json'] = _product_costs_json()
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            # vat_included 변경 시 모든 항목 VAT 재계산
            for item in self.object.quote_items.all():
                item.save()
            self.object.update_total()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class QuotationConvertView(ManagerRequiredMixin, View):
    """견적서 → 주문 전환"""

    def post(self, request, slug):
        quote = get_object_or_404(Quotation, quote_number=slug)
        if quote.status == Quotation.Status.CONVERTED:
            messages.error(request, '이미 주문으로 전환된 견적입니다.')
            return HttpResponseRedirect(
                reverse_lazy('sales:quote_detail', kwargs={'slug': slug})
            )
        if quote.valid_until and quote.valid_until < date.today():
            messages.error(request, '유효기한이 만료된 견적서는 주문으로 전환할 수 없습니다.')
            return HttpResponseRedirect(
                reverse_lazy('sales:quote_detail', kwargs={'slug': slug})
            )
        if quote.status in (Quotation.Status.REJECTED, Quotation.Status.EXPIRED):
            messages.error(request, f'{quote.get_status_display()} 상태의 견적서는 주문으로 전환할 수 없습니다.')
            return HttpResponseRedirect(
                reverse_lazy('sales:quote_detail', kwargs={'slug': slug})
            )

        from apps.core.utils import generate_document_number

        with transaction.atomic():
            order = Order.objects.create(
                order_number=generate_document_number(
                    Order, 'order_number', 'ORD',
                    reference_date=quote.quote_date,
                ),
                partner=quote.partner,
                customer=quote.customer,
                order_date=date.today(),
                delivery_date=quote.valid_until,
                status=Order.Status.DRAFT,
                vat_included=quote.vat_included,
                bank_account=quote.bank_account,
                shipping_address=getattr(quote, 'shipping_address', ''),
                created_by=request.user,
            )

            for qi in quote.quote_items.filter(is_active=True):
                OrderItem.objects.create(
                    order=order,
                    product=qi.product,
                    quantity=qi.quantity,
                    cost_price=qi.cost_price,
                    unit_price=qi.unit_price,
                    discount_rate=qi.discount_rate,
                    discount_amount=qi.discount_amount,
                    created_by=request.user,
                )

            order.update_total()
            quote.status = Quotation.Status.CONVERTED
            quote.converted_order = order
            quote.save(update_fields=[
                'status', 'converted_order', 'updated_at',
            ])

            # 마켓플레이스 주문 연결 (견적서 경유 → ERP 주문) + 채널/결제수단 자동매핑
            from apps.marketplace.models import MarketplaceOrder
            mkt_qs = MarketplaceOrder.objects.filter(
                erp_quotation=quote, is_active=True,
            ).select_related('import_session')
            mkt_qs.update(erp_order=order)
            first_mkt = mkt_qs.first()
            if first_mkt and first_mkt.import_session:
                platform = (first_mkt.import_session.platform or '').upper()
                channel_map = {
                    'NAVER': (Order.SalesChannel.NAVER, Order.PaymentMethod.NAVER_PAY),
                    'COUPANG': (Order.SalesChannel.COUPANG, Order.PaymentMethod.PLATFORM),
                }
                if platform in channel_map:
                    order.sales_channel, order.payment_method = channel_map[platform]
                    order.save(update_fields=[
                        'sales_channel', 'payment_method', 'updated_at',
                    ])

        messages.success(
            request,
            f'주문 {order.order_number}으로 전환되었습니다.',
        )
        return HttpResponseRedirect(
            reverse_lazy('sales:order_detail', kwargs={'slug': order.order_number})
        )


class QuotationImportView(ManagerRequiredMixin, TemplateView):
    template_name = 'core/import.html'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            from apps.core.import_views import export_resource_data
            from .resources import QuotationResource
            return export_resource_data(QuotationResource(), '견적_데이터')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = '견적서 일괄 가져오기'
        ctx['cancel_url'] = reverse_lazy('sales:quote_list')
        ctx['sample_url'] = reverse_lazy(
            'sales:quote_import_sample',
        )
        ctx['field_hints'] = [
            '견적번호(quote_number)가 동일하면 수정됩니다.',
            'partner_code: 거래처코드',
            'customer_name: 고객명',
            'status: DRAFT(작성중), SENT(발송), '
            'ACCEPTED(수락), REJECTED(거절)',
        ]
        ctx['supports_export'] = True
        return ctx

    def post(self, request, *args, **kwargs):
        from .resources import QuotationResource
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)
        resource = QuotationResource()
        try:
            data = _parse_import_file(request, import_file)
            if data is None:
                messages.error(
                    request, '지원하지 않는 파일 형식입니다.',
                )
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError,
                UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)
        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = _build_preview(result, data)
            ctx['errors'] = _collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = _collect_errors(result)
                return self.render_to_response(ctx)
            total = (result.totals.get('new', 0)
                     + result.totals.get('update', 0))
            messages.success(
                request, f'{total}건 가져오기 완료.',
            )
            return HttpResponseRedirect(
                str(reverse_lazy('sales:quote_list')),
            )


class QuotationImportSampleView(ManagerRequiredMixin, View):
    def get(self, request):
        from apps.core.excel import export_to_excel
        headers = [
            ('quote_number', 18), ('partner_code', 15),
            ('customer_code', 15), ('quote_date', 12),
            ('valid_until', 12), ('status', 10),
            ('notes', 30),
        ]
        rows = [
            ['QT-2026-001', 'P-001', 'CST-0001',
             '2026-03-01', '2026-04-01', 'DRAFT', ''],
        ]
        return export_to_excel(
            '견적서_가져오기_양식', headers, rows,
            filename='견적서_가져오기_양식.xlsx',
            required_columns=[0, 3, 4],  # quote_number, quote_date, valid_until
        )


# ── 배송 추적 ─────────────────────────────────────────────

class ShipmentListView(LoginRequiredMixin, ListView):
    model = Shipment
    template_name = 'sales/shipment_list.html'
    context_object_name = 'shipments'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True).select_related('order', 'order__partner')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(shipment_number__icontains=q)
                | Q(tracking_number__icontains=q)
                | Q(order__order_number__icontains=q)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs


class ShipmentCreateView(ManagerRequiredMixin, CreateView):
    model = Shipment
    fields = [
        'shipment_number', 'shipping_type', 'carrier', 'tracking_number',
        'receiver_name', 'receiver_phone',
        'receiver_address', 'receiver_address_road', 'receiver_address_detail',
        'notes',
    ]
    template_name = 'sales/shipment_form.html'

    def get_initial(self):
        initial = super().get_initial()
        from apps.core.utils import generate_document_number
        initial['shipment_number'] = generate_document_number(Shipment, 'shipment_number', 'SH')
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        for field in form.fields.values():
            field.widget.attrs.setdefault('class', 'form-input')
        return form

    def form_valid(self, form):
        order = get_object_or_404(Order, order_number=self.kwargs['order_slug'])
        form.instance.order = order
        form.instance.created_by = self.request.user
        form.instance.status = Shipment.Status.PREPARING
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            'sales:order_detail',
            kwargs={'slug': self.kwargs['order_slug']},
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order = get_object_or_404(
            Order.objects.select_related('customer'),
            order_number=self.kwargs['order_slug'],
        )
        ctx['order'] = order
        # 주문서 배송정보 (JS 자동채우기용)
        customer = order.customer
        name = customer.name if customer else ''
        phone = customer.phone if customer else ''
        road = order.shipping_address_road or ''
        detail = order.shipping_address_detail or ''
        if not road and order.shipping_address:
            parts = order.shipping_address.split('\n')
            road = parts[0].strip() if parts else ''
            detail = parts[1].strip() if len(parts) > 1 else ''
        if name or road:
            ctx['orderer_info'] = {
                'name': name,
                'phone': phone,
                'address_road': road,
                'address_detail': detail,
                'address': order.shipping_address or f'{road} {detail}'.strip(),
            }
        else:
            ctx['orderer_info'] = None
        return ctx


class ShipmentDetailView(LoginRequiredMixin, DetailView):
    model = Shipment
    template_name = 'sales/shipment_detail.html'
    context_object_name = 'shipment'
    slug_field = 'shipment_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.inventory.models import SerialNumber
        serial_info = []
        items = self.object.items.filter(
            is_active=True,
        ).select_related('order_item__product')
        for item in items:
            serials = SerialNumber.objects.filter(
                shipment_item=item, is_active=True,
            ).order_by('serial') if hasattr(SerialNumber, 'shipment_item') else SerialNumber.objects.none()
            if serials.exists():
                serial_info.append({
                    'item': item,
                    'serials': serials,
                    'range_start': serials.first().serial,
                    'range_end': serials.last().serial,
                    'count': serials.count(),
                })
        ctx['serial_info'] = serial_info
        return ctx


class ShipmentUpdateView(ManagerRequiredMixin, UpdateView):
    model = Shipment
    fields = [
        'shipping_type', 'carrier', 'tracking_number', 'status',
        'shipped_date', 'delivered_date',
        'receiver_name', 'receiver_phone',
        'receiver_address', 'receiver_address_road', 'receiver_address_detail',
        'notes',
    ]
    template_name = 'sales/shipment_form.html'
    slug_field = 'shipment_number'
    slug_url_kwarg = 'slug'

    # Shipment 모델에 STATUS_TRANSITIONS가 없으므로 뷰에서 검증
    SHIPMENT_STATUS_TRANSITIONS = {
        'PREPARING': ['SHIPPED'],
        'SHIPPED': ['IN_TRANSIT', 'DELIVERED'],
        'IN_TRANSIT': ['DELIVERED'],
        'DELIVERED': [],
        'RETURNED': [],
    }

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        for field in form.fields.values():
            field.widget.attrs.setdefault('class', 'form-input')
        return form

    def form_valid(self, form):
        old_status = self.get_object().status
        new_status = form.cleaned_data.get('status', old_status)
        if old_status != new_status:
            allowed = self.SHIPMENT_STATUS_TRANSITIONS.get(old_status, [])
            if new_status not in allowed:
                form.add_error(
                    'status',
                    f'{self.object.get_status_display()}에서 '
                    f'{dict(Shipment.Status.choices).get(new_status, new_status)}(으)로 '
                    f'변경할 수 없습니다.',
                )
                return self.form_invalid(form)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            'sales:shipment_detail', kwargs={'slug': self.object.shipment_number},
        )


class PartialShipmentView(ManagerRequiredMixin, TemplateView):
    """부분 출고 — 주문항목별 출고수량 지정"""
    template_name = 'sales/partial_shipment.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order = get_object_or_404(
            Order.objects.prefetch_related('items__product'),
            order_number=self.kwargs['slug'],
        )
        ctx['order'] = order
        ctx['items'] = [
            item for item in order.items.all()
            if item.remaining_quantity > 0
        ]
        return ctx

    def post(self, request, slug):
        order = get_object_or_404(Order, order_number=slug)
        if order.status not in ('CONFIRMED', 'PARTIAL_SHIPPED'):
            messages.error(request, '출고 가능한 상태가 아닙니다.')
            return redirect('sales:order_detail', slug=slug)

        from apps.sales.signals import InsufficientStockError
        from apps.core.utils import generate_document_number

        try:
            with transaction.atomic():
                shipment = Shipment.objects.create(
                    order=order,
                    shipping_type='PARCEL',
                    status='SHIPPED',
                    shipped_date=date.today(),
                    receiver_name=request.POST.get(
                        'receiver_name', '',
                    ),
                    tracking_number=request.POST.get(
                        'tracking_number', '',
                    ),
                    created_by=request.user,
                )

                created_any = False
                for item in order.items.all():
                    qty_str = request.POST.get(
                        f'qty_{item.pk}', '0',
                    )
                    qty = int(qty_str) if qty_str else 0
                    if qty <= 0:
                        continue
                    remaining = item.remaining_quantity
                    if qty > remaining:
                        qty = remaining
                    ShipmentItem.objects.create(
                        shipment=shipment,
                        order_item=item,
                        quantity=qty,
                        created_by=request.user,
                    )
                    created_any = True

                if not created_any:
                    raise ValueError(
                        '출고할 항목이 없습니다.',
                    )

            messages.success(
                request,
                f'부분 출고 완료 ({shipment.shipment_number})',
            )
        except InsufficientStockError as e:
            messages.error(request, str(e))
        except ValueError as e:
            messages.error(request, str(e))

        return redirect('sales:order_detail', slug=slug)


class PartnerAnalysisView(ManagerRequiredMixin, TemplateView):
    """거래처별 매출/수익 분석"""
    template_name = 'sales/partner_analysis.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # 기간 필터 (기본: 올해)
        date_from = self.request.GET.get('date_from', f'{today.year}-01-01')
        date_to = self.request.GET.get('date_to', today.strftime('%Y-%m-%d'))
        sort_by = self.request.GET.get('sort', '-revenue')

        context['date_from'] = date_from
        context['date_to'] = date_to
        context['sort_by'] = sort_by

        # 거래처별 집계: OrderItem 기반
        partner_stats = (
            OrderItem.objects.filter(
                order__is_active=True,
                order__partner__isnull=False,
                order__status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED', 'CLOSED'],
                order__order_date__gte=date_from,
                order__order_date__lte=date_to,
                is_active=True,
            )
            .values(
                'order__partner__id',
                'order__partner__name',
                'order__partner__code',
            )
            .annotate(
                order_count=Count('order', distinct=True),
                revenue=Sum('amount'),
                total_cost=Sum(
                    ExpressionWrapper(
                        F('quantity') * F('cost_price'),
                        output_field=DecimalField(max_digits=20, decimal_places=0),
                    )
                ),
            )
            .order_by()  # clear default ordering
        )

        # 후처리: 이익, 이익률 계산
        results = []
        total_revenue = Decimal('0')
        total_cost_sum = Decimal('0')
        total_orders = 0
        for row in partner_stats:
            revenue = row['revenue'] or Decimal('0')
            cost = row['total_cost'] or Decimal('0')
            profit = revenue - cost
            profit_rate = (
                round(float(profit) / float(revenue) * 100, 1)
                if revenue > 0 else 0
            )
            results.append({
                'partner_id': row['order__partner__id'],
                'partner_name': row['order__partner__name'],
                'partner_code': row['order__partner__code'],
                'order_count': row['order_count'],
                'revenue': revenue,
                'cost': cost,
                'profit': profit,
                'profit_rate': profit_rate,
            })
            total_revenue += revenue
            total_cost_sum += cost
            total_orders += row['order_count']

        # 정렬
        sort_map = {
            'name': ('partner_name', False),
            '-name': ('partner_name', True),
            'orders': ('order_count', False),
            '-orders': ('order_count', True),
            'revenue': ('revenue', False),
            '-revenue': ('revenue', True),
            'cost': ('cost', False),
            '-cost': ('cost', True),
            'profit': ('profit', False),
            '-profit': ('profit', True),
            'profit_rate': ('profit_rate', False),
            '-profit_rate': ('profit_rate', True),
        }
        sort_key, sort_reverse = sort_map.get(sort_by, ('revenue', True))
        results.sort(key=lambda x: x[sort_key], reverse=sort_reverse)

        context['partner_stats'] = results
        context['total_revenue'] = total_revenue
        context['total_cost'] = total_cost_sum
        context['total_profit'] = total_revenue - total_cost_sum
        context['total_profit_rate'] = (
            round(float(total_revenue - total_cost_sum) / float(total_revenue) * 100, 1)
            if total_revenue > 0 else 0
        )
        context['total_orders'] = total_orders

        return context


# ── 택배사 관리 ────────────────────────────────────────────


class ShippingCarrierListView(LoginRequiredMixin, ListView):
    model = ShippingCarrier
    template_name = 'sales/carrier_list.html'
    context_object_name = 'carriers'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)


class ShippingCarrierCreateView(ManagerRequiredMixin, CreateView):
    model = ShippingCarrier
    form_class = ShippingCarrierForm
    template_name = 'sales/carrier_form.html'
    success_url = reverse_lazy('sales:carrier_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class ShippingCarrierUpdateView(ManagerRequiredMixin, UpdateView):
    model = ShippingCarrier
    form_class = ShippingCarrierForm
    template_name = 'sales/carrier_form.html'
    success_url = reverse_lazy('sales:carrier_list')


# ── 배송 추적 ─────────────────────────────────────────────


class ShipmentTrackingView(LoginRequiredMixin, DetailView):
    """배송 추적 조회 뷰 (Shipment 상세에서 접근)"""
    model = Shipment
    template_name = 'sales/shipment_tracking.html'
    context_object_name = 'shipment'
    slug_field = 'shipment_number'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tracking_list'] = self.object.tracking_history.filter(
            is_active=True,
        ).order_by('-tracked_at')
        return ctx


# ── 가격규칙 ─────────────────────────────────────────────


class PriceRuleListView(ManagerRequiredMixin, ListView):
    model = PriceRule
    template_name = 'sales/price_rule_list.html'
    context_object_name = 'price_rules'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('product', 'partner', 'customer')
        product_id = self.request.GET.get('product')
        partner_id = self.request.GET.get('partner')
        if product_id:
            qs = qs.filter(product_id=product_id)
        if partner_id:
            qs = qs.filter(partner_id=partner_id)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['products'] = Product.objects.filter(is_active=True).order_by('name')
        ctx['partners'] = Partner.objects.filter(is_active=True).order_by('name')
        return ctx


class PriceRuleCreateView(ManagerRequiredMixin, CreateView):
    model = PriceRule
    form_class = PriceRuleForm
    template_name = 'sales/price_rule_form.html'
    success_url = reverse_lazy('sales:price_rule_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, '가격규칙이 등록되었습니다.')
        return super().form_valid(form)


class PriceRuleUpdateView(ManagerRequiredMixin, UpdateView):
    model = PriceRule
    form_class = PriceRuleForm
    template_name = 'sales/price_rule_form.html'
    success_url = reverse_lazy('sales:price_rule_list')

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True)

    def form_valid(self, form):
        messages.success(self.request, '가격규칙이 수정되었습니다.')
        return super().form_valid(form)


class PriceRuleDeleteView(_SoftDeleteView):
    model = PriceRule
    success_url = reverse_lazy('sales:price_rule_list')


class PriceLookupView(LoginRequiredMixin, View):
    def get(self, request):
        from django.http import JsonResponse
        from apps.sales.pricing import get_applicable_price

        product_id = request.GET.get('product_id')
        partner_id = request.GET.get('partner_id')
        customer_id = request.GET.get('customer_id')
        try:
            quantity = int(request.GET.get('quantity', 1))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid quantity'}, status=400)

        try:
            product = Product.objects.get(pk=product_id, is_active=True)
        except (Product.DoesNotExist, TypeError, ValueError):
            return JsonResponse({'error': 'Invalid product'}, status=400)

        partner = None
        customer = None
        if partner_id:
            partner = Partner.objects.filter(pk=partner_id, is_active=True).first()
        if customer_id:
            customer = Customer.objects.filter(pk=customer_id, is_active=True).first()

        result = get_applicable_price(product, partner, customer, quantity)
        return JsonResponse(result)


class CustomerAddressView(LoginRequiredMixin, View):
    """고객 주소 정보 반환 (주문폼 AJAX용)"""
    def get(self, request):
        from django.http import JsonResponse
        customer_id = request.GET.get('customer_id')
        if not customer_id:
            return JsonResponse({'address': '', 'phone': ''})
        customer = Customer.objects.filter(pk=customer_id, is_active=True).first()
        if not customer:
            return JsonResponse({'address': '', 'phone': ''})
        road = customer.address_road or ''
        detail = customer.address_detail or ''
        # address_road가 없고 address만 있으면 줄바꿈으로 분리
        if not road and customer.address:
            parts = customer.address.split('\n')
            road = parts[0].strip() if parts else ''
            detail = parts[1].strip() if len(parts) > 1 else ''
        return JsonResponse({
            'address': customer.address or f'{road} {detail}'.strip(),
            'address_road': road,
            'address_detail': detail,
            'phone': customer.phone or '',
        })


# ───────────────────────────────────────────────
# 고객등급 (CustomerTier)
# ───────────────────────────────────────────────

class CustomerTierListView(ManagerRequiredMixin, ListView):
    model = CustomerTier
    template_name = 'sales/customer_tier_list.html'
    context_object_name = 'tiers'
    paginate_by = 20

    def get_queryset(self):
        return super().get_queryset().filter(is_active=True).order_by('sort_order')


class CustomerTierCreateView(ManagerRequiredMixin, CreateView):
    model = CustomerTier
    form_class = CustomerTierForm
    template_name = 'sales/customer_tier_form.html'
    success_url = reverse_lazy('sales:customer_tier_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CustomerTierUpdateView(ManagerRequiredMixin, UpdateView):
    model = CustomerTier
    form_class = CustomerTierForm
    template_name = 'sales/customer_tier_form.html'
    success_url = reverse_lazy('sales:customer_tier_list')


# ───────────────────────────────────────────────
# 영업목표 (SalesTarget)
# ───────────────────────────────────────────────

class SalesTargetListView(ManagerRequiredMixin, ListView):
    model = SalesTarget
    template_name = 'sales/sales_target_list.html'
    context_object_name = 'targets'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('salesperson').order_by('-year', 'quarter')
        year = self.request.GET.get('year')
        if year:
            qs = qs.filter(year=year)
        return qs


class SalesTargetCreateView(ManagerRequiredMixin, CreateView):
    model = SalesTarget
    form_class = SalesTargetForm
    template_name = 'sales/sales_target_form.html'
    success_url = reverse_lazy('sales:sales_target_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class SalesTargetDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'sales/sales_target_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.utils import timezone
        year = int(self.request.GET.get('year', timezone.now().year))
        targets = SalesTarget.objects.filter(
            is_active=True, year=year,
        ).select_related('salesperson')
        ctx['year'] = year
        ctx['targets'] = targets
        ctx['years'] = sorted(
            SalesTarget.objects.filter(is_active=True)
            .values_list('year', flat=True).distinct(),
            reverse=True,
        )
        return ctx


# ───────────────────────────────────────────────
# 영업리드 (SalesLead)
# ───────────────────────────────────────────────

class SalesLeadListView(ManagerRequiredMixin, ListView):
    model = SalesLead
    template_name = 'sales/sales_lead_list.html'
    context_object_name = 'leads'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('assigned_to').order_by('-created_at')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(company_name__icontains=q) | Q(contact_name__icontains=q)
                | Q(lead_number__icontains=q)
            )
        return qs


class SalesLeadCreateView(ManagerRequiredMixin, CreateView):
    model = SalesLead
    form_class = SalesLeadForm
    template_name = 'sales/sales_lead_form.html'
    success_url = reverse_lazy('sales:lead_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class SalesLeadDetailView(ManagerRequiredMixin, DetailView):
    model = SalesLead
    template_name = 'sales/sales_lead_detail.html'
    context_object_name = 'lead'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['activities'] = self.object.activities.filter(
            is_active=True,
        ).order_by('-activity_date')
        ctx['activity_form'] = LeadActivityForm()
        return ctx

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = LeadActivityForm(request.POST)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.lead = self.object
            activity.created_by = request.user
            activity.save()
            messages.success(request, '활동이 기록되었습니다.')
            return redirect('sales:lead_detail', pk=self.object.pk)
        ctx = self.get_context_data()
        ctx['activity_form'] = form
        return self.render_to_response(ctx)


class SalesLeadUpdateView(ManagerRequiredMixin, UpdateView):
    model = SalesLead
    form_class = SalesLeadForm
    template_name = 'sales/sales_lead_form.html'

    def get_success_url(self):
        return reverse_lazy('sales:lead_detail', kwargs={'pk': self.object.pk})


class LeadConvertView(ManagerRequiredMixin, View):
    """Lead -> Quotation 전환"""

    def post(self, request, pk):
        lead = get_object_or_404(SalesLead, pk=pk, is_active=True)
        with transaction.atomic():
            partner, _ = Partner.objects.get_or_create(
                name=lead.company_name,
                defaults={
                    'code': Partner.generate_next_code('CUSTOMER'),
                    'partner_type': 'CUSTOMER',
                    'contact_name': lead.contact_name,
                    'email': lead.contact_email,
                    'phone': lead.contact_phone,
                    'created_by': request.user,
                },
            )
            quote = Quotation.objects.create(
                partner=partner,
                quote_date=date.today(),
                valid_until=date.today(),
                created_by=request.user,
            )
            lead.status = 'WON'
            lead.save(update_fields=['status', 'updated_at'])
        messages.success(
            request,
            f'리드 {lead.lead_number}이(가) 견적서 {quote.quote_number}(으)로 전환되었습니다.',
        )
        return redirect('sales:quote_detail', slug=quote.quote_number)


class PipelineBoardView(ManagerRequiredMixin, TemplateView):
    """영업 파이프라인 칸반 보드"""
    template_name = 'sales/pipeline_board.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.http import JsonResponse
        stages = dict(SalesLead.STATUS_CHOICES)
        pipeline = {}
        for code, label in stages.items():
            leads = SalesLead.objects.filter(
                is_active=True, status=code,
            ).select_related('assigned_to').order_by('-expected_amount')
            pipeline[code] = {
                'label': label,
                'leads': leads,
                'count': leads.count(),
                'total': sum(int(l.expected_amount) for l in leads),
            }
        ctx['pipeline'] = pipeline
        ctx['stages'] = stages
        return ctx


# ───────────────────────────────────────────────
# 고객만족도 (CustomerSatisfaction)
# ───────────────────────────────────────────────

class CustomerSatisfactionListView(ManagerRequiredMixin, ListView):
    model = CustomerSatisfaction
    template_name = 'sales/satisfaction_list.html'
    context_object_name = 'surveys'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_active=True,
        ).select_related('partner', 'order').order_by('-survey_date')
        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(category=category)
        return qs


class CustomerSatisfactionCreateView(ManagerRequiredMixin, CreateView):
    model = CustomerSatisfaction
    form_class = CustomerSatisfactionForm
    template_name = 'sales/satisfaction_form.html'
    success_url = reverse_lazy('sales:satisfaction_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class SatisfactionDashboardView(ManagerRequiredMixin, TemplateView):
    """고객만족도 대시보드 - NPS, 카테고리별 평균"""
    template_name = 'sales/satisfaction_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.db.models import Avg
        qs = CustomerSatisfaction.objects.filter(is_active=True)
        ctx['avg_score'] = qs.aggregate(avg=Avg('score'))['avg'] or 0
        ctx['avg_nps'] = qs.aggregate(avg=Avg('nps_score'))['avg'] or 0
        ctx['category_stats'] = list(
            qs.values('category').annotate(
                avg_score=Avg('score'),
                avg_nps=Avg('nps_score'),
                count=Count('id'),
            ).order_by('category')
        )
        ctx['category_labels'] = dict(CustomerSatisfaction.CATEGORY_CHOICES)
        ctx['total_count'] = qs.count()
        return ctx
