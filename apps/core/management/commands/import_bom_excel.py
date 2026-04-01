"""BOM-BEP Excel 파일에서 실제 제품 데이터를 일괄 임포트하는 커맨드"""
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

User = get_user_model()


class Command(BaseCommand):
    help = 'BOM-BEP Excel 파일(.xlsx)에서 카테고리, 거래처, 제품, BOM, 고정비를 일괄 임포트합니다.'

    def add_arguments(self, parser):
        parser.add_argument(
            'file', type=str,
            help='임포트할 Excel 파일 경로 (예: local/BOM-BEP-PnL예상\\ 계산-V2.xlsx)',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='실제 DB에 저장하지 않고 파싱 결과만 출력',
        )
        parser.add_argument(
            '--price-scenario', type=str, default='B', choices=['A', 'B', 'C'],
            help='BEP 시나리오별 판매가 선택 (A=79000, B=89000, C=129000, 기본: B)',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl이 설치되어 있지 않습니다.')

        file_path = options['file']
        dry_run = options['dry_run']
        scenario = options['price_scenario']

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'파일을 찾을 수 없습니다: {file_path}')

        self.stdout.write(f'파일 로드 완료: {file_path}')
        self.stdout.write(f'시트: {", ".join(wb.sheetnames)}')
        self.stdout.write(f'판매가 시나리오: {scenario}')
        self.stdout.write('')

        # 시트 파싱
        bom_data = self._parse_bom_sheet(wb['BOM'])
        bep_data = self._parse_bep_sheet(wb['PRICE and BEP'], scenario)
        parts_data = self._parse_parts_sheet(wb['PARTS LIST'])

        # 파싱 결과 출력
        self._print_summary(bom_data, bep_data, parts_data)

        if dry_run:
            self.stdout.write(self.style.WARNING('\n--dry-run: DB에 저장하지 않습니다.'))
            return

        # DB 임포트
        admin = User.objects.filter(role='admin').first()
        if not admin:
            raise CommandError('admin 계정이 없습니다. 먼저 init_prod 또는 createsuperuser를 실행하세요.')

        with transaction.atomic():
            categories = self._create_categories(admin, parts_data)
            partners = self._create_partners(admin, bom_data)
            products = self._create_products(admin, bom_data, bep_data, parts_data, categories)
            bom = self._create_bom(admin, bom_data, products)
            fixed_costs = self._create_fixed_costs(admin, bep_data)

        # 리포트 가이드 생성
        self._generate_report(bom_data, bep_data, parts_data, scenario)

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('임포트 완료!'))
        self.stdout.write(f'  카테고리: {len(categories)}개')
        self.stdout.write(f'  거래처: {len(partners)}개')
        self.stdout.write(f'  제품: {len(products)}개 (완제품 1 + 원자재 {len(products) - 1})')
        self.stdout.write(f'  BOM: 1개, BOM 항목: {bom.items.count()}개')
        self.stdout.write(f'  고정비: {len(fixed_costs)}개')

    # ── 파싱 ──

    def _parse_bom_sheet(self, ws):
        """BOM 시트 파싱: 완제품명, 버전, 부품 목록"""
        data = {
            'product_name': ws['C2'].value,           # DiveChecker V1
            'version': ws['C4'].value,                # V1_20260205
            'author': ws['C5'].value,                 # 작성자
            'unit_cost': ws['I6'].value,              # 25007 (1대 비용 단가합)
            'total_cost': ws['J6'].value,             # 1654530 (1대 비용 전체합)
            'materials': [],
        }

        for row_num in range(10, ws.max_row + 1):
            name = ws[f'B{row_num}'].value
            qty = ws[f'D{row_num}'].value
            if not name or name == 'TEST' or not qty:
                continue
            data['materials'].append({
                'name': name,
                'base_qty': ws[f'C{row_num}'].value or 1,
                'qty_per_unit': qty,
                'unit_cost': ws[f'E{row_num}'].value or 0,
                'total_cost': ws[f'F{row_num}'].value or 0,
                'supplier': ws[f'M{row_num}'].value or '',
                'purchase_source': ws[f'R{row_num}'].value or '',
                'link': ws[f'S{row_num}'].value or '',
                'wholesale_cost': ws[f'I{row_num}'].value or 0,
                'wholesale_qty': ws[f'J{row_num}'].value or 1,
            })

        return data

    def _parse_bep_sheet(self, ws, scenario):
        """BEP 시트 파싱: 고정비, 변동비, 판매가 시나리오"""
        scenario_col = {'A': 'E', 'B': 'F', 'C': 'G'}[scenario]

        data = {
            'selling_price': ws[f'{scenario_col}24'].value or 0,
            'selling_qty': ws[f'{scenario_col}23'].value or 0,
            'fixed_costs': [],
            'variable_costs': [],
            'bep_qty': ws[f'{scenario_col}40'].value or 0,
            'profit_per_unit': ws[f'{scenario_col}43'].value or 0,
        }

        # 고정비 (rows 26-32)
        for row_num in range(26, 33):
            name = ws[f'B{row_num}'].value
            category = ws[f'C{row_num}'].value
            amount = ws[f'D{row_num}'].value
            if name and amount:
                data['fixed_costs'].append({
                    'name': name,
                    'category': category or '고정비',
                    'amount': int(amount),
                })

        # 변동비 (rows 33-36)
        for row_num in range(33, 37):
            name = ws[f'B{row_num}'].value
            category = ws[f'C{row_num}'].value
            per_unit = ws[f'D{row_num}'].value
            if name and per_unit is not None:
                data['variable_costs'].append({
                    'name': name,
                    'category': category or '변동비',
                    'per_unit': per_unit,
                })

        return data

    def _parse_parts_sheet(self, ws):
        """PARTS LIST 시트 파싱: 부품 카테고리, 상세 규격"""
        data = []
        for row_num in range(10, ws.max_row + 1):
            name = ws[f'D{row_num}'].value
            if not name:
                continue
            data.append({
                'group': ws[f'B{row_num}'].value or '',      # 본체
                'category': ws[f'C{row_num}'].value or '',    # 보드, 레진, 기타 ...
                'name': name,
                'spec_a': ws[f'E{row_num}'].value or '',
                'spec_b': ws[f'F{row_num}'].value or '',
                'spec_c': ws[f'G{row_num}'].value or '',
                'qty': ws[f'H{row_num}'].value or 1,
            })
        return data

    # ── 출력 ──

    def _print_summary(self, bom_data, bep_data, parts_data):
        self.stdout.write(self.style.MIGRATE_HEADING('── BOM ──'))
        self.stdout.write(f'  완제품: {bom_data["product_name"]}')
        self.stdout.write(f'  버전: {bom_data["version"]}')
        self.stdout.write(f'  1대 원가: {int(bom_data["unit_cost"] or 0):,}원')
        self.stdout.write(f'  부품 {len(bom_data["materials"])}종:')
        for m in bom_data['materials']:
            self.stdout.write(
                f'    - {m["name"]}: {m["qty_per_unit"]}EA × '
                f'{int(m["unit_cost"]):,}원 = {int(m["total_cost"]):,}원  [{m["supplier"]}]'
            )

        self.stdout.write('')
        self.stdout.write(self.style.MIGRATE_HEADING('── BEP ──'))
        self.stdout.write(f'  판매가: {int(bep_data["selling_price"]):,}원')
        self.stdout.write(f'  BEP 수량: {bep_data["bep_qty"]:.1f}대')
        self.stdout.write(f'  개당 순이익: {int(bep_data["profit_per_unit"]):,}원')
        self.stdout.write(f'  고정비 {len(bep_data["fixed_costs"])}항목:')
        for fc in bep_data['fixed_costs']:
            self.stdout.write(f'    - {fc["name"]}: {fc["amount"]:,}원')
        self.stdout.write(f'  변동비 {len(bep_data["variable_costs"])}항목:')
        for vc in bep_data['variable_costs']:
            self.stdout.write(f'    - {vc["name"]}: {vc["per_unit"]}')

        self.stdout.write('')
        self.stdout.write(self.style.MIGRATE_HEADING('── PARTS LIST ──'))
        categories = set(p['category'] for p in parts_data)
        self.stdout.write(f'  카테고리: {", ".join(categories)}')
        self.stdout.write(f'  부품: {len(parts_data)}종')

    # ── DB 생성 ──

    def _create_categories(self, admin, parts_data):
        from apps.inventory.models import Category

        cat_names = sorted(set(p['category'] for p in parts_data if p['category']))
        categories = {}
        for name in cat_names:
            cat, _ = Category.objects.get_or_create(
                name=name,
                defaults={'created_by': admin},
            )
            categories[name] = cat
        return categories

    def _create_partners(self, admin, bom_data):
        from apps.sales.models import Partner

        supplier_names = sorted(set(
            m['supplier'] for m in bom_data['materials']
            if m['supplier'] and m['supplier'] not in ('(미정)', '.', '직접인쇄')
        ))

        partners = {}
        for idx, name in enumerate(supplier_names, 1):
            code = f'SUP-{idx:03d}'
            partner, _ = Partner.objects.get_or_create(
                name=name,
                defaults={
                    'code': code,
                    'partner_type': 'SUPPLIER',
                    'created_by': admin,
                },
            )
            partners[name] = partner
        return partners

    def _create_products(self, admin, bom_data, bep_data, parts_data, categories):
        from apps.inventory.models import Product

        products = {}

        # 부품 카테고리 매핑 (PARTS LIST → name: category)
        part_cat_map = {p['name']: p['category'] for p in parts_data}
        part_spec_map = {
            p['name']: ' / '.join(
                str(v) for v in [p['spec_a'], p['spec_b'], p['spec_c']] if v
            )
            for p in parts_data
        }

        # 원자재 생성
        for idx, m in enumerate(bom_data['materials'], 1):
            code = f'MAT-{idx:03d}'
            cost = Decimal(str(m['unit_cost'])).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            cat_name = part_cat_map.get(m['name'], '')
            notes_parts = []
            if m['supplier']:
                notes_parts.append(f'거래처: {m["supplier"]}')
            if m['link'] and m['link'] not in ('(미정)', '직접인쇄', '오프라인'):
                notes_parts.append(f'구매링크: {m["link"]}')
            if m['purchase_source']:
                notes_parts.append(f'구매처: {m["purchase_source"]}')

            prod, _ = Product.objects.get_or_create(
                name=m['name'],
                defaults={
                    'code': code,
                    'product_type': 'RAW',
                    'category': categories.get(cat_name),
                    'cost_price': cost,
                    'unit_price': 0,
                    'specification': part_spec_map.get(m['name'], ''),
                    'notes': '\n'.join(notes_parts),
                    'created_by': admin,
                },
            )
            products[m['name']] = prod

        # 완제품 생성
        selling_price = Decimal(str(int(bep_data['selling_price'])))
        total_cost = Decimal(str(int(bom_data['unit_cost'] or 0)))
        finished, _ = Product.objects.get_or_create(
            name=bom_data['product_name'],
            defaults={
                'code': 'FIN-001',
                'product_type': 'FINISHED',
                'unit_price': selling_price,
                'cost_price': total_cost,
                'specification': f'버전: {bom_data["version"]}',
                'notes': f'작성자: {bom_data["author"]}',
                'created_by': admin,
            },
        )
        products[bom_data['product_name']] = finished

        return products

    def _create_bom(self, admin, bom_data, products):
        from apps.production.models import BOM, BOMItem

        finished = products[bom_data['product_name']]
        bom, created = BOM.objects.get_or_create(
            product=finished,
            version=bom_data['version'] or '1.0',
            defaults={
                'is_default': True,
                'created_by': admin,
            },
        )

        if created:
            for m in bom_data['materials']:
                material = products.get(m['name'])
                if material:
                    BOMItem.objects.create(
                        bom=bom,
                        material=material,
                        quantity=Decimal(str(m['qty_per_unit'])),
                        created_by=admin,
                    )

        return bom

    def _create_fixed_costs(self, admin, bep_data):
        from apps.accounting.models import FixedCost

        # 카테고리 매핑
        cat_map = {
            'Anycubic Photon M3 Max': 'EQUIPMENT',
            'KC 인증비용': 'OTHER',
            'Mac Mini': 'EQUIPMENT',
            'Apple Developer Program': 'SUBSCRIPTION',
            'Android Developers': 'SUBSCRIPTION',
            '포장비 고정비용': 'OTHER',
            '기타비용': 'OTHER',
        }

        today = date.today()
        month_first = today.replace(day=1)
        costs = []

        for fc in bep_data['fixed_costs']:
            category = cat_map.get(fc['name'], 'OTHER')
            cost, _ = FixedCost.objects.get_or_create(
                name=fc['name'],
                month=month_first,
                defaults={
                    'category': category,
                    'amount': Decimal(str(fc['amount'])),
                    'is_recurring': False,
                    'created_by': admin,
                },
            )
            costs.append(cost)

        return costs

    def _generate_report(self, bom_data, bep_data, parts_data, scenario):
        """Prod 수동 입력 가이드용 리포트 JSON 생성"""
        from apps.core.report import update_report_section

        cat_map = {
            'EQUIPMENT': '장비/감가상각',
            'SUBSCRIPTION': '구독/라이선스',
            'OTHER': '기타 고정비',
        }
        fc_cat_map = {
            'Anycubic Photon M3 Max': 'EQUIPMENT',
            'KC 인증비용': 'OTHER',
            'Mac Mini': 'EQUIPMENT',
            'Apple Developer Program': 'SUBSCRIPTION',
            'Android Developers': 'SUBSCRIPTION',
            '포장비 고정비용': 'OTHER',
            '기타비용': 'OTHER',
        }

        # 1) 거래처
        supplier_names = sorted(set(
            m['supplier'] for m in bom_data['materials']
            if m['supplier'] and m['supplier'] not in ('(미정)', '.', '직접인쇄')
        ))
        partner_items = []
        for idx, name in enumerate(supplier_names, 1):
            partner_items.append({
                'group': '거래처 (공급처)',
                'label': name,
                'nav_path': '판매관리 → 거래처 → + 거래처 등록',
                'fields': [
                    {'name': '거래처코드', 'value': f'SUP-{idx:03d}'},
                    {'name': '거래처명', 'value': name},
                    {'name': '유형', 'value': 'SUPPLIER (공급처)'},
                ],
            })
        update_report_section(
            'partners', '거래처 등록', f'공급처 {len(supplier_names)}개',
            partner_items,
        )

        # 2) 카테고리
        part_cat_map = {p['name']: p['category'] for p in parts_data}
        cat_names = sorted(set(p['category'] for p in parts_data if p['category']))
        cat_items = [{
            'group': '카테고리',
            'label': name,
            'nav_path': '재고관리 → 카테고리 (관리자 페이지)',
            'fields': [{'name': '카테고리명', 'value': name}],
        } for name in cat_names]
        update_report_section(
            'categories', '카테고리 등록', f'{len(cat_names)}개', cat_items,
        )

        # 3) 제품 (원자재)
        product_items = []
        for idx, m in enumerate(bom_data['materials'], 1):
            cat_name = part_cat_map.get(m['name'], '')
            product_items.append({
                'group': '원자재',
                'label': m['name'],
                'nav_path': '재고관리 → 제품 → + 제품 등록',
                'fields': [
                    {'name': '제품코드', 'value': f'MAT-{idx:03d}'},
                    {'name': '제품명', 'value': m['name']},
                    {'name': '유형', 'value': 'RAW (원자재)'},
                    {'name': '카테고리', 'value': cat_name or '-'},
                    {'name': '원가', 'value': f'{int(m["unit_cost"]):,}원'},
                    {'name': '거래처', 'value': m['supplier'] or '-'},
                    {'name': '구매링크', 'value': m.get('link', '-'),
                     'help': m.get('purchase_source', '')},
                ],
            })

        # 완제품
        product_items.append({
            'group': '완제품',
            'label': bom_data['product_name'],
            'nav_path': '재고관리 → 제품 → + 제품 등록',
            'fields': [
                {'name': '제품코드', 'value': 'FIN-001'},
                {'name': '제품명', 'value': bom_data['product_name']},
                {'name': '유형', 'value': 'FINISHED (완제품)'},
                {'name': '판매단가', 'value': f'{int(bep_data["selling_price"]):,}원',
                 'help': f'시나리오 {scenario}'},
                {'name': '원가', 'value': f'{int(bom_data["unit_cost"] or 0):,}원'},
                {'name': '버전', 'value': bom_data['version']},
            ],
        })
        update_report_section(
            'products', '제품 등록',
            f'원자재 {len(bom_data["materials"])}종 + 완제품 1종',
            product_items,
        )

        # 4) BOM
        bom_items = [{
            'group': 'BOM 헤더',
            'label': f'{bom_data["product_name"]} BOM',
            'nav_path': '생산관리 → BOM → + BOM 등록',
            'fields': [
                {'name': '완제품', 'value': bom_data['product_name']},
                {'name': '버전', 'value': bom_data['version']},
                {'name': '기본BOM', 'value': '예'},
            ],
        }]
        for m in bom_data['materials']:
            bom_items.append({
                'group': 'BOM 자재 (위 BOM에 추가)',
                'label': m['name'],
                'nav_path': 'BOM 상세 → 자재 추가',
                'fields': [
                    {'name': '자재', 'value': m['name']},
                    {'name': '소요량', 'value': str(m['qty_per_unit'])},
                    {'name': '단가', 'value': f'{int(m["unit_cost"]):,}원'},
                ],
            })
        update_report_section(
            'bom', 'BOM 등록',
            f'{bom_data["product_name"]} — 자재 {len(bom_data["materials"])}종',
            bom_items,
        )

        # 5) 고정비
        fc_items = []
        for fc in bep_data['fixed_costs']:
            cat_code = fc_cat_map.get(fc['name'], 'OTHER')
            fc_items.append({
                'group': '고정비',
                'label': fc['name'],
                'nav_path': '회계관리 → 고정비 → + 고정비 등록',
                'fields': [
                    {'name': '비용명', 'value': fc['name']},
                    {'name': '비용구분', 'value': f'{cat_code} ({cat_map.get(cat_code, cat_code)})'},
                    {'name': '금액', 'value': f'{fc["amount"]:,}원'},
                    {'name': '해당월', 'value': '당월 1일'},
                    {'name': '반복비용', 'value': '아니오'},
                ],
            })
        update_report_section(
            'fixed_costs', '고정비 등록',
            f'{len(bep_data["fixed_costs"])}항목 — BEP 손익분기 계산에 사용',
            fc_items,
        )

        # 6) BEP 요약
        bep_items = [{
            'group': 'BEP 참고 정보',
            'label': f'시나리오 {scenario} 요약',
            'nav_path': '회계관리 → 손익분기점',
            'fields': [
                {'name': '판매가', 'value': f'{int(bep_data["selling_price"]):,}원'},
                {'name': 'BEP 수량', 'value': f'{bep_data["bep_qty"]:.1f}대'},
                {'name': '개당 순이익', 'value': f'{int(bep_data["profit_per_unit"]):,}원'},
                {'name': '고정비 합계',
                 'value': f'{sum(fc["amount"] for fc in bep_data["fixed_costs"]):,}원'},
            ],
        }]
        for vc in bep_data['variable_costs']:
            bep_items.append({
                'group': 'BEP 변동비 참고',
                'label': vc['name'],
                'fields': [
                    {'name': '구분', 'value': vc['category']},
                    {'name': '단가/비율', 'value': str(vc['per_unit'])},
                ],
            })
        update_report_section(
            'bep_summary', 'BEP 손익분기 참고',
            '자동계산 — 제품/고정비 등록 후 회계관리 > 손익분기점에서 확인',
            bep_items,
        )

        self.stdout.write(self.style.SUCCESS('  리포트 가이드 생성 완료 → /report/'))
