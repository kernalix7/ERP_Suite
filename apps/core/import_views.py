"""
범용 Excel/CSV 일괄 가져오기 뷰 기반 클래스.

사용법:
    class MyImportView(BaseImportView):
        resource_class = MyResource
        page_title = '○○ 가져오기'
        cancel_url = reverse_lazy('app:list')
        sample_url = reverse_lazy('app:import_sample')
        field_hints = ['code가 동일하면 기존 데이터가 수정됩니다.']
"""
import io

import tablib
from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic import TemplateView
from openpyxl import load_workbook

from apps.core.mixins import ManagerRequiredMixin


def _find_header_row(ws):
    """export_to_excel 양식의 장식행(제목/날짜/빈행)을 건너뛰고 헤더 행 번호를 반환."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        non_empty = [c for c in row if c.value is not None]
        if len(non_empty) < 2:
            continue
        # 병합 셀(제목/날짜행)이면 건너뛰기
        if any(type(c).__name__ == 'MergedCell' for c in row):
            continue
        return row_idx
    return 1


def _load_xlsx(file_bytes):
    """xlsx를 읽어 장식행을 건너뛴 tablib Dataset을 반환."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    header_row = _find_header_row(ws)

    raw_headers = [
        str(c.value).strip() if c.value is not None else ''
        for c in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    ]
    # 필수 표시 (*) 제거 — 샘플 양식에서 "code (*)" → "code"
    headers = [
        h.replace(' (*)', '').replace('(*)', '') for h in raw_headers
    ]

    ds = tablib.Dataset(headers=headers)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        # None → '' 변환 (CSV와 동일하게, NOT NULL CharField 오류 방지)
        ds.append(['' if v is None else v for v in row])

    wb.close()
    return ds


def parse_import_file(import_file):
    """업로드된 파일을 tablib Dataset으로 파싱"""
    ext = import_file.name.rsplit('.', 1)[-1].lower()
    if ext == 'csv':
        return tablib.Dataset().load(
            import_file.read().decode('utf-8-sig'), format='csv',
        )
    elif ext in ('xlsx', 'xls'):
        return _load_xlsx(import_file.read())
    return None


def build_preview(result, data):
    """import-export 결과를 UI용 미리보기로 변환"""
    headers = list(data.headers) if data.headers else []
    rows = []
    for i, row_result in enumerate(result.rows):
        values = list(data[i]) if i < len(data) else []
        if row_result.errors:
            status = 'error'
        elif row_result.import_type == 'new':
            status = 'new'
        elif row_result.import_type == 'update':
            status = 'update'
        else:
            status = 'skip'
        rows.append({'status': status, 'values': values})

    totals = {
        'new': result.totals.get('new', 0),
        'update': result.totals.get('update', 0),
        'skip': result.totals.get('skip', 0),
        'error': result.totals.get('error', 0) + result.totals.get('invalid', 0),
    }
    return {
        'headers': headers,
        'rows': rows,
        'totals': totals,
        'has_valid_rows': totals['new'] > 0 or totals['update'] > 0,
        'file_name': '',
    }


def collect_errors(result):
    """import-export 결과에서 에러만 추출"""
    errors = []
    if result.base_errors:
        for err in result.base_errors:
            errors.append(str(err.error))
    for i, row in enumerate(result.rows):
        if row.errors:
            for err in row.errors:
                errors.append(f'{i + 1}행: {err.error}')
        if hasattr(row, 'validation_error') and row.validation_error:
            errors.append(f'{i + 1}행: {row.validation_error}')
    return errors


def export_resource_data(resource, filename='data_export'):
    """Resource 인스턴스를 이용해 기존 데이터를 가져오기 양식과 동일한 형식으로 내보내기."""
    qs = resource.get_queryset()
    if hasattr(qs.model, 'is_active'):
        qs = qs.filter(is_active=True)
    dataset = resource.export(queryset=qs)
    response = HttpResponse(
        dataset.xlsx,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


class BaseImportView(ManagerRequiredMixin, TemplateView):
    """범용 일괄 가져오기 뷰. 서브클래스에서 아래 속성을 지정하세요."""
    template_name = 'core/import.html'
    resource_class = None
    page_title = '일괄 가져오기'
    cancel_url = '/'
    sample_url = None
    field_hints = []
    success_message = '{count}건이 성공적으로 가져오기 되었습니다.'
    export_filename = 'data_export'

    def get(self, request, *args, **kwargs):
        if request.GET.get('action') == 'export':
            resource = self.get_resource()
            if resource:
                return export_resource_data(resource, self.export_filename)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = self.page_title
        ctx['cancel_url'] = self.cancel_url
        ctx['sample_url'] = self.sample_url
        ctx['field_hints'] = self.field_hints
        ctx['supports_export'] = True
        return ctx

    def get_resource(self):
        return self.resource_class()

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action', 'preview')
        import_file = request.FILES.get('import_file')

        if not import_file:
            messages.error(request, '파일을 선택해주세요.')
            return self.get(request, *args, **kwargs)

        resource = self.get_resource()

        try:
            data = parse_import_file(import_file)
            if data is None:
                messages.error(request, '지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv)')
                return self.get(request, *args, **kwargs)
        except (ValueError, KeyError, TypeError, UnicodeDecodeError) as e:
            messages.error(request, f'파일 읽기 오류: {e}')
            return self.get(request, *args, **kwargs)

        if action == 'preview':
            result = resource.import_data(data, dry_run=True)
            ctx = self.get_context_data(**kwargs)
            ctx['preview'] = build_preview(result, data)
            ctx['errors'] = collect_errors(result)
            return self.render_to_response(ctx)
        else:
            result = resource.import_data(data, dry_run=False)
            if result.has_errors():
                ctx = self.get_context_data(**kwargs)
                ctx['errors'] = collect_errors(result)
                return self.render_to_response(ctx)
            total = result.totals.get('new', 0) + result.totals.get('update', 0)
            messages.success(request, self.success_message.format(count=total))
            return HttpResponseRedirect(str(self.cancel_url))
