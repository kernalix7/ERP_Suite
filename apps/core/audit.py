"""감사 증적 관리 (Audit Trail) — ISMS 증적 / 회계 증빙 / 시스템 감사"""
import json
import logging
import re
from collections import defaultdict
from datetime import timedelta
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import models
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, TemplateView
from simple_history.models import HistoricalRecords

from apps.core.models import BaseModel

User = get_user_model()
logger = logging.getLogger('audit')


# ─────────────────────────────────────────────
# Model
# ─────────────────────────────────────────────

class AuditAccessLog(models.Model):
    """감사 증적 열람 기록 — 누가 어떤 감사 데이터를 언제 열람했는지"""

    class Section(models.TextChoices):
        DASHBOARD = 'DASHBOARD', '감사 대시보드'
        ACCESS_LOG = 'ACCESS_LOG', '시스템 접근 로그'
        DATA_CHANGE = 'DATA_CHANGE', '데이터 변경 이력'
        LOGIN_HISTORY = 'LOGIN_HISTORY', '로그인/보안 이벤트'
        AUDIT_LOG = 'AUDIT_LOG', '감사 열람 기록'

    user = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        verbose_name='열람자',
        on_delete=models.PROTECT,
        related_name='audit_access_logs',
    )
    section = models.CharField('열람 섹션', max_length=30, choices=Section.choices)
    ip_address = models.GenericIPAddressField('IP 주소', null=True, blank=True)
    user_agent = models.TextField('User-Agent', blank=True)
    query_params = models.TextField('조회 조건', blank=True)
    accessed_at = models.DateTimeField('열람일시', auto_now_add=True, db_index=True)

    class Meta:
        verbose_name = '감사 열람 기록'
        verbose_name_plural = '감사 열람 기록'
        ordering = ['-accessed_at']
        indexes = [
            models.Index(fields=['user', 'accessed_at'], name='idx_auditlog_user_date'),
            models.Index(fields=['section'], name='idx_auditlog_section'),
        ]

    def __str__(self):
        return f'{self.user} → {self.get_section_display()} ({self.accessed_at:%Y-%m-%d %H:%M})'


# ─────────────────────────────────────────────
# Mixin
# ─────────────────────────────────────────────

class AuditRequiredMixin(LoginRequiredMixin):
    """감사권한(is_auditor) 확인 + 열람 기록 자동 저장"""

    audit_section = ''  # 하위 클래스에서 지정

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.is_auditor:
            raise PermissionDenied('감사 증적 열람 권한이 없습니다. 관리자에게 감사권한을 요청하세요.')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        self._log_audit_access(request)
        return super().get(request, *args, **kwargs)

    def _log_audit_access(self, request):
        AuditAccessLog.objects.create(
            user=request.user,
            section=self.audit_section,
            ip_address=self._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
            query_params=request.GET.urlencode()[:500],
        )
        logger.info(
            'AUDIT_ACCESS user=%s section=%s ip=%s',
            request.user.username,
            self.audit_section,
            self._get_client_ip(request),
        )

    @staticmethod
    def _get_client_ip(request):
        xff = request.META.get('HTTP_X_FORWARDED_FOR')
        if xff:
            return xff.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR')


# ─────────────────────────────────────────────
# Views
# ─────────────────────────────────────────────

class AuditDashboardView(AuditRequiredMixin, TemplateView):
    """감사 증적 대시보드"""
    template_name = 'audit/dashboard.html'
    audit_section = AuditAccessLog.Section.DASHBOARD

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        today = now.date()
        week_ago = now - timedelta(days=7)

        # 오늘 접근 로그 수 (파일 기반)
        ctx['today_access_count'] = self._count_today_access(today)

        # 최근 7일 데이터 변경 건수
        ctx['week_changes'] = self._count_recent_changes(week_ago)

        # 로그인 실패 (axes)
        try:
            from axes.models import AccessAttempt
            ctx['failed_logins_today'] = AccessAttempt.objects.filter(
                attempt_time__date=today,
            ).count()
            ctx['locked_accounts'] = AccessAttempt.objects.filter(
                attempt_time__gte=now - timedelta(hours=1),
                failures_since_start__gte=settings.AXES_FAILURE_LIMIT,
            ).values('username').distinct().count()
        except Exception:
            ctx['failed_logins_today'] = 0
            ctx['locked_accounts'] = 0

        # 감사 열람 기록 최근 5건
        ctx['recent_audit_access'] = AuditAccessLog.objects.select_related(
            'user'
        ).order_by('-accessed_at')[:5]

        # 감사권한 보유자 수
        ctx['auditor_count'] = User.objects.filter(is_auditor=True, is_active=True).count()

        # ISMS 자동 진단
        ctx['isms_checks'] = self._run_isms_checks()

        # 결재 현황
        ctx.update(self._get_approval_stats())

        # 권한 변경 이력
        ctx['recent_role_changes'] = self._get_recent_role_changes()

        # Chart.js 시계열 데이터
        ctx.update(self._build_chart_data(today, week_ago))

        return ctx

    @staticmethod
    def _count_today_access(today):
        log_path = Path(settings.BASE_DIR) / 'local' / 'erp.log'
        if not log_path.exists():
            return 0
        today_str = today.strftime('%Y-%m-%d')
        count = 0
        try:
            with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
                for line in f:
                    if today_str in line and 'access' in line.lower():
                        count += 1
        except OSError:
            pass
        return count

    @staticmethod
    def _count_recent_changes(since):
        """simple_history 테이블에서 최근 변경 건수 집계"""
        from django.apps import apps
        total = 0
        for model in apps.get_models():
            if hasattr(model, 'history'):
                try:
                    total += model.history.filter(history_date__gte=since).count()
                except Exception:
                    pass
        return total

    @staticmethod
    def _run_isms_checks():
        """ISMS 증적 자동 진단"""
        checks = []

        # 1. 접근통제
        allowed_hosts_ok = bool(getattr(settings, 'ALLOWED_HOSTS', [])) and settings.ALLOWED_HOSTS != ['*']
        csp_ok = 'csp.middleware.CSPMiddleware' in getattr(settings, 'MIDDLEWARE', [])
        checks.append({
            'name': '접근통제',
            'desc': 'ALLOWED_HOSTS 설정, CSP 헤더',
            'passed': allowed_hosts_ok or csp_ok,
            'details': f'ALLOWED_HOSTS: {"설정됨" if allowed_hosts_ok else "미설정"}, CSP: {"활성" if csp_ok else "미활성"}',
        })

        # 2. 인증
        axes_ok = 'axes.middleware.AxesMiddleware' in getattr(settings, 'MIDDLEWARE', [])
        hashers = getattr(settings, 'PASSWORD_HASHERS', [])
        pbkdf2_ok = any('PBKDF2' in h for h in hashers) if hashers else True
        checks.append({
            'name': '인증',
            'desc': 'AXES 브루트포스 방지, 비밀번호 해시',
            'passed': axes_ok and pbkdf2_ok,
            'details': f'AXES: {"활성" if axes_ok else "미활성"}, PBKDF2: {"사용" if pbkdf2_ok else "미사용"}',
        })

        # 3. 개인정보
        encryption_key = bool(getattr(settings, 'FIELD_ENCRYPTION_KEY', ''))
        from django.apps import apps as django_apps
        encrypted_count = 0
        for model in django_apps.get_models():
            for field in model._meta.get_fields():
                if type(field).__name__ == 'EncryptedCharField':
                    encrypted_count += 1
                    break
        checks.append({
            'name': '개인정보 보호',
            'desc': '민감데이터 암호화, 암호화 키 설정',
            'passed': encryption_key and encrypted_count > 0,
            'details': f'암호화 키: {"설정됨" if encryption_key else "미설정"}, 암호화 모델: {encrypted_count}개',
        })

        # 4. 감사 증적
        audit_table_ok = True
        try:
            AuditAccessLog.objects.count()
        except Exception:
            audit_table_ok = False
        auditor_separated = User.objects.filter(is_auditor=True).exists()
        checks.append({
            'name': '감사 증적',
            'desc': '열람 기록 테이블, 감사권한 분리',
            'passed': audit_table_ok and auditor_separated,
            'details': f'열람기록: {"정상" if audit_table_ok else "오류"}, 감사권한분리: {"적용" if auditor_separated else "미적용"}',
        })

        # 5. 데이터 무결성
        history_count = 0
        for model in django_apps.get_models():
            if hasattr(model, 'history'):
                history_count += 1
        checks.append({
            'name': '데이터 무결성',
            'desc': 'HistoricalRecords 이력 추적',
            'passed': history_count >= 10,
            'details': f'이력추적 모델: {history_count}개',
        })

        # 6. 암호화
        secure_ssl = getattr(settings, 'SECURE_SSL_REDIRECT', False)
        fernet_ok = bool(getattr(settings, 'FIELD_ENCRYPTION_KEY', ''))
        checks.append({
            'name': '암호화',
            'desc': 'HTTPS 설정, Fernet 키',
            'passed': fernet_ok,
            'details': f'HTTPS: {"강제" if secure_ssl else "미강제(dev)"}, Fernet: {"설정됨" if fernet_ok else "미설정"}',
        })

        return checks

    @staticmethod
    def _get_approval_stats():
        """결재 현황 통계"""
        from apps.approval.models import ApprovalRequest
        qs = ApprovalRequest.objects.filter(is_active=True)
        return {
            'approval_pending': qs.filter(status__in=['DRAFT', 'SUBMITTED']).count(),
            'approval_approved': qs.filter(status='APPROVED').count(),
            'approval_rejected': qs.filter(status='REJECTED').count(),
            'recent_approvals': qs.select_related('requester').order_by('-created_at')[:10],
        }

    @staticmethod
    def _get_recent_role_changes():
        """최근 권한 변경 이력 (User.history에서 role 변경 추출)"""
        changes = []
        try:
            records = User.history.order_by('-history_date')[:200]
            for record in records:
                prev = record.prev_record
                if prev and prev.role != record.role:
                    role_display = dict(User.Role.choices)
                    changes.append({
                        'username': record.username,
                        'name': record.name,
                        'old_role': role_display.get(prev.role, prev.role),
                        'new_role': role_display.get(record.role, record.role),
                        'changed_at': record.history_date,
                        'changed_by': str(record.history_user) if record.history_user else '(시스템)',
                    })
                if len(changes) >= 10:
                    break
        except Exception:
            pass
        return changes

    @staticmethod
    def _build_chart_data(today, week_ago):
        """Chart.js 시계열 데이터"""
        from django.apps import apps as django_apps

        labels = []
        access_data = []
        change_data = []

        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            labels.append(d.strftime('%m/%d'))

            # 접근 로그 (AuditAccessLog 활용)
            access_data.append(
                AuditAccessLog.objects.filter(accessed_at__date=d).count()
            )

            # 데이터 변경 (simple_history)
            day_start = timezone.make_aware(
                timezone.datetime.combine(d, timezone.datetime.min.time())
            )
            day_end = day_start + timedelta(days=1)
            day_changes = 0
            for model in django_apps.get_models():
                if hasattr(model, 'history'):
                    try:
                        day_changes += model.history.filter(
                            history_date__gte=day_start,
                            history_date__lt=day_end,
                        ).count()
                    except Exception:
                        pass
            change_data.append(day_changes)

        # 로그인 성공/실패 (axes)
        login_success = []
        login_fail = []
        try:
            from axes.models import AccessAttempt, AccessLog
            for i in range(6, -1, -1):
                d = today - timedelta(days=i)
                login_fail.append(
                    AccessAttempt.objects.filter(attempt_time__date=d).count()
                )
                login_success.append(
                    AccessLog.objects.filter(attempt_time__date=d).count()
                )
        except Exception:
            login_success = [0] * 7
            login_fail = [0] * 7

        return {
            'chart_labels': json.dumps(labels),
            'chart_access_data': json.dumps(access_data),
            'chart_change_data': json.dumps(change_data),
            'chart_login_success': json.dumps(login_success),
            'chart_login_fail': json.dumps(login_fail),
        }


class AccessLogView(AuditRequiredMixin, TemplateView):
    """시스템 접근 로그 — erp.log 파싱"""
    template_name = 'audit/access_log.html'
    audit_section = AuditAccessLog.Section.ACCESS_LOG

    LOG_PATTERN = re.compile(
        r'\[(?P<ts>[^\]]+)\]\s+INFO\s+access\s+'
        r'(?P<user>\S+)\s+(?P<method>\S+)\s+(?P<path>\S+)\s+'
        r'(?P<status>\d+)\s+(?P<duration>\d+)ms'
    )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        log_path = Path(settings.BASE_DIR) / 'local' / 'erp.log'
        entries = []

        filter_user = self.request.GET.get('user', '')
        filter_path = self.request.GET.get('path', '')
        filter_status = self.request.GET.get('status', '')

        if log_path.exists():
            try:
                lines = log_path.read_text(
                    encoding='utf-8', errors='replace'
                ).splitlines()
                for line in reversed(lines):
                    m = self.LOG_PATTERN.match(line)
                    if not m:
                        continue
                    entry = m.groupdict()
                    if filter_user and filter_user not in entry['user']:
                        continue
                    if filter_path and filter_path not in entry['path']:
                        continue
                    if filter_status and entry['status'] != filter_status:
                        continue
                    entries.append(entry)
                    if len(entries) >= 500:
                        break
            except OSError:
                pass

        paginator = Paginator(entries, 50)
        page = self.request.GET.get('page', 1)
        ctx['page_obj'] = paginator.get_page(page)
        ctx['filter_user'] = filter_user
        ctx['filter_path'] = filter_path
        ctx['filter_status'] = filter_status
        ctx['total_entries'] = len(entries)
        return ctx


class DataChangeLogView(AuditRequiredMixin, TemplateView):
    """데이터 변경 이력 — simple_history 통합 조회"""
    template_name = 'audit/data_change_log.html'
    audit_section = AuditAccessLog.Section.DATA_CHANGE

    TRACKED_MODELS = [
        ('accounts', 'User', '사용자'),
        ('inventory', 'Product', '제품'),
        ('inventory', 'StockMovement', '입출고'),
        ('sales', 'Order', '주문'),
        ('sales', 'Quotation', '견적'),
        ('sales', 'Partner', '거래처'),
        ('purchase', 'PurchaseOrder', '발주'),
        ('production', 'ProductionPlan', '생산계획'),
        ('production', 'WorkOrder', '작업지시'),
        ('accounting', 'Voucher', '전표'),
        ('accounting', 'TaxInvoice', '세금계산서'),
        ('accounting', 'Payment', '입출금'),
        ('approval', 'ApprovalRequest', '결재'),
        ('hr', 'EmployeeProfile', '직원정보'),
        ('attendance', 'LeaveRequest', '휴가신청'),
        ('investment', 'Investment', '투자'),
    ]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.apps import apps

        filter_model = self.request.GET.get('model', '')
        filter_user = self.request.GET.get('user', '')
        filter_type = self.request.GET.get('type', '')  # +, ~, -

        entries = []
        for app_label, model_name, display in self.TRACKED_MODELS:
            if filter_model and model_name != filter_model:
                continue
            try:
                model = apps.get_model(app_label, model_name)
                if not hasattr(model, 'history'):
                    continue
                qs = model.history.select_related('history_user').order_by('-history_date')
                if filter_user:
                    qs = qs.filter(history_user__username__icontains=filter_user)
                if filter_type:
                    qs = qs.filter(history_type=filter_type)
                for record in qs[:100]:
                    entries.append({
                        'model_name': display,
                        'model_key': model_name,
                        'object_id': record.history_object.pk if hasattr(record, 'history_object') else record.pk,
                        'object_repr': str(record.instance) if hasattr(record, 'instance') else str(record.pk),
                        'history_type': record.history_type,
                        'history_type_display': {'+': '생성', '~': '수정', '-': '삭제'}.get(record.history_type, '?'),
                        'history_user': str(record.history_user) if record.history_user else '(시스템)',
                        'history_date': record.history_date,
                    })
            except Exception:
                continue

        entries.sort(key=lambda x: x['history_date'], reverse=True)
        entries = entries[:500]

        paginator = Paginator(entries, 50)
        page = self.request.GET.get('page', 1)
        ctx['page_obj'] = paginator.get_page(page)
        ctx['tracked_models'] = self.TRACKED_MODELS
        ctx['filter_model'] = filter_model
        ctx['filter_user'] = filter_user
        ctx['filter_type'] = filter_type
        ctx['total_entries'] = len(entries)
        return ctx


class LoginHistoryView(AuditRequiredMixin, TemplateView):
    """로그인/보안 이벤트 — django-axes"""
    template_name = 'audit/login_history.html'
    audit_section = AuditAccessLog.Section.LOGIN_HISTORY

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        filter_user = self.request.GET.get('user', '')
        filter_result = self.request.GET.get('result', '')  # success, fail

        entries = []
        try:
            from axes.models import AccessAttempt, AccessLog
            # 실패 기록
            if filter_result != 'success':
                qs = AccessAttempt.objects.order_by('-attempt_time')
                if filter_user:
                    qs = qs.filter(username__icontains=filter_user)
                for a in qs[:300]:
                    entries.append({
                        'username': a.username,
                        'ip_address': a.ip_address,
                        'user_agent': a.user_agent[:80] if a.user_agent else '',
                        'timestamp': a.attempt_time,
                        'result': 'FAIL',
                        'failures': a.failures_since_start,
                    })

            # 성공 기록
            if filter_result != 'fail':
                qs = AccessLog.objects.order_by('-attempt_time')
                if filter_user:
                    qs = qs.filter(username__icontains=filter_user)
                for a in qs[:300]:
                    entries.append({
                        'username': a.username,
                        'ip_address': a.ip_address,
                        'user_agent': a.user_agent[:80] if a.user_agent else '',
                        'timestamp': a.attempt_time,
                        'result': 'OK',
                        'failures': 0,
                    })
        except Exception:
            pass

        entries.sort(key=lambda x: x['timestamp'], reverse=True)
        entries = entries[:500]

        paginator = Paginator(entries, 50)
        page = self.request.GET.get('page', 1)
        ctx['page_obj'] = paginator.get_page(page)
        ctx['filter_user'] = filter_user
        ctx['filter_result'] = filter_result
        ctx['total_entries'] = len(entries)
        return ctx


class AuditAccessLogView(AuditRequiredMixin, ListView):
    """감사 열람 기록 (메타 감사) — 누가 이 시스템을 열람했는지"""
    template_name = 'audit/audit_access_log.html'
    audit_section = AuditAccessLog.Section.AUDIT_LOG
    model = AuditAccessLog
    context_object_name = 'logs'
    paginate_by = 50

    def get_queryset(self):
        qs = AuditAccessLog.objects.select_related('user').order_by('-accessed_at')
        filter_user = self.request.GET.get('user', '')
        filter_section = self.request.GET.get('section', '')
        if filter_user:
            qs = qs.filter(user__username__icontains=filter_user)
        if filter_section:
            qs = qs.filter(section=filter_section)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_user'] = self.request.GET.get('user', '')
        ctx['filter_section'] = self.request.GET.get('section', '')
        ctx['section_choices'] = AuditAccessLog.Section.choices
        return ctx


class AuditExcelExportView(AuditRequiredMixin, View):
    """증적 보고서 Excel 다운로드"""
    audit_section = AuditAccessLog.Section.DASHBOARD

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

        # Sheet 1: 감사 열람 기록
        ws = wb.active
        ws.title = '감사열람기록'
        headers = ['열람자', '섹션', 'IP', 'User-Agent', '일시']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row, log in enumerate(AuditAccessLog.objects.select_related('user').order_by('-accessed_at')[:500], 2):
            ws.cell(row=row, column=1, value=str(log.user))
            ws.cell(row=row, column=2, value=log.get_section_display())
            ws.cell(row=row, column=3, value=log.ip_address or '')
            ws.cell(row=row, column=4, value=(log.user_agent or '')[:100])
            ws.cell(row=row, column=5, value=log.accessed_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Sheet 2: 데이터 변경 이력
        ws2 = wb.create_sheet('데이터변경이력')
        headers2 = ['모델', '유형', '변경자', '일시']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        from django.apps import apps as django_apps
        type_map = {'+': '생성', '~': '수정', '-': '삭제'}
        for model in django_apps.get_models():
            if hasattr(model, 'history') and row_idx < 1000:
                try:
                    for rec in model.history.select_related('history_user').order_by('-history_date')[:50]:
                        ws2.cell(row=row_idx, column=1, value=model._meta.verbose_name)
                        ws2.cell(row=row_idx, column=2, value=type_map.get(rec.history_type, '?'))
                        ws2.cell(row=row_idx, column=3, value=str(rec.history_user) if rec.history_user else '(시스템)')
                        ws2.cell(row=row_idx, column=4, value=rec.history_date.strftime('%Y-%m-%d %H:%M:%S'))
                        row_idx += 1
                except Exception:
                    pass

        # Sheet 3: 로그인 이력
        ws3 = wb.create_sheet('로그인이력')
        headers3 = ['사용자', 'IP', '결과', '시도횟수', '일시']
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        try:
            from axes.models import AccessAttempt, AccessLog
            for a in AccessAttempt.objects.order_by('-attempt_time')[:250]:
                ws3.cell(row=row_idx, column=1, value=a.username)
                ws3.cell(row=row_idx, column=2, value=a.ip_address or '')
                ws3.cell(row=row_idx, column=3, value='실패')
                ws3.cell(row=row_idx, column=4, value=a.failures_since_start)
                ws3.cell(row=row_idx, column=5, value=a.attempt_time.strftime('%Y-%m-%d %H:%M:%S'))
                row_idx += 1
            for a in AccessLog.objects.order_by('-attempt_time')[:250]:
                ws3.cell(row=row_idx, column=1, value=a.username)
                ws3.cell(row=row_idx, column=2, value=a.ip_address or '')
                ws3.cell(row=row_idx, column=3, value='성공')
                ws3.cell(row=row_idx, column=4, value=0)
                ws3.cell(row=row_idx, column=5, value=a.attempt_time.strftime('%Y-%m-%d %H:%M:%S'))
                row_idx += 1
        except Exception:
            pass

        # Sheet 4: 권한 변경
        ws4 = wb.create_sheet('권한변경이력')
        headers4 = ['사용자', '이전역할', '변경역할', '변경자', '일시']
        for col, h in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        role_display = dict(User.Role.choices)
        row_idx = 2
        try:
            for rec in User.history.order_by('-history_date')[:500]:
                prev = rec.prev_record
                if prev and prev.role != rec.role:
                    ws4.cell(row=row_idx, column=1, value=rec.username)
                    ws4.cell(row=row_idx, column=2, value=role_display.get(prev.role, prev.role))
                    ws4.cell(row=row_idx, column=3, value=role_display.get(rec.role, rec.role))
                    ws4.cell(row=row_idx, column=4, value=str(rec.history_user) if rec.history_user else '(시스템)')
                    ws4.cell(row=row_idx, column=5, value=rec.history_date.strftime('%Y-%m-%d %H:%M:%S'))
                    row_idx += 1
        except Exception:
            pass

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        now_str = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="audit_report_{now_str}.xlsx"'
        wb.save(response)
        return response
