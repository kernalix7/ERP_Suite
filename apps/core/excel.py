import io
from datetime import date

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='맑은 고딕', size=10)
TITLE_FONT = Font(name='맑은 고딕', size=14, bold=True, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
MONEY_FORMAT = '#,##0'


def export_to_excel(title, headers, rows, filename=None, money_columns=None):
    """
    스타일링된 Excel 파일 생성.

    Args:
        title: 시트 제목
        headers: [(컬럼명, 너비), ...]
        rows: [[값, 값, ...], ...]
        filename: 다운로드 파일명
        money_columns: 금액 서식 적용할 컬럼 인덱스 리스트 (0-based)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    money_columns = money_columns or []

    # 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 30

    # 날짜
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f'출력일: {date.today().strftime("%Y-%m-%d")}')
    date_cell.font = Font(name='맑은 고딕', size=9, color='666666')
    date_cell.alignment = Alignment(horizontal='right')

    # 헤더
    for col_idx, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # 데이터
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if (col_idx - 1) in money_columns:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(vertical='center')

        # 줄무늬
        if row_idx % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color='F2F7FB', end_color='F2F7FB', fill_type='solid'
                )

    # 자동필터
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{4 + len(rows)}'

    # Disclaimer 시트
    ds = wb.create_sheet(title='개인정보 취급 안내')
    disclaimer_lines = [
        '[ 개인정보 취급 안내 (Disclaimer) ]',
        '',
        '본 파일에는 개인정보보호법 및 관련 법령에 의해 보호되는',
        '개인정보가 포함되어 있을 수 있습니다.',
        '',
        '1. 본 자료는 업무 목적으로만 사용해야 하며,',
        '   무단 복제·배포·전송을 금지합니다.',
        '',
        '2. 업무 완료 후 파일을 즉시 삭제하거나',
        '   안전하게 보관해야 합니다.',
        '',
        '3. 개인정보를 제3자에게 제공하거나 목적 외로',
        '   사용할 경우, 개인정보보호법 제71조에 따라',
        '   5년 이하의 징역 또는 5천만원 이하의',
        '   벌금에 처해질 수 있습니다.',
        '',
        '4. 본 자료에서 개인정보가 유출된 경우,',
        '   즉시 관리자에게 보고하여 주시기 바랍니다.',
        '',
        f'출력일시: {date.today().strftime("%Y-%m-%d")}',
        '출력시스템: ERP Suite',
    ]
    warn_font = Font(name='맑은 고딕', size=11)
    warn_title_font = Font(
        name='맑은 고딕', size=14, bold=True, color='CC0000',
    )
    ds.column_dimensions['A'].width = 60
    for idx, line in enumerate(disclaimer_lines, 1):
        cell = ds.cell(row=idx, column=1, value=line)
        cell.font = warn_title_font if idx == 1 else warn_font

    # 응답
    if not filename:
        filename = f'{title}_{date.today().strftime("%Y%m%d")}.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response
